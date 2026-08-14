"""KNM Work Orders v2 — Fault Report + Tech Support module.

This file is a side-by-side rewrite of workorders.py for the 2026-06-03 schema
migration. It is NOT loaded until cutover (rename to workorders.py).

Changes vs v1:
  • Status / Priority stored as TINYINT (StatusCode / PriorityCode).
  • Complaint: + FirstReportedAt, ReportedBy, EventCode, PerceivedUrgency, DisplayID.
  • JobOrder:  + EventCode, Diagnosis, ProposedFix, LastBlockReason, DisplayID.
  • Images now stored in SharePoint (Documents/ComplaintUploads/ and
    Documents/WorkOrderUploads/) via sharepoint_helper. DB stores SPItemID +
    SPWebURL only; raw bytes proxied through /api/wo/images/<id> for Easy Auth.
  • New tables: WO_JobOrderTasks (tickbox checklist), WO_KB_Entries +
    WO_KB_Tickboxes (knowledge base), WO_Counters (per-month NNNN allocator).
  • New endpoints: /joborders/<id>/tasks*, /kb*, /kb/suggest.
  • DisplayID format: KNM-CMP-NNNN-YYMM (complaints), KNM-WkO-NNNN-YYMM (WOs).

Heartbeat hook deferred per Yash's directive (2026-06-03).
"""
from __future__ import annotations

from functools import wraps
from datetime import datetime, timedelta
import base64
import io
import uuid

from flask import Blueprint, request, jsonify, Response

# Reuse helpers from the existing vending app so auth and DB stay consistent.
from app import (
    get_current_user, get_role, get_connection,
    to_ole_date, from_ole_date, log_deletion,
)
import json

# SharePoint helper — module-level import; nothing executes here.
import sharepoint_helper as sp

workorders_bp = Blueprint("workorders", __name__)


# ── Roles ─────────────────────────────────────────────────────────────────────

ROLE_ADMIN         = "admin"
ROLE_OPERATOR      = "operator"
ROLE_FIELD_MANAGER = "field_manager"

ROLE_DISPATCH      = "dispatch"
ROLE_SALES         = "sales"

# 2026-08-11 — dispatch and field_manager are the same person at KNM today, so
# the two roles are MERGED: `dispatch` is an alias of `field_manager` and gets
# exactly the same authority everywhere, including the operator surfaces.
#
# To split them again later, delete ROLE_DISPATCH from these two sets. The
# assignment routes are already tagged DISPATCH_ROLES, so they keep working for
# a narrowed dispatch role without further edits — that is the whole point of
# keeping the name around after the merge.
OPERATOR_ROLES = {ROLE_OPERATOR, ROLE_FIELD_MANAGER, ROLE_ADMIN, ROLE_DISPATCH}
MANAGER_ROLES  = {ROLE_FIELD_MANAGER, ROLE_ADMIN, ROLE_DISPATCH}
DISPATCH_ROLES = MANAGER_ROLES
# 2026-08-11 — sales keys the week's top-up/delivery schedule. Sales may create
# and cancel UNASSIGNED, still-open stops through the /schedule routes only; it
# is deliberately NOT folded into DISPATCH_ROLES, so sales can never name a
# driver, reorder a round, touch started work, or reach any other /api/wo/*.
SALES_ROLES    = MANAGER_ROLES | {ROLE_SALES}


# ── Status / Priority code maps ───────────────────────────────────────────────
# Stored as TINYINT in DB; expanded to strings in API responses.

COMPLAINT_STATUS = {0: "fresh", 1: "assigned", 2: "closed", 3: "unresolved"}
JOBORDER_STATUS  = {
    0: "assigned",
    1: "needs_assistance",
    2: "pending_review",   # operator submitted, awaiting manager review
    3: "closed",           # manager accepted (work done) or superseded
}
PRIORITY         = {0: "low",      1: "normal",           2: "high"}
MOVEMENT_STATUS  = {0: "scheduled", 1: "in_progress", 2: "completed"}
MOVEMENT_TYPES   = ("deploy", "relocate", "retrieve")


def _label(mapping: dict, code) -> str:
    if code is None:
        return ""
    try:
        return mapping.get(int(code), str(code))
    except (TypeError, ValueError):
        return str(code)


# ── 8 standard delivery items (names provided by Yash; placeholder) ───────────

DELIVERY_ITEMS = [
    "Item 1", "Item 2", "Item 3", "Item 4",
    "Item 5", "Item 6", "Item 7", "Item 8",
]


# ── Auth helpers (API-style: JSON 401/403, never redirect) ────────────────────

def api_login_required(f):
    """Any signed-in user with any app role can access."""
    @wraps(f)
    def inner(*args, **kwargs):
        email = get_current_user()
        if not email:
            return jsonify({"error": "Not signed in."}), 401
        if not get_role(email):
            return jsonify({"error": "You are not authorised to use this app."}), 403
        return f(*args, **kwargs)
    return inner


def require_roles(*roles):
    allowed = set(roles)

    def deco(f):
        @wraps(f)
        def inner(*args, **kwargs):
            email = get_current_user()
            if not email:
                return jsonify({"error": "Not signed in."}), 401
            if get_role(email) not in allowed:
                return jsonify({"error": "You do not have permission for this action."}), 403
            return f(*args, **kwargs)
        return inner
    return deco


# ── Schema init (idempotent; runs on app startup) ─────────────────────────────

def init_workorders_db():
    """Create WO_* tables if missing. Safe to run on every startup.

    NOTE: schema migrations (column additions, type changes, indexes) are
    handled by migration_2026-06-03.sql. This function only handles the
    fresh-install path so a new environment starts up cleanly.
    """
    tables = [
        ("WO_Complaints", """
            CREATE TABLE WO_Complaints (
                ComplaintID       INT IDENTITY(1,1) PRIMARY KEY,
                Description       NVARCHAR(MAX)  NOT NULL,
                Source            NVARCHAR(20)   NOT NULL DEFAULT 'self',
                ImpactDescription NVARCHAR(MAX)  NULL,
                ImpactAmount      DECIMAL(18,2)  NULL,
                ImpactSeverity    TINYINT        NULL,
                MachineName       NVARCHAR(255)  NULL,
                MachineCode       NVARCHAR(50)   NULL,
                Status            NVARCHAR(20)   NOT NULL DEFAULT 'open',
                StatusCode        TINYINT        NOT NULL DEFAULT 0,
                JobOrderID        INT            NULL,
                SubmitterEmail    NVARCHAR(255)  NOT NULL,
                SubmittedAt       DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME(),
                FirstReportedAt   DATETIME2      NULL,
                ReportedBy        NVARCHAR(255)  NULL,
                EventCode         INT            NULL,
                PerceivedUrgency  TINYINT        NOT NULL DEFAULT 1,
                DisplayID         NVARCHAR(30)   NULL,
                RefundIssued      BIT            NOT NULL DEFAULT 0,
                ClosedReason      NVARCHAR(255)  NULL,
                ClosedBy          NVARCHAR(255)  NULL,
                ClosedAt          DATETIME2      NULL,
                GroupID           INT            NULL
            )
        """),
        ("WO_JobOrders", """
            CREATE TABLE WO_JobOrders (
                JobOrderID        INT IDENTITY(1,1) PRIMARY KEY,
                ComplaintID       INT            NULL,
                MachineName       NVARCHAR(255)  NOT NULL,
                MachineCode       NVARCHAR(50)   NULL,
                Notes             NVARCHAR(MAX)  NULL,
                AssignedTo        NVARCHAR(255)  NULL,
                Priority          NVARCHAR(10)   NOT NULL DEFAULT 'normal',
                PriorityCode      TINYINT        NOT NULL DEFAULT 1,
                Status            NVARCHAR(20)   NOT NULL DEFAULT 'open',
                StatusCode        TINYINT        NOT NULL DEFAULT 0,
                Report            NVARCHAR(MAX)  NULL,
                RootCause         NVARCHAR(MAX)  NULL,
                CorrectiveAction  NVARCHAR(MAX)  NULL,
                PreventiveAction  NVARCHAR(MAX)  NULL,
                EventCode         INT            NULL,
                Diagnosis         NVARCHAR(MAX)  NULL,
                ProposedFix       NVARCHAR(MAX)  NULL,
                LastBlockReason   NVARCHAR(MAX)  NULL,
                DisplayID         NVARCHAR(30)   NULL,
                AttachedKBID      INT            NULL,
                OnSiteObservations NVARCHAR(MAX) NULL,
                OnSiteChanges      NVARCHAR(MAX) NULL,
                TechnicianComments NVARCHAR(MAX) NULL,
                CreatedBy         NVARCHAR(255)  NOT NULL,
                CreatedAt         DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME(),
                CompletedBy       NVARCHAR(255)  NULL,
                CompletedAt       DATETIME2      NULL
            )
        """),
        ("WO_DeliveryOrders", """
            CREATE TABLE WO_DeliveryOrders (
                DeliveryOrderID   INT IDENTITY(1,1) PRIMARY KEY,
                MachineName       NVARCHAR(255)  NOT NULL,
                MachineCode       NVARCHAR(50)   NULL,
                Notes             NVARCHAR(MAX)  NULL,
                AssignedTo        NVARCHAR(255)  NULL,
                Priority          NVARCHAR(10)   NOT NULL DEFAULT 'normal',
                Status            NVARCHAR(20)   NOT NULL DEFAULT 'open',
                Item1Qty INT NOT NULL DEFAULT 0, Item2Qty INT NOT NULL DEFAULT 0,
                Item3Qty INT NOT NULL DEFAULT 0, Item4Qty INT NOT NULL DEFAULT 0,
                Item5Qty INT NOT NULL DEFAULT 0, Item6Qty INT NOT NULL DEFAULT 0,
                Item7Qty INT NOT NULL DEFAULT 0, Item8Qty INT NOT NULL DEFAULT 0,
                RecipientName     NVARCHAR(255)  NULL,
                CreatedBy         NVARCHAR(255)  NOT NULL,
                CreatedAt         DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME(),
                CompletedBy       NVARCHAR(255)  NULL,
                CompletedAt       DATETIME2      NULL
            )
        """),
        ("WO_Images", """
            CREATE TABLE WO_Images (
                ImageID      INT IDENTITY(1,1) PRIMARY KEY,
                ParentType   NVARCHAR(20)   NOT NULL,
                ParentID     INT            NOT NULL,
                Stage        NVARCHAR(20)   NOT NULL,
                ImageData    VARBINARY(MAX) NULL,
                SPItemID     NVARCHAR(255)  NULL,
                SPWebURL     NVARCHAR(1024) NULL,
                ContentType  NVARCHAR(100)  NOT NULL,
                FileName     NVARCHAR(255)  NULL,
                UploadedBy   NVARCHAR(255)  NOT NULL,
                UploadedAt   DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME()
            )
        """),
        ("WO_Activity", """
            CREATE TABLE WO_Activity (
                ActivityID   INT IDENTITY(1,1) PRIMARY KEY,
                ParentType   NVARCHAR(20)   NOT NULL,
                ParentID     INT            NOT NULL,
                Action       NVARCHAR(50)   NOT NULL,
                Detail       NVARCHAR(MAX)  NULL,
                ByUser       NVARCHAR(255)  NOT NULL,
                AtTime       DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME()
            )
        """),
        ("WO_JobOrderTasks", """
            CREATE TABLE WO_JobOrderTasks (
                TaskID        INT IDENTITY(1,1) PRIMARY KEY,
                JobOrderID    INT             NOT NULL,
                SeqNum        INT             NOT NULL,
                Label         NVARCHAR(500)   NOT NULL,
                Done          BIT             NOT NULL DEFAULT 0,
                BlockedNote   NVARCHAR(MAX)   NULL,
                CompletedBy   NVARCHAR(255)   NULL,
                CompletedAt   DATETIME2       NULL
            )
        """),
        ("WO_KB_Entries", """
            CREATE TABLE WO_KB_Entries (
                KBID          INT IDENTITY(1,1) PRIMARY KEY,
                EventCode     INT             NULL,
                Title         NVARCHAR(255)   NOT NULL,
                Diagnosis     NVARCHAR(MAX)   NULL,
                SuggestedFix  NVARCHAR(MAX)   NULL,
                Symptom                   NVARCHAR(MAX) NULL,
                DiagnosticConfirmation    NVARCHAR(MAX) NULL,
                RootCause                 NVARCHAR(MAX) NULL,
                CorrectiveAction          NVARCHAR(MAX) NULL,
                PreventiveAction          NVARCHAR(MAX) NULL,
                VerificationOfCompletion  NVARCHAR(MAX) NULL,
                UseCount      INT             NOT NULL DEFAULT 0,
                CreatedBy     NVARCHAR(255)   NOT NULL,
                CreatedAt     DATETIME2       NOT NULL DEFAULT SYSUTCDATETIME(),
                UpdatedBy     NVARCHAR(255)   NULL,
                UpdatedAt     DATETIME2       NULL
            )
        """),
        ("WO_KB_Tickboxes", """
            CREATE TABLE WO_KB_Tickboxes (
                TBID    INT IDENTITY(1,1) PRIMARY KEY,
                KBID    INT             NOT NULL,
                SeqNum  INT             NOT NULL,
                Label   NVARCHAR(500)   NOT NULL
            )
        """),
        ("WO_Counters", """
            CREATE TABLE WO_Counters (
                Kind      NVARCHAR(10)  NOT NULL,
                YYMM      CHAR(4)       NOT NULL,
                NextSeq   INT           NOT NULL,
                CONSTRAINT PK_WO_Counters PRIMARY KEY (Kind, YYMM)
            )
        """),
        ("WO_MovementOrders", """
            CREATE TABLE WO_MovementOrders (
                MovementOrderID    INT IDENTITY(1,1) PRIMARY KEY,
                MovementType       NVARCHAR(20)   NOT NULL,
                MachineCode        NVARCHAR(50)   NOT NULL,
                FromLocation       NVARCHAR(255)  NULL,
                FromLat            FLOAT          NULL,
                FromLon            FLOAT          NULL,
                ToLocation         NVARCHAR(255)  NULL,
                ToLat              FLOAT          NULL,
                ToLon              FLOAT          NULL,
                Notes              NVARCHAR(MAX)  NULL,
                AssignedTo         NVARCHAR(255)  NULL,
                StatusCode         TINYINT        NOT NULL DEFAULT 0,
                DisplayID          NVARCHAR(30)   NULL,
                ReasonForRetrieval NVARCHAR(255)  NULL,
                CreatedBy          NVARCHAR(255)  NOT NULL,
                CreatedAt          DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME(),
                CompletedBy        NVARCHAR(255)  NULL,
                CompletedAt        DATETIME2      NULL
            )
        """),
        ("WO_DeletedLog", """
            CREATE TABLE WO_DeletedLog (
                LogID       INT IDENTITY(1,1) PRIMARY KEY,
                EntityType  NVARCHAR(40)   NOT NULL,
                EntityKey   NVARCHAR(100)  NULL,
                Summary     NVARCHAR(500)  NULL,
                Snapshot    NVARCHAR(MAX)  NULL,
                Reversible  BIT            NOT NULL DEFAULT 0,
                DeletedBy   NVARCHAR(255)  NOT NULL,
                DeletedAt   DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME(),
                RestoredBy  NVARCHAR(255)  NULL,
                RestoredAt  DATETIME2      NULL
            )
        """),
    ]
    indexes = [
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOImg_Parent') "
        "CREATE INDEX IX_WOImg_Parent ON WO_Images (ParentType, ParentID)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOAct_Parent') "
        "CREATE INDEX IX_WOAct_Parent ON WO_Activity (ParentType, ParentID)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOJO_Assigned') "
        "CREATE INDEX IX_WOJO_Assigned ON WO_JobOrders (AssignedTo, StatusCode)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WODO_Assigned') "
        "CREATE INDEX IX_WODO_Assigned ON WO_DeliveryOrders (AssignedTo, Status)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOJOTask_JobOrder') "
        "CREATE INDEX IX_WOJOTask_JobOrder ON WO_JobOrderTasks (JobOrderID, SeqNum)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOKB_EventCode') "
        "CREATE INDEX IX_WOKB_EventCode ON WO_KB_Entries (EventCode, UseCount DESC)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOKBTB_KB') "
        "CREATE INDEX IX_WOKBTB_KB ON WO_KB_Tickboxes (KBID, SeqNum)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOMVO_Machine') "
        "CREATE INDEX IX_WOMVO_Machine ON WO_MovementOrders (MachineCode, CreatedAt DESC)",
        "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_WOMVO_Assigned') "
        "CREATE INDEX IX_WOMVO_Assigned ON WO_MovementOrders (AssignedTo, StatusCode)",
    ]

    # Idempotent column adds, same shape as the table and index loops below.
    # ScheduledDate turns "every open order" into "today's round"; RouteSeq is
    # the driver's stop order within that round. Both nullable, so every
    # existing row keeps working exactly as it does now.
    columns = [
        ("WO_DeliveryOrders", "ScheduledDate", "DATE"),
        ("WO_DeliveryOrders", "RouteSeq",      "INT"),
        # 2026-08-11 — a service request is now a flag and a note ON the delivery
        # order rather than a second WO_JobOrders row. ServiceNote is separate
        # from Notes so folding an existing job order in cannot clobber a
        # top-up note that is already there.
        ("WO_DeliveryOrders", "NeedsService",  "BIT"),
        ("WO_DeliveryOrders", "ServiceNote",   "NVARCHAR(MAX)"),
        # 2026-08-11 — sales schedule. A repeat is MATERIALISED: one real row
        # per occurrence, all sharing SeriesID, so every existing read path
        # (driver list, board, PDF, top-up sync) sees ordinary stops and needs
        # no knowledge of recurrence. SeriesRule is kept only so the UI can say
        # "weekly" and offer "cancel the rest of the series".
        ("WO_DeliveryOrders", "SeriesID",      "NVARCHAR(40)"),
        ("WO_DeliveryOrders", "SeriesRule",    "NVARCHAR(24)"),
        ("WO_DeliveryOrders", "RequestedBy",   "NVARCHAR(256)"),
        ("WO_JobOrders",      "ScheduledDate", "DATE"),
        ("WO_JobOrders",      "RouteSeq",      "INT"),
        ("WO_MovementOrders", "ScheduledDate", "DATE"),
        ("WO_MovementOrders", "RouteSeq",      "INT"),
    ]

    conn = get_connection()
    cursor = conn.cursor()
    for name, create_sql in tables:
        try:
            cursor.execute(f"""
                IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{name}')
                {create_sql}
            """)
            conn.commit()
        except Exception as e:
            print(f"[init_workorders_db] create {name} failed: {e}")
    for tbl, col, dtype in columns:
        try:
            cursor.execute(f"""
                IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                               WHERE TABLE_NAME = '{tbl}' AND COLUMN_NAME = '{col}')
                ALTER TABLE {tbl} ADD [{col}] {dtype} NULL
            """)
            conn.commit()
        except Exception as e:
            print(f"[init_workorders_db] column {tbl}.{col} skipped: {e}")
    for idx_sql in indexes:
        try:
            cursor.execute(idx_sql)
            conn.commit()
        except Exception as e:
            print(f"[init_workorders_db] index skipped: {e}")
    conn.close()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _iso(dt):
    if not dt:
        return None
    return dt.replace(microsecond=0).isoformat() + "Z"


def _log_activity(cursor, parent_type, parent_id, action, detail, user):
    cursor.execute(
        "INSERT INTO WO_Activity (ParentType, ParentID, Action, Detail, ByUser) "
        "VALUES (%s, %s, %s, %s, %s)",
        (parent_type, parent_id, action, detail, user),
    )


def _images_for(cursor, parent_type, parent_id, stage=None):
    if stage:
        cursor.execute(
            "SELECT ImageID, Stage, ContentType, FileName, UploadedBy, UploadedAt, SPWebURL "
            "FROM WO_Images WHERE ParentType=%s AND ParentID=%s AND Stage=%s "
            "ORDER BY UploadedAt, ImageID",
            (parent_type, parent_id, stage),
        )
    else:
        cursor.execute(
            "SELECT ImageID, Stage, ContentType, FileName, UploadedBy, UploadedAt, SPWebURL "
            "FROM WO_Images WHERE ParentType=%s AND ParentID=%s "
            "ORDER BY UploadedAt, ImageID",
            (parent_type, parent_id),
        )
    return [{
        "id": r[0], "stage": r[1], "content_type": r[2],
        "file_name": r[3], "uploaded_by": r[4], "uploaded_at": _iso(r[5]),
        "sp_web_url": r[6],
    } for r in cursor.fetchall()]


def _activity_for(cursor, parent_type, parent_id):
    cursor.execute(
        "SELECT Action, Detail, ByUser, AtTime FROM WO_Activity "
        "WHERE ParentType=%s AND ParentID=%s ORDER BY AtTime DESC, ActivityID DESC",
        (parent_type, parent_id),
    )
    return [{
        "action": r[0], "detail": r[1], "by": r[2], "at": _iso(r[3]),
    } for r in cursor.fetchall()]


def _decode_data_url(data_url):
    """('data:image/jpeg;base64,XXXX', ...) -> (bytes, content_type)."""
    if not isinstance(data_url, str) or not data_url.startswith("data:"):
        return None, None
    try:
        header, b64 = data_url.split(",", 1)
        content_type = header.split(";")[0].replace("data:", "").strip() or "image/jpeg"
        if not content_type.startswith("image/"):
            return None, None
        return base64.b64decode(b64), content_type
    except Exception:
        return None, None


# ── DisplayID generator (atomic per-month NNNN allocator) ─────────────────────

def allocate_display_id(cursor, kind: str) -> str:
    """
    kind: 'CMP' (complaints), 'WkO' (work orders), 'MVO' (movement orders)
    Returns: 'KNM-{Kind}-NNNN-YYMM'.

    Uses WO_Counters with row-level locking to atomically allocate a new
    sequence number for the current YYMM. Caller must commit.
    """
    if kind not in ("CMP", "WkO", "MVO", "VIS"):
        raise ValueError(f"Unknown DisplayID kind: {kind!r}")
    now = datetime.utcnow()
    yymm = now.strftime("%y%m")

    # MERGE upsert with OUTPUT — atomic next-seq under WO_Counters PK lock.
    cursor.execute("""
        MERGE WO_Counters WITH (HOLDLOCK) AS tgt
        USING (SELECT %s AS Kind, %s AS YYMM) AS src
           ON tgt.Kind = src.Kind AND tgt.YYMM = src.YYMM
        WHEN MATCHED THEN
            UPDATE SET NextSeq = tgt.NextSeq + 1
        WHEN NOT MATCHED THEN
            INSERT (Kind, YYMM, NextSeq) VALUES (src.Kind, src.YYMM, 2)
        OUTPUT
            CASE WHEN $action = 'INSERT' THEN 1 ELSE inserted.NextSeq - 1 END AS Seq;
    """, (kind, yymm))
    row = cursor.fetchone()
    if not row:
        raise RuntimeError("Counter allocation returned no row.")
    seq = int(row[0])
    return f"KNM-{kind}-{seq:04d}-{yymm}"


# ── Image upload helpers (SP-backed) ──────────────────────────────────────────

def _kind_for_parent(parent_type: str) -> str:
    """Map parent_type → SP folder kind."""
    if parent_type == "complaint":
        return "complaint"
    # joborder, task → workorder folder
    return "workorder"


def _display_for_parent(cursor, parent_type: str, parent_id: int) -> str:
    """Fetch the DisplayID for the parent so SP files land in the right folder."""
    if parent_type == "complaint":
        cursor.execute("SELECT DisplayID FROM WO_Complaints WHERE ComplaintID=%s", (parent_id,))
    elif parent_type == "joborder":
        cursor.execute("SELECT DisplayID FROM WO_JobOrders WHERE JobOrderID=%s", (parent_id,))
    elif parent_type == "task":
        cursor.execute("""
            SELECT j.DisplayID FROM WO_JobOrderTasks t
            INNER JOIN WO_JobOrders j ON j.JobOrderID = t.JobOrderID
            WHERE t.TaskID = %s
        """, (parent_id,))
    else:
        return "UNKNOWN"
    row = cursor.fetchone()
    return (row[0] if row and row[0] else f"{parent_type}-{parent_id}")


def _save_image_to_sp(cursor, parent_type, parent_id, stage, file_name,
                      content_type, raw_bytes, uploaded_by) -> int:
    """Upload to SP, insert WO_Images row, return ImageID. Caller commits."""
    display_id = _display_for_parent(cursor, parent_type, parent_id)
    now = datetime.utcnow()
    sp_item_id, web_url, _path = sp.upload_bytes(
        kind        = _kind_for_parent(parent_type),
        display_id  = display_id,
        year        = now.year,
        month       = now.month,
        file_name   = file_name or f"{stage}.bin",
        data        = raw_bytes,
        content_type = content_type or "application/octet-stream",
    )
    cursor.execute("""
        INSERT INTO WO_Images
            (ParentType, ParentID, Stage, ImageData, SPItemID, SPWebURL,
             ContentType, FileName, UploadedBy)
        OUTPUT INSERTED.ImageID
        VALUES (%s, %s, %s, NULL, %s, %s, %s, %s, %s)
    """, (parent_type, parent_id, stage, sp_item_id, web_url,
          content_type, file_name, uploaded_by))
    return int(cursor.fetchone()[0])


def _ingest_data_urls(cursor, parent_type, parent_id, stage, data_urls, user):
    """Accept a list of data-URL strings; upload each to SP. Returns count."""
    if not data_urls:
        return 0
    count = 0
    for du in data_urls:
        raw, ctype = _decode_data_url(du)
        if not raw:
            continue
        _save_image_to_sp(
            cursor, parent_type, parent_id, stage,
            file_name=f"{stage}-{count+1}.jpg",
            content_type=ctype or "image/jpeg",
            raw_bytes=raw, uploaded_by=user,
        )
        count += 1
    return count


# ── Technicians (AAD-backed via Graph) ────────────────────────────────────────

@workorders_bp.route("/technicians")
@require_roles(*DISPATCH_ROLES)
def api_technicians():
    """Returns users with the 'operator' app role (rendered as 'Technician' in UI).
    Pulled live from Microsoft Graph via sharepoint_helper.list_users_by_role
    using the dashboard's Easy Auth service principal + the Operator role's GUID.

    Required env vars on App Service:
        MS_EASYAUTH_SP_OBJECT_ID  — Object ID of the dashboard's enterprise app
        MS_OPERATOR_ROLE_ID       — GUID of the 'operator' app role
    """
    import os
    import config as _cfg
    sp_id   = os.environ.get("MS_EASYAUTH_SP_OBJECT_ID") or getattr(_cfg, "MS_EASYAUTH_SP_OBJECT_ID", "")
    role_id = os.environ.get("MS_OPERATOR_ROLE_ID")       or getattr(_cfg, "MS_OPERATOR_ROLE_ID", "")
    if not sp_id or not role_id:
        return jsonify({
            "error": "Technician lookup not configured. Set MS_EASYAUTH_SP_OBJECT_ID and MS_OPERATOR_ROLE_ID env vars.",
            "technicians": [],
        }), 200  # 200 so dropdown falls back to free text
    try:
        users = sp.list_users_by_role(sp_id.strip(), role_id.strip())
        return jsonify({"technicians": users})
    except Exception as e:
        return jsonify({
            "error": f"Graph lookup failed: {str(e)}",
            "technicians": [],
        }), 200


@workorders_bp.route("/technicians/refresh", methods=["POST"])
@require_roles(*DISPATCH_ROLES)
def api_technicians_refresh():
    """Bust the 10-min cache to force a fresh Graph fetch."""
    try:
        sp.clear_user_cache()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@workorders_bp.route("/technicians/diag")
@require_roles(ROLE_ADMIN)
def api_technicians_diag():
    """Admin diagnostic: shows raw Graph response so we can see why the dropdown
    is empty. Returns the env-var presence, the appRoleAssignedTo response, and
    the role-filter logic."""
    import os
    import requests as _req
    import config as _cfg
    sp_id   = (os.environ.get("MS_EASYAUTH_SP_OBJECT_ID") or getattr(_cfg, "MS_EASYAUTH_SP_OBJECT_ID", "") or "").strip()
    role_id = (os.environ.get("MS_OPERATOR_ROLE_ID")     or getattr(_cfg, "MS_OPERATOR_ROLE_ID", "")     or "").strip()
    out = {
        "env": {
            "MS_EASYAUTH_SP_OBJECT_ID_present": bool(sp_id),
            "MS_EASYAUTH_SP_OBJECT_ID_value":   sp_id,
            "MS_OPERATOR_ROLE_ID_present":      bool(role_id),
            "MS_OPERATOR_ROLE_ID_value":        role_id,
        },
    }
    if not sp_id or not role_id:
        out["status"] = "env_vars_missing"
        return jsonify(out)
    try:
        token_present = bool(sp._acquire_token())
        out["token_acquired"] = token_present
        url = f"{sp.GRAPH_BASE}/servicePrincipals/{sp_id}/appRoleAssignedTo"
        r = _req.get(url, headers=sp._auth_header(), timeout=20)
        out["graph_status_code"] = r.status_code
        try:
            j = r.json()
        except Exception:
            j = {"raw": r.text[:400]}
        if r.status_code != 200:
            out["status"] = "graph_error"
            out["graph_response"] = j
            return jsonify(out)
        assignments = j.get("value", [])
        out["total_assignments"] = len(assignments)
        out["distinct_app_role_ids_seen"] = sorted({a.get("appRoleId") for a in assignments})
        target_matches = [a for a in assignments if a.get("appRoleId") == role_id]
        out["matches_for_target_role"] = len(target_matches)
        out["sample_assignment"] = assignments[0] if assignments else None
        # Try to look up the FIRST user-type match and show what /users/{id} returns
        user_match = next((a for a in target_matches if (a.get("principalType") or "") == "User"), None)
        if user_match:
            pid = user_match.get("principalId")
            try:
                ur = _req.get(
                    f"{sp.GRAPH_BASE}/users/{pid}?$select=id,mail,userPrincipalName,displayName,accountEnabled",
                    headers=sp._auth_header(), timeout=15,
                )
                out["user_lookup_status"]   = ur.status_code
                try:
                    out["user_lookup_body"] = ur.json()
                except Exception:
                    out["user_lookup_body"] = ur.text[:400]
            except Exception as e2:
                out["user_lookup_exception"] = str(e2)
        out["status"] = "ok"
        return jsonify(out)
    except Exception as e:
        out["status"] = "exception"
        out["error"] = str(e)
        return jsonify(out)


# ── Bootstrap ─────────────────────────────────────────────────────────────────

@workorders_bp.route("/bootstrap")
@api_login_required
def api_bootstrap():
    """Tells the UI who the user is, what role, what tabs to show, delivery
    item labels, and status/priority code maps."""
    email = get_current_user()
    role  = get_role(email)
    tabs = ["fault_report"]
    if role in MANAGER_ROLES:
        tabs += ["tech_support", "kb_admin", "field_ops", "delivery", "manager"]
    elif role in OPERATOR_ROLES:
        tabs += ["tech_support", "field_ops", "delivery"]
    return jsonify({
        "email": email,
        "role":  role,
        "tabs":  tabs,
        "delivery_items": DELIVERY_ITEMS,
        "complaint_status": COMPLAINT_STATUS,
        "joborder_status":  JOBORDER_STATUS,
        "priority":         PRIORITY,
    })


# ── Machines (for dropdowns in submission forms) ──────────────────────────────

@workorders_bp.route("/machines")
@api_login_required
def api_machines():
    """Active machines only. Pass ?include_inactive=1 to include decommissioned."""
    include_inactive = request.args.get("include_inactive") in ("1", "true", "yes")
    try:
        conn = get_connection()
        cursor = conn.cursor()
        sql = "SELECT MachineName, MachineCode, ISNULL(IsActive, 1) FROM MachineLookup"
        if not include_inactive:
            sql += " WHERE ISNULL(IsActive, 1) = 1"
        sql += " ORDER BY MachineName"
        cursor.execute(sql)
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "name": r[0], "code": str(r[1]), "is_active": bool(r[2]),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/locations/historic")
@require_roles(*MANAGER_ROLES)
def api_locations_historic():
    """Decommissioned machines with lifetime vend counts."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                ml.MachineName, ml.MachineCode, ml.Latitude, ml.Longitude,
                ml.DecommissionedAt, ml.DecommissionReason,
                (
                    SELECT COUNT(DISTINCT CAST(mdt.[Date Time] AS FLOAT)) FROM [MasterData Table] mdt
                    WHERE CAST(mdt.[Machine Code] AS NVARCHAR(50)) = CAST(ml.MachineCode AS NVARCHAR(50))
                      AND LEN(CAST(mdt.[Event Code] AS NVARCHAR(20))) = 6
                      AND CAST(mdt.[Event Code] AS NVARCHAR(20)) LIKE '1%'
                ) AS LifetimeVends
            FROM MachineLookup ml
            WHERE ISNULL(ml.IsActive, 1) = 0
            ORDER BY ml.DecommissionedAt DESC
        """)
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "name": r[0], "code": str(r[1]),
            "lat": r[2], "lon": r[3],
            "decommissioned_at": _iso(r[4]),
            "decommission_reason": r[5],
            "lifetime_vends": int(r[6] or 0),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/locations/<path:code>/decommission", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_location_decommission(code):
    data = request.get_json(silent=True) or {}
    reason = (data.get("reason") or "").strip() or None
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE MachineLookup
            SET IsActive = 0,
                DecommissionedAt = SYSUTCDATETIME(),
                DecommissionReason = %s
            WHERE MachineCode = %s
        """, (reason, code))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/locations/<path:code>/recommission", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_location_recommission(code):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE MachineLookup
            SET IsActive = 1, DecommissionedAt = NULL, DecommissionReason = NULL
            WHERE MachineCode = %s
        """, (code,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Complaints  →  "Fault Report" tab ─────────────────────────────────────────

@workorders_bp.route("/complaints", methods=["POST"])
@api_login_required
def api_complaint_create():
    """Customer-service rep submits a fault report.
    Body (JSON):
        description        (required)
        machine_code       (recommended)
        machine_name       (optional — looked up if missing)
        first_reported_at  (ISO8601 — when customer originally reported)
        reported_by        (free text — who called in)
        event_code         (optional int 600000–699999)
        perceived_urgency  (0=low, 1=normal, 2=high; default 1)
        impact_description (optional)
        impact_severity    (optional int 1-5; 1=minimal, 5=critical)
        source             ('self' | 'customer_chat'; default 'self')
        images             (optional list of data: URLs)
    """
    data = request.get_json(silent=True) or {}
    desc = (data.get("description") or "").strip()
    if not desc:
        return jsonify({"error": "Please describe the complaint."}), 400

    source = (data.get("source") or "self").strip().lower()
    if source not in ("self", "customer_chat"):
        source = "self"

    impact_desc = (data.get("impact_description") or "").strip() or None
    impact_severity = data.get("impact_severity")
    if impact_severity in (None, "", "null"):
        impact_severity = None
    else:
        try:
            impact_severity = int(impact_severity)
            if not (1 <= impact_severity <= 5):
                return jsonify({"error": "Impact severity must be 1-5."}), 400
        except (TypeError, ValueError):
            return jsonify({"error": "Impact severity must be an integer 1-5."}), 400

    # Refund issued flag (CS rep can tick at submission; does NOT close).
    refund_issued = bool(data.get("refund_issued"))

    machine_name = (data.get("machine_name") or "").strip() or None
    machine_code = (data.get("machine_code") or "").strip() or None
    reported_by  = (data.get("reported_by") or "").strip() or None

    first_reported_at = None
    fr_raw = data.get("first_reported_at")
    if fr_raw:
        try:
            first_reported_at = datetime.fromisoformat(str(fr_raw).replace("Z", ""))
        except ValueError:
            return jsonify({"error": "first_reported_at must be ISO8601."}), 400

    event_code = data.get("event_code")
    if event_code not in (None, "", "null"):
        try:
            event_code = int(event_code)
            if not (600000 <= event_code <= 699999):
                return jsonify({"error": "Complaint EventCode must be 600000–699999."}), 400
        except (TypeError, ValueError):
            return jsonify({"error": "EventCode must be an integer."}), 400
    else:
        event_code = None

    urgency = data.get("perceived_urgency", 1)
    try:
        urgency = int(urgency)
        if urgency not in (0, 1, 2):
            urgency = 1
    except (TypeError, ValueError):
        urgency = 1

    images = data.get("images") or []

    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Resolve MachineName if only code given
        if machine_code and not machine_name:
            cursor.execute(
                "SELECT MachineName FROM MachineLookup WHERE MachineCode=%s",
                (machine_code,),
            )
            row = cursor.fetchone()
            if row:
                machine_name = row[0]

        display_id = allocate_display_id(cursor, "CMP")

        cursor.execute("""
            INSERT INTO WO_Complaints
                (Description, Source, ImpactDescription, ImpactSeverity,
                 MachineName, MachineCode, SubmitterEmail,
                 FirstReportedAt, ReportedBy, EventCode, PerceivedUrgency,
                 StatusCode, DisplayID, RefundIssued)
            OUTPUT INSERTED.ComplaintID
            VALUES (%s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    0, %s, %s)
        """, (desc, source, impact_desc, impact_severity,
              machine_name, machine_code, user,
              first_reported_at, reported_by, event_code, urgency,
              display_id, 1 if refund_issued else 0))
        new_id = int(cursor.fetchone()[0])

        _log_activity(cursor, "complaint", new_id, "submitted",
                      f"Submitted as {display_id}", user)

        # Upload images (if any) to SP
        img_count = _ingest_data_urls(cursor, "complaint", new_id, "before", images, user)

        conn.commit()
        conn.close()
        return jsonify({
            "ok": True,
            "id": new_id,
            "display_id": display_id,
            "images_uploaded": img_count,
        })
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/complaints")
@api_login_required
def api_complaint_list():
    """?scope=mine|all  ?status=fresh|assigned|closed|all  ?machine_code=…"""
    scope        = (request.args.get("scope") or "mine").strip().lower()
    status_filter = (request.args.get("status") or "all").strip().lower()
    machine_code = (request.args.get("machine_code") or "").strip() or None
    user         = get_current_user()

    # Managers can scope=all; ordinary submitters default to their own.
    if scope == "all" and get_role(user) not in MANAGER_ROLES:
        scope = "mine"

    where, params = [], []
    if scope == "mine":
        where.append("SubmitterEmail = %s")
        params.append(user)

    rev = {v: k for k, v in COMPLAINT_STATUS.items()}
    if status_filter in rev:
        where.append("StatusCode = %s")
        params.append(rev[status_filter])

    if machine_code:
        where.append("MachineCode = %s")
        params.append(machine_code)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sql = f"""
        SELECT TOP 200
            ComplaintID, DisplayID, Description, Source,
            ImpactDescription, ImpactSeverity,
            MachineName, MachineCode,
            StatusCode, JobOrderID,
            SubmitterEmail, SubmittedAt,
            FirstReportedAt, ReportedBy, EventCode, PerceivedUrgency,
            RefundIssued, GroupID, ClosedReason, ClosedBy, ClosedAt,
            (SELECT COUNT(*) FROM WO_Images
                WHERE ParentType='complaint' AND ParentID = wc.ComplaintID) AS ImgCount
        FROM WO_Complaints wc
        {where_sql}
        ORDER BY SubmittedAt DESC, ComplaintID DESC
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "id": r[0], "display_id": r[1], "description": r[2], "source": r[3],
            "impact_description": r[4],
            "impact_severity": int(r[5]) if r[5] is not None else None,
            "machine_name": r[6], "machine_code": r[7],
            "status_code":  int(r[8]),
            "status_label": _label(COMPLAINT_STATUS, r[8]),
            "job_order_id": r[9],
            "submitter": r[10], "submitted_at": _iso(r[11]),
            "first_reported_at": _iso(r[12]),
            "reported_by": r[13],
            "event_code": int(r[14]) if r[14] is not None else None,
            "perceived_urgency_code":  int(r[15]),
            "perceived_urgency_label": _label(PRIORITY, r[15]),
            "refund_issued": bool(r[16]),
            "group_id":      int(r[17]) if r[17] is not None else None,
            "closed_reason": r[18],
            "closed_by":     r[19],
            "closed_at":     _iso(r[20]),
            "image_count":   int(r[21]),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/complaints/<int:cid>")
@api_login_required
def api_complaint_detail(cid):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ComplaintID, DisplayID, Description, Source,
                   ImpactDescription, ImpactSeverity,
                   MachineName, MachineCode,
                   StatusCode, JobOrderID,
                   SubmitterEmail, SubmittedAt,
                   FirstReportedAt, ReportedBy, EventCode, PerceivedUrgency,
                   RefundIssued, GroupID, ClosedReason, ClosedBy, ClosedAt
            FROM WO_Complaints WHERE ComplaintID = %s
        """, (cid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Complaint not found."}), 404
        images   = _images_for(cursor, "complaint", cid)
        activity = _activity_for(cursor, "complaint", cid)
        conn.close()
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500
    return jsonify({
        "id": row[0], "display_id": row[1], "description": row[2], "source": row[3],
        "impact_description": row[4],
        "impact_severity": int(row[5]) if row[5] is not None else None,
        "machine_name": row[6], "machine_code": row[7],
        "status_code":  int(row[8]),
        "status_label": _label(COMPLAINT_STATUS, row[8]),
        "job_order_id": row[9],
        "submitter": row[10], "submitted_at": _iso(row[11]),
        "first_reported_at": _iso(row[12]),
        "reported_by": row[13],
        "event_code": int(row[14]) if row[14] is not None else None,
        "perceived_urgency_code":  int(row[15]),
        "perceived_urgency_label": _label(PRIORITY, row[15]),
        "refund_issued": bool(row[16]),
        "group_id":      int(row[17]) if row[17] is not None else None,
        "closed_reason": row[18],
        "closed_by":     row[19],
        "closed_at":     _iso(row[20]),
        "images": images, "activity": activity,
    })


# ── Complaint actions: refund flag, manager close, link group, suggestions ────

@workorders_bp.route("/complaints/<int:cid>/refund", methods=["POST"])
@api_login_required
def api_complaint_refund(cid):
    """CS rep toggles RefundIssued. Does NOT close the complaint.
    Body: {refund_issued: bool}"""
    data = request.get_json(silent=True) or {}
    flag = bool(data.get("refund_issued"))
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE WO_Complaints SET RefundIssued = %s WHERE ComplaintID = %s",
            (1 if flag else 0, cid),
        )
        _log_activity(cursor, "complaint", cid, "refund_flag",
                      f"refund_issued = {flag}", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "refund_issued": flag})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/complaints/close", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_complaint_close():
    """Manager-only. Bulk-close fresh complaints (no WO needed).
    Body: {complaint_ids: [int, ...], reason: str}"""
    data = request.get_json(silent=True) or {}
    ids = data.get("complaint_ids") or []
    reason = (data.get("reason") or "").strip() or "Closed by manager"
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "complaint_ids list required."}), 400
    try:
        ids = [int(i) for i in ids]
    except (TypeError, ValueError):
        return jsonify({"error": "complaint_ids must be integers."}), 400
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        ph = ",".join(["%s"] * len(ids))
        cursor.execute(f"""
            UPDATE WO_Complaints
            SET StatusCode = 2,
                ClosedReason = %s,
                ClosedBy = %s,
                ClosedAt = SYSUTCDATETIME()
            WHERE ComplaintID IN ({ph}) AND StatusCode <> 2
        """, (reason, user, *ids))
        for cid in ids:
            _log_activity(cursor, "complaint", cid, "closed",
                          f"manager close: {reason}", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "closed_count": len(ids)})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/complaints/link", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_complaint_link():
    """Manager-only. Link multiple complaints into one issue group.
    GroupID = smallest ComplaintID in the selection (or current GroupID if any).
    Body: {complaint_ids: [int, ...]}"""
    data = request.get_json(silent=True) or {}
    ids = data.get("complaint_ids") or []
    if not isinstance(ids, list) or len(ids) < 2:
        return jsonify({"error": "Need 2+ complaint_ids to link."}), 400
    try:
        ids = sorted({int(i) for i in ids})
    except (TypeError, ValueError):
        return jsonify({"error": "complaint_ids must be integers."}), 400
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        ph = ",".join(["%s"] * len(ids))
        # If any selected complaint already belongs to a group, adopt the smallest existing GroupID
        cursor.execute(
            f"SELECT MIN(ISNULL(GroupID, ComplaintID)) FROM WO_Complaints WHERE ComplaintID IN ({ph})",
            tuple(ids),
        )
        group_id = int(cursor.fetchone()[0])
        cursor.execute(f"""
            UPDATE WO_Complaints SET GroupID = %s
            WHERE ComplaintID IN ({ph})
        """, (group_id, *ids))
        for cid in ids:
            _log_activity(cursor, "complaint", cid, "linked",
                          f"Linked into group {group_id}", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "group_id": group_id, "linked_count": len(ids)})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/complaints/oneoff-suggestions")
@require_roles(*MANAGER_ROLES)
def api_complaint_oneoff_suggestions():
    """Manager-only. Returns fresh complaints >48h old where the same MachineCode
    has had NO subsequent complaints (i.e. likely a one-off worth closing)."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT c.ComplaintID, c.DisplayID, c.MachineName, c.MachineCode,
                   c.Description, c.SubmittedAt, c.RefundIssued
            FROM WO_Complaints c
            WHERE c.StatusCode = 0
              AND c.SubmittedAt < DATEADD(hour, -48, SYSUTCDATETIME())
              AND NOT EXISTS (
                  SELECT 1 FROM WO_Complaints c2
                  WHERE c2.MachineCode = c.MachineCode
                    AND c2.ComplaintID <> c.ComplaintID
                    AND c2.SubmittedAt > c.SubmittedAt
              )
            ORDER BY c.SubmittedAt
        """)
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "id": r[0], "display_id": r[1],
            "machine_name": r[2], "machine_code": r[3],
            "description": (r[4] or "")[:160],
            "submitted_at": _iso(r[5]),
            "refund_issued": bool(r[6]),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Job Orders  →  "Tech Support" tab ─────────────────────────────────────────

@workorders_bp.route("/joborders", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_joborder_create():
    """Field manager creates a work order. Optionally linked from a complaint.
    Body:
        complaint_id   (optional int)
        machine_name   (required if no complaint)
        machine_code   (optional)
        diagnosis      (manager's initial hypothesis)
        proposed_fix   (suggested fix plan — free text)
        notes          (free text)
        priority       (0/1/2 — TINYINT; default 1)
        assigned_to    (email of operator)
        event_code     (optional int 800000–899999)
        tasks          (optional list of tickbox label strings — creates WO_JobOrderTasks)
    """
    data = request.get_json(silent=True) or {}
    complaint_id = data.get("complaint_id")
    machine_name = (data.get("machine_name") or "").strip()
    machine_code = (data.get("machine_code") or "").strip() or None
    diagnosis    = (data.get("diagnosis") or "").strip() or None
    proposed_fix = (data.get("proposed_fix") or "").strip() or None
    notes        = (data.get("notes") or "").strip() or None
    assigned     = (data.get("assigned_to") or "").strip().lower() or None
    tasks        = data.get("tasks") or []

    priority = data.get("priority", 1)
    try:
        priority = int(priority)
        if priority not in (0, 1, 2):
            priority = 1
    except (TypeError, ValueError):
        priority = 1

    event_code = data.get("event_code")
    if event_code not in (None, "", "null"):
        try:
            event_code = int(event_code)
            if not (800000 <= event_code <= 899999):
                return jsonify({"error": "JobOrder EventCode must be 800000–899999."}), 400
        except (TypeError, ValueError):
            return jsonify({"error": "EventCode must be an integer."}), 400
    else:
        event_code = None

    attached_kb_id = data.get("attached_kb_id")
    if attached_kb_id in (None, "", "null"):
        attached_kb_id = None
    else:
        try:
            attached_kb_id = int(attached_kb_id)
        except (TypeError, ValueError):
            return jsonify({"error": "attached_kb_id must be an integer."}), 400

    if complaint_id:
        try:
            complaint_id = int(complaint_id)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid complaint id."}), 400

    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # If from complaint, pull machine + event_code if not supplied
        if complaint_id:
            cursor.execute(
                "SELECT MachineName, MachineCode, EventCode FROM WO_Complaints WHERE ComplaintID = %s",
                (complaint_id,),
            )
            crow = cursor.fetchone()
            if not crow:
                conn.close()
                return jsonify({"error": "Complaint not found."}), 404
            if not machine_name and crow[0]:
                machine_name = crow[0]
            if not machine_code and crow[1]:
                machine_code = crow[1]

        if not machine_name:
            conn.close()
            return jsonify({"error": "Machine name is required."}), 400

        display_id = allocate_display_id(cursor, "WkO")

        cursor.execute("""
            INSERT INTO WO_JobOrders
                (ComplaintID, MachineName, MachineCode, Notes,
                 AssignedTo, PriorityCode, StatusCode,
                 EventCode, Diagnosis, ProposedFix, AttachedKBID,
                 DisplayID, CreatedBy)
            OUTPUT INSERTED.JobOrderID
            VALUES (%s, %s, %s, %s,
                    %s, %s, 0,
                    %s, %s, %s, %s,
                    %s, %s)
        """, (complaint_id, machine_name, machine_code, notes,
              assigned, priority,
              event_code, diagnosis, proposed_fix, attached_kb_id,
              display_id, user))
        new_id = int(cursor.fetchone()[0])

        _log_activity(cursor, "joborder", new_id, "created",
                      f"Created as {display_id} for {machine_name}", user)
        if assigned:
            _log_activity(cursor, "joborder", new_id, "assigned",
                          f"Assigned to {assigned}", user)

        # Tickboxes
        for i, label in enumerate(tasks):
            label = str(label).strip()
            if not label:
                continue
            cursor.execute("""
                INSERT INTO WO_JobOrderTasks (JobOrderID, SeqNum, Label, Done)
                VALUES (%s, %s, %s, 0)
            """, (new_id, i + 1, label[:500]))

        # If created from complaint, mark it (and ALL group-linked complaints) as assigned
        if complaint_id:
            cursor.execute(
                "SELECT GroupID FROM WO_Complaints WHERE ComplaintID = %s",
                (complaint_id,),
            )
            grow = cursor.fetchone()
            group_id = int(grow[0]) if (grow and grow[0] is not None) else None
            if group_id is not None:
                cursor.execute("""
                    UPDATE WO_Complaints SET JobOrderID = %s, StatusCode = 1
                    WHERE GroupID = %s
                """, (new_id, group_id))
                cursor.execute(
                    "SELECT ComplaintID FROM WO_Complaints WHERE GroupID = %s",
                    (group_id,),
                )
                linked_ids = [r[0] for r in cursor.fetchall()]
                for cid in linked_ids:
                    _log_activity(cursor, "complaint", cid, "linked",
                                  f"Group {group_id} → {display_id}", user)
            else:
                cursor.execute("""
                    UPDATE WO_Complaints SET JobOrderID = %s, StatusCode = 1
                    WHERE ComplaintID = %s
                """, (new_id, complaint_id))
                _log_activity(cursor, "complaint", complaint_id, "linked",
                              f"Linked to {display_id}", user)

        conn.commit()
        conn.close()
        return jsonify({
            "ok": True, "id": new_id, "display_id": display_id,
            "tasks_created": len([t for t in tasks if str(t).strip()]),
        })
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/joborders")
@require_roles(*OPERATOR_ROLES)
def api_joborder_list():
    """?scope=mine|all  ?status=assigned|needs_assistance|closed|all  ?machine_code=…
    Operators always see mine. Managers can see all.
    """
    scope         = (request.args.get("scope") or "mine").strip().lower()
    status_filter = (request.args.get("status") or "all").strip().lower()
    machine_code  = (request.args.get("machine_code") or "").strip() or None
    user          = get_current_user()

    if get_role(user) not in MANAGER_ROLES:
        scope = "mine"

    where, params = [], []
    if scope == "mine":
        where.append("AssignedTo = %s")
        params.append(user)

    rev = {v: k for k, v in JOBORDER_STATUS.items()}
    if status_filter in rev:
        where.append("StatusCode = %s")
        params.append(rev[status_filter])

    if machine_code:
        where.append("MachineCode = %s")
        params.append(machine_code)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sql = f"""
        SELECT TOP 200
            JobOrderID, DisplayID, ComplaintID,
            MachineName, MachineCode,
            AssignedTo, PriorityCode, StatusCode,
            EventCode, Diagnosis, ProposedFix, LastBlockReason,
            CreatedBy, CreatedAt, CompletedBy, CompletedAt,
            (SELECT COUNT(*) FROM WO_JobOrderTasks WHERE JobOrderID = wj.JobOrderID) AS TaskCount,
            (SELECT COUNT(*) FROM WO_JobOrderTasks
                WHERE JobOrderID = wj.JobOrderID AND Done = 1) AS TasksDone
        FROM WO_JobOrders wj
        {where_sql}
        ORDER BY CreatedAt DESC, JobOrderID DESC
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "id": r[0], "display_id": r[1], "complaint_id": r[2],
            "machine_name": r[3], "machine_code": r[4],
            "assigned_to": r[5],
            "priority_code":  int(r[6]),
            "priority_label": _label(PRIORITY, r[6]),
            "status_code":    int(r[7]),
            "status_label":   _label(JOBORDER_STATUS, r[7]),
            "event_code":     int(r[8]) if r[8] is not None else None,
            "diagnosis":      r[9],
            "proposed_fix":   r[10],
            "last_block_reason": r[11],
            "created_by": r[12], "created_at": _iso(r[13]),
            "completed_by": r[14], "completed_at": _iso(r[15]),
            "task_count": int(r[16] or 0),
            "tasks_done": int(r[17] or 0),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/joborders/<int:jid>")
@require_roles(*OPERATOR_ROLES)
def api_joborder_detail(jid):
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT JobOrderID, DisplayID, ComplaintID,
                   MachineName, MachineCode, Notes,
                   AssignedTo, PriorityCode, StatusCode,
                   Report, RootCause, CorrectiveAction, PreventiveAction,
                   EventCode, Diagnosis, ProposedFix, LastBlockReason,
                   CreatedBy, CreatedAt, CompletedBy, CompletedAt,
                   AttachedKBID, OnSiteObservations, OnSiteChanges, TechnicianComments
            FROM WO_JobOrders WHERE JobOrderID = %s
        """, (jid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Job order not found."}), 404

        # Operators can only see their own
        if get_role(user) not in MANAGER_ROLES and (row[6] or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not assigned to you."}), 403

        cursor.execute("""
            SELECT TaskID, SeqNum, Label, Done, BlockedNote, CompletedBy, CompletedAt
            FROM WO_JobOrderTasks WHERE JobOrderID = %s ORDER BY SeqNum, TaskID
        """, (jid,))
        tasks = [{
            "id": tr[0], "seq": tr[1], "label": tr[2],
            "done": bool(tr[3]), "blocked_note": tr[4],
            "completed_by": tr[5], "completed_at": _iso(tr[6]),
        } for tr in cursor.fetchall()]

        images   = _images_for(cursor, "joborder", jid)
        activity = _activity_for(cursor, "joborder", jid)

        # Attached KB article (if manager pre-attached one).
        attached_kb = None
        kb_id = row[21]
        if kb_id is not None:
            cursor.execute("""
                SELECT KBID, EventCode, Title, Symptom, DiagnosticConfirmation, RootCause,
                       CorrectiveAction, PreventiveAction, VerificationOfCompletion
                FROM WO_KB_Entries WHERE KBID = %s
            """, (kb_id,))
            kbr = cursor.fetchone()
            if kbr:
                cursor.execute(
                    "SELECT SeqNum, Label FROM WO_KB_Tickboxes WHERE KBID=%s ORDER BY SeqNum, TBID",
                    (kb_id,),
                )
                tbs = [{"seq": x[0], "label": x[1]} for x in cursor.fetchall()]
                attached_kb = {
                    "id": kbr[0],
                    "event_code": int(kbr[1]) if kbr[1] is not None else None,
                    "title": kbr[2],
                    "symptom": kbr[3],
                    "diagnostic_confirmation": kbr[4],
                    "root_cause": kbr[5],
                    "corrective_action": kbr[6],
                    "preventive_action": kbr[7],
                    "verification_of_completion": kbr[8],
                    "tickboxes": tbs,
                }

        # Linked complaints — show ALL in the group if grouped, else just the primary.
        linked = []
        if row[2] is not None:
            cursor.execute(
                "SELECT GroupID FROM WO_Complaints WHERE ComplaintID = %s",
                (row[2],),
            )
            grow = cursor.fetchone()
            gid = int(grow[0]) if (grow and grow[0] is not None) else None
            if gid is not None:
                cursor.execute("""
                    SELECT ComplaintID, DisplayID, Description, MachineName, MachineCode,
                           SubmitterEmail, SubmittedAt, ReportedBy, PerceivedUrgency,
                           ImpactSeverity, ImpactDescription, RefundIssued, StatusCode
                    FROM WO_Complaints WHERE GroupID = %s
                    ORDER BY SubmittedAt, ComplaintID
                """, (gid,))
            else:
                cursor.execute("""
                    SELECT ComplaintID, DisplayID, Description, MachineName, MachineCode,
                           SubmitterEmail, SubmittedAt, ReportedBy, PerceivedUrgency,
                           ImpactSeverity, ImpactDescription, RefundIssued, StatusCode
                    FROM WO_Complaints WHERE ComplaintID = %s
                """, (row[2],))
            for cr in cursor.fetchall():
                linked.append({
                    "id": cr[0], "display_id": cr[1], "description": cr[2],
                    "machine_name": cr[3], "machine_code": cr[4],
                    "submitter": cr[5], "submitted_at": _iso(cr[6]),
                    "reported_by": cr[7],
                    "perceived_urgency_code":  int(cr[8]),
                    "perceived_urgency_label": _label(PRIORITY, cr[8]),
                    "impact_severity":   int(cr[9]) if cr[9] is not None else None,
                    "impact_description": cr[10],
                    "refund_issued":     bool(cr[11]),
                    "status_code":       int(cr[12]),
                    "status_label":      _label(COMPLAINT_STATUS, cr[12]),
                })

        conn.close()
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500
    return jsonify({
        "id": row[0], "display_id": row[1], "complaint_id": row[2],
        "machine_name": row[3], "machine_code": row[4], "notes": row[5],
        "assigned_to": row[6],
        "priority_code": int(row[7]),  "priority_label": _label(PRIORITY, row[7]),
        "status_code":   int(row[8]),  "status_label":   _label(JOBORDER_STATUS, row[8]),
        "report":            row[9],
        "root_cause":        row[10],
        "corrective_action": row[11],
        "preventive_action": row[12],
        "event_code":     int(row[13]) if row[13] is not None else None,
        "diagnosis":      row[14],
        "proposed_fix":   row[15],
        "last_block_reason": row[16],
        "created_by": row[17], "created_at": _iso(row[18]),
        "completed_by": row[19], "completed_at": _iso(row[20]),
        "attached_kb_id":        int(row[21]) if row[21] is not None else None,
        "on_site_observations":  row[22],
        "on_site_changes":       row[23],
        "technician_comments":   row[24],
        "attached_kb":           attached_kb,
        "tasks": tasks,
        "linked_complaints": linked,
        "images": images, "activity": activity,
    })


@workorders_bp.route("/joborders/<int:jid>", methods=["PATCH"])
@require_roles(*MANAGER_ROLES)
def api_joborder_update(jid):
    """Manager edits a WO. Updatable fields: notes, priority, diagnosis,
    proposed_fix, event_code, report, root_cause, corrective_action,
    preventive_action."""
    data = request.get_json(silent=True) or {}
    sets, params = [], []

    if "notes" in data:
        sets.append("Notes = %s"); params.append(data["notes"] or None)
    if "diagnosis" in data:
        sets.append("Diagnosis = %s"); params.append(data["diagnosis"] or None)
    if "proposed_fix" in data:
        sets.append("ProposedFix = %s"); params.append(data["proposed_fix"] or None)
    if "report" in data:
        sets.append("Report = %s"); params.append(data["report"] or None)
    if "root_cause" in data:
        sets.append("RootCause = %s"); params.append(data["root_cause"] or None)
    if "corrective_action" in data:
        sets.append("CorrectiveAction = %s"); params.append(data["corrective_action"] or None)
    if "preventive_action" in data:
        sets.append("PreventiveAction = %s"); params.append(data["preventive_action"] or None)
    if "priority" in data:
        try:
            p = int(data["priority"])
            if p in (0, 1, 2):
                sets.append("PriorityCode = %s"); params.append(p)
        except (TypeError, ValueError):
            pass
    if "event_code" in data:
        ev = data["event_code"]
        if ev in (None, "", "null"):
            sets.append("EventCode = NULL")
        else:
            try:
                ev = int(ev)
                if not (800000 <= ev <= 899999):
                    return jsonify({"error": "EventCode must be 800000–899999."}), 400
                sets.append("EventCode = %s"); params.append(ev)
            except (TypeError, ValueError):
                return jsonify({"error": "EventCode must be an integer."}), 400
    if "attached_kb_id" in data:
        kb = data["attached_kb_id"]
        if kb in (None, "", "null"):
            sets.append("AttachedKBID = NULL")
        else:
            try:
                sets.append("AttachedKBID = %s"); params.append(int(kb))
            except (TypeError, ValueError):
                return jsonify({"error": "attached_kb_id must be an integer."}), 400

    if not sets:
        return jsonify({"error": "No updatable fields supplied."}), 400

    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        params.append(jid)
        cursor.execute(
            f"UPDATE WO_JobOrders SET {', '.join(sets)} WHERE JobOrderID = %s",
            tuple(params),
        )
        _log_activity(cursor, "joborder", jid, "updated",
                      "; ".join(s.split(" = ")[0] for s in sets), user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/joborders/<int:jid>/assign", methods=["POST"])
@require_roles(*DISPATCH_ROLES)
def api_joborder_assign(jid):
    data = request.get_json(silent=True) or {}
    assigned = (data.get("assigned_to") or "").strip().lower() or None
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        # This route used to force StatusCode=0 unconditionally. Harmless while
        # it was manager-only, dangerous once dispatch could call it: it would
        # resurrect a job order the manager had already closed (3) or accepted
        # into review (2), leaving an open WO hanging off a closed complaint and
        # silently emptying the review queue. Reassignment is now refused for
        # those two states, and only clears needs_assistance (1) back to
        # assigned (0), which is the one case where a new owner is the point.
        cursor.execute("SELECT StatusCode FROM WO_JobOrders WHERE JobOrderID = %s", (jid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Job order not found."}), 404
        sc = int(row[0] or 0)
        if sc == 2:
            conn.close()
            return jsonify({"error": "Job order is pending manager review; it cannot be reassigned."}), 400
        if sc == 3:
            conn.close()
            return jsonify({"error": "Job order is closed; it cannot be reassigned."}), 400
        sets, params = ["AssignedTo=%s", "StatusCode=0"], [assigned]
        _day = _parse_date(data.get("scheduled_date"))
        if _day:
            sets.append("ScheduledDate=%s"); params.append(_day)
        params.append(jid)
        cursor.execute(
            f"UPDATE WO_JobOrders SET {', '.join(sets)} WHERE JobOrderID=%s",
            tuple(params),
        )
        _log_activity(cursor, "joborder", jid, "assigned",
                      f"Assigned to {assigned or '(unassigned)'}"
                      + (" (cleared needs_assistance)" if sc == 1 else ""), user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/joborders/<int:jid>/status", methods=["POST"])
@require_roles(*OPERATOR_ROLES)
def api_joborder_status(jid):
    """Operator updates status:
        0 assigned, 1 needs_assistance, 2 pending_review.
    Status 3 (closed) is set ONLY by the manager via /review endpoint.

    For status=1: body MUST include 'block_reason' (stored in LastBlockReason).
    For status=2: operator marks work done, awaiting manager review.
                  Does NOT close the complaint — manager decides at /review.
    """
    data = request.get_json(silent=True) or {}
    user = get_current_user()
    try:
        status_code = int(data.get("status_code"))
    except (TypeError, ValueError):
        return jsonify({"error": "status_code (0/1/2) required."}), 400
    if status_code not in (0, 1, 2):
        return jsonify({"error": "Operator can only set 0, 1, or 2. Manager closes via /review."}), 400

    block_reason = (data.get("block_reason") or "").strip() or None
    if status_code == 1 and not block_reason:
        return jsonify({"error": "block_reason required when marking needs_assistance."}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT AssignedTo, DisplayID FROM WO_JobOrders WHERE JobOrderID=%s",
            (jid,),
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Job order not found."}), 404
        if get_role(user) not in MANAGER_ROLES and (row[0] or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not assigned to you."}), 403

        if status_code == 2:
            # Operator marks "done — pending review". No complaint close yet.
            cursor.execute("""
                UPDATE WO_JobOrders
                SET StatusCode=2, LastBlockReason=NULL,
                    CompletedBy=%s, CompletedAt=SYSUTCDATETIME()
                WHERE JobOrderID=%s
            """, (user, jid))
        else:
            cursor.execute("""
                UPDATE WO_JobOrders
                SET StatusCode=%s, LastBlockReason=%s
                WHERE JobOrderID=%s
            """, (status_code, block_reason, jid))

        _log_activity(cursor, "joborder", jid, "status",
                      f"{_label(JOBORDER_STATUS, status_code)}"
                      + (f" — {block_reason}" if block_reason else ""),
                      user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Driver input: on-site observations, changes done, comments ────────────────

@workorders_bp.route("/joborders/<int:jid>/driver-input", methods=["PATCH"])
@require_roles(*OPERATOR_ROLES)
def api_joborder_driver_input(jid):
    """Driver (operator) saves their on-site write-up. Does NOT change status —
    use /status for that (typically status=2 'pending_review' on submit).

    Body any of:
        on_site_observations: str
        on_site_changes:      str
        technician_comments:  str
    """
    data = request.get_json(silent=True) or {}
    user = get_current_user()
    sets, params = [], []
    if "on_site_observations" in data:
        sets.append("OnSiteObservations = %s")
        params.append((data["on_site_observations"] or "").strip() or None)
    if "on_site_changes" in data:
        sets.append("OnSiteChanges = %s")
        params.append((data["on_site_changes"] or "").strip() or None)
    if "technician_comments" in data:
        sets.append("TechnicianComments = %s")
        params.append((data["technician_comments"] or "").strip() or None)
    if not sets:
        return jsonify({"error": "No fields to save."}), 400
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT AssignedTo FROM WO_JobOrders WHERE JobOrderID = %s",
            (jid,),
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Job order not found."}), 404
        if get_role(user) not in MANAGER_ROLES and (row[0] or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not assigned to you."}), 403
        params.append(jid)
        cursor.execute(
            f"UPDATE WO_JobOrders SET {', '.join(sets)} WHERE JobOrderID = %s",
            tuple(params),
        )
        _log_activity(cursor, "joborder", jid, "driver_input",
                      "; ".join(s.split(" = ")[0] for s in sets), user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Manager review (the "review loop" for WO completion) ─────────────────────

@workorders_bp.route("/joborders/<int:jid>/review", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_joborder_review(jid):
    """Manager reviews an operator-submitted WO (StatusCode=2 pending_review).
    Body: {decision: 'accept'|'reject', notes: str (optional)}

    accept → WO StatusCode=3 (closed). If linked complaint, close it (StatusCode=2).
    reject → WO StatusCode=3 (closed, superseded). Create a NEW WO inheriting the
             complaint link + machine + assignment. Manager later edits the new WO
             with revised diagnosis/proposed_fix.
    """
    data = request.get_json(silent=True) or {}
    decision = (data.get("decision") or "").strip().lower()
    notes    = (data.get("notes") or "").strip() or None
    if decision not in ("accept", "reject"):
        return jsonify({"error": "decision must be 'accept' or 'reject'."}), 400

    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT JobOrderID, ComplaintID, MachineName, MachineCode,
                   AssignedTo, PriorityCode, EventCode, StatusCode, DisplayID
            FROM WO_JobOrders WHERE JobOrderID = %s
        """, (jid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Job order not found."}), 404
        if int(row[7]) != 2:
            conn.close()
            return jsonify({"error": "WO is not pending_review."}), 400

        (_, complaint_id, machine_name, machine_code,
         assigned, priority_code, event_code, _, old_display) = row

        # Close the reviewed WO regardless of decision.
        cursor.execute("UPDATE WO_JobOrders SET StatusCode = 3 WHERE JobOrderID = %s", (jid,))
        _log_activity(cursor, "joborder", jid, "manager_review",
                      f"{decision}" + (f" — {notes}" if notes else ""), user)

        new_wo_id = None
        new_display = None
        if decision == "accept":
            if complaint_id:
                # If complaint is in a group, close ALL complaints in that group
                cursor.execute(
                    "SELECT GroupID FROM WO_Complaints WHERE ComplaintID = %s",
                    (complaint_id,),
                )
                grow = cursor.fetchone()
                group_id = int(grow[0]) if (grow and grow[0] is not None) else None
                if group_id is not None:
                    cursor.execute("""
                        UPDATE WO_Complaints
                        SET StatusCode = 2,
                            ClosedReason = %s,
                            ClosedBy = %s,
                            ClosedAt = SYSUTCDATETIME()
                        WHERE GroupID = %s
                    """, (f"Resolved via {old_display}", user, group_id))
                    cursor.execute(
                        "SELECT ComplaintID FROM WO_Complaints WHERE GroupID = %s",
                        (group_id,),
                    )
                    for r in cursor.fetchall():
                        _log_activity(cursor, "complaint", r[0], "closed",
                                      f"Group close via accept of {old_display}", user)
                else:
                    cursor.execute("""
                        UPDATE WO_Complaints
                        SET StatusCode = 2,
                            ClosedReason = %s,
                            ClosedBy = %s,
                            ClosedAt = SYSUTCDATETIME()
                        WHERE ComplaintID = %s
                    """, (f"Resolved via {old_display}", user, complaint_id))
                    _log_activity(cursor, "complaint", complaint_id, "closed",
                                  f"Closed via accept of {old_display}", user)
        else:
            # reject: WO is closed; complaints go to UNRESOLVED (StatusCode=3).
            # JobOrderID stays pointing to the failed WO so manager can review it.
            # No auto-follow-up — manager re-triages and creates a new WO when ready.
            if complaint_id:
                cursor.execute(
                    "SELECT GroupID FROM WO_Complaints WHERE ComplaintID = %s",
                    (complaint_id,),
                )
                grow = cursor.fetchone()
                group_id = int(grow[0]) if (grow and grow[0] is not None) else None
                if group_id is not None:
                    cursor.execute("""
                        UPDATE WO_Complaints SET StatusCode = 3
                        WHERE GroupID = %s
                    """, (group_id,))
                    cursor.execute(
                        "SELECT ComplaintID FROM WO_Complaints WHERE GroupID = %s",
                        (group_id,),
                    )
                    for r in cursor.fetchall():
                        _log_activity(cursor, "complaint", r[0], "unresolved",
                                      f"Group {group_id}: {old_display} rejected — {notes or 'no notes'}", user)
                else:
                    cursor.execute("""
                        UPDATE WO_Complaints SET StatusCode = 3
                        WHERE ComplaintID = %s
                    """, (complaint_id,))
                    _log_activity(cursor, "complaint", complaint_id, "unresolved",
                                  f"{old_display} rejected — {notes or 'no notes'}", user)

        conn.commit()
        conn.close()
        return jsonify({
            "ok": True,
            "decision": decision,
            "new_wo_id": new_wo_id,
            "new_display_id": new_display,
        })
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── JobOrder Tasks (tickboxes per WO) ─────────────────────────────────────────

@workorders_bp.route("/joborders/<int:jid>/tasks", methods=["GET"])
@require_roles(*OPERATOR_ROLES)
def api_tasks_list(jid):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TaskID, SeqNum, Label, Done, BlockedNote, CompletedBy, CompletedAt
            FROM WO_JobOrderTasks WHERE JobOrderID = %s ORDER BY SeqNum, TaskID
        """, (jid,))
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "id": r[0], "seq": r[1], "label": r[2],
            "done": bool(r[3]), "blocked_note": r[4],
            "completed_by": r[5], "completed_at": _iso(r[6]),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/joborders/<int:jid>/tasks", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_tasks_add(jid):
    """Body: {label, seq?}  — manager adds a tickbox to an existing WO."""
    data = request.get_json(silent=True) or {}
    label = (data.get("label") or "").strip()
    if not label:
        return jsonify({"error": "Task label required."}), 400
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        seq = data.get("seq")
        if seq is None:
            cursor.execute(
                "SELECT ISNULL(MAX(SeqNum), 0) + 1 FROM WO_JobOrderTasks WHERE JobOrderID = %s",
                (jid,),
            )
            seq = int(cursor.fetchone()[0])
        else:
            try:
                seq = int(seq)
            except (TypeError, ValueError):
                seq = 1
        cursor.execute("""
            INSERT INTO WO_JobOrderTasks (JobOrderID, SeqNum, Label, Done)
            OUTPUT INSERTED.TaskID
            VALUES (%s, %s, %s, 0)
        """, (jid, seq, label[:500]))
        new_id = int(cursor.fetchone()[0])
        _log_activity(cursor, "joborder", jid, "task_added", label[:200], user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id, "seq": seq})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/tasks/<int:tid>", methods=["PATCH"])
@require_roles(*OPERATOR_ROLES)
def api_task_update(tid):
    """Body: {done?: bool, blocked_note?: str, label?: str (manager only)}
    Optional 'image_data_url' uploads a per-step photo (stage='task_done' or
    'task_blocked' depending on Done state).
    """
    data = request.get_json(silent=True) or {}
    user = get_current_user()
    role = get_role(user)
    sets, params = [], []
    stage = None

    if "label" in data:
        if role not in MANAGER_ROLES:
            return jsonify({"error": "Only managers can edit task labels."}), 403
        sets.append("Label = %s")
        params.append(str(data["label"])[:500])

    if "done" in data:
        done = bool(data["done"])
        sets.append("Done = %s"); params.append(1 if done else 0)
        if done:
            sets.append("CompletedBy = %s"); params.append(user)
            sets.append("CompletedAt = SYSUTCDATETIME()")
            stage = "task_done"
        else:
            sets.append("CompletedBy = NULL")
            sets.append("CompletedAt = NULL")

    if "blocked_note" in data:
        bn = (data["blocked_note"] or "").strip() or None
        sets.append("BlockedNote = %s"); params.append(bn)
        if bn:
            stage = stage or "task_blocked"

    if not sets:
        return jsonify({"error": "No updatable fields supplied."}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor()
        # Ensure task exists; capture parent JobOrderID for activity log
        cursor.execute("SELECT JobOrderID FROM WO_JobOrderTasks WHERE TaskID=%s", (tid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Task not found."}), 404
        jid = int(row[0])

        # Operators must be assigned to the parent WO
        if role not in MANAGER_ROLES:
            cursor.execute("SELECT AssignedTo FROM WO_JobOrders WHERE JobOrderID=%s", (jid,))
            ar = cursor.fetchone()
            if not ar or (ar[0] or "").lower() != user.lower():
                conn.close()
                return jsonify({"error": "Not assigned to you."}), 403

        params_tuple = tuple(params) + (tid,)
        # Stitch '= SYSUTCDATETIME()' literal back into the SET clause
        cursor.execute(
            f"UPDATE WO_JobOrderTasks SET {', '.join(sets)} WHERE TaskID = %s",
            params_tuple,
        )

        # Optional per-step image
        img_data_url = data.get("image_data_url")
        if img_data_url and stage:
            raw, ctype = _decode_data_url(img_data_url)
            if raw:
                _save_image_to_sp(
                    cursor, parent_type="task", parent_id=tid, stage=stage,
                    file_name=f"task-{tid}-{stage}.jpg",
                    content_type=ctype or "image/jpeg",
                    raw_bytes=raw, uploaded_by=user,
                )

        _log_activity(cursor, "joborder", jid, "task_update",
                      f"task#{tid}: " + ", ".join(s.split(" = ")[0] for s in sets), user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/tasks/<int:tid>", methods=["DELETE"])
@require_roles(*MANAGER_ROLES)
def api_task_delete(tid):
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT JobOrderID FROM WO_JobOrderTasks WHERE TaskID=%s", (tid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Task not found."}), 404
        jid = int(row[0])
        cursor.execute("DELETE FROM WO_JobOrderTasks WHERE TaskID=%s", (tid,))
        _log_activity(cursor, "joborder", jid, "task_deleted", f"task#{tid}", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Knowledge Base (Manage-KB sub-tab) ────────────────────────────────────────

@workorders_bp.route("/kb")
@api_login_required
def api_kb_list():
    """All KB entries. Manager-only edit; everyone can read so suggestions work."""
    event_code = request.args.get("event_code")
    where, params = [], []
    if event_code:
        try:
            where.append("EventCode = %s")
            params.append(int(event_code))
        except (TypeError, ValueError):
            return jsonify({"error": "event_code must be an integer."}), 400
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(f"""
            SELECT KBID, EventCode, Title, UseCount,
                   CreatedBy, CreatedAt, UpdatedBy, UpdatedAt,
                   Symptom, DiagnosticConfirmation, RootCause,
                   CorrectiveAction, PreventiveAction, VerificationOfCompletion
            FROM WO_KB_Entries {where_sql}
            ORDER BY UseCount DESC, KBID DESC
        """, tuple(params))
        rows = cursor.fetchall()
        # Pre-fetch tickboxes grouped by KBID
        cursor.execute("""
            SELECT TBID, KBID, SeqNum, Label
            FROM WO_KB_Tickboxes ORDER BY KBID, SeqNum, TBID
        """)
        by_kb = {}
        for tb in cursor.fetchall():
            by_kb.setdefault(int(tb[1]), []).append(
                {"id": tb[0], "seq": tb[2], "label": tb[3]}
            )
        conn.close()
        return jsonify([{
            "id": r[0],
            "event_code": int(r[1]) if r[1] is not None else None,
            "title": r[2],
            "use_count": int(r[3]),
            "created_by": r[4], "created_at": _iso(r[5]),
            "updated_by": r[6], "updated_at": _iso(r[7]),
            "symptom":                    r[8],
            "diagnostic_confirmation":    r[9],
            "root_cause":                 r[10],
            "corrective_action":          r[11],
            "preventive_action":          r[12],
            "verification_of_completion": r[13],
            "tickboxes": by_kb.get(int(r[0]), []),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/kb", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_kb_create():
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"error": "Title required."}), 400
    ev = data.get("event_code")
    if ev not in (None, "", "null"):
        try:
            ev = int(ev)
            if not ((600000 <= ev <= 699999) or (800000 <= ev <= 899999)):
                return jsonify({"error": "EventCode must be 6xxxxx or 8xxxxx."}), 400
        except (TypeError, ValueError):
            return jsonify({"error": "EventCode must be an integer."}), 400
    else:
        ev = None
    car = {
        "symptom":                    (data.get("symptom") or "").strip() or None,
        "diagnostic_confirmation":    (data.get("diagnostic_confirmation") or "").strip() or None,
        "root_cause":                 (data.get("root_cause") or "").strip() or None,
        "corrective_action":          (data.get("corrective_action") or "").strip() or None,
        "preventive_action":          (data.get("preventive_action") or "").strip() or None,
        "verification_of_completion": (data.get("verification_of_completion") or "").strip() or None,
    }
    tickboxes = data.get("tickboxes") or []
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO WO_KB_Entries
                (EventCode, Title, CreatedBy,
                 Symptom, DiagnosticConfirmation, RootCause,
                 CorrectiveAction, PreventiveAction, VerificationOfCompletion)
            OUTPUT INSERTED.KBID
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (ev, title[:255], user,
              car["symptom"], car["diagnostic_confirmation"], car["root_cause"],
              car["corrective_action"], car["preventive_action"], car["verification_of_completion"]))
        new_id = int(cursor.fetchone()[0])
        for i, lbl in enumerate(tickboxes):
            lbl = str(lbl).strip()
            if not lbl:
                continue
            cursor.execute("""
                INSERT INTO WO_KB_Tickboxes (KBID, SeqNum, Label)
                VALUES (%s, %s, %s)
            """, (new_id, i + 1, lbl[:500]))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/kb/<int:kid>", methods=["PATCH"])
@require_roles(*MANAGER_ROLES)
def api_kb_update(kid):
    data = request.get_json(silent=True) or {}
    sets, params = [], []
    if "title" in data:
        sets.append("Title = %s"); params.append(str(data["title"])[:255])
    _car_map = {
        "symptom":                    "Symptom",
        "diagnostic_confirmation":    "DiagnosticConfirmation",
        "root_cause":                 "RootCause",
        "corrective_action":          "CorrectiveAction",
        "preventive_action":          "PreventiveAction",
        "verification_of_completion": "VerificationOfCompletion",
    }
    for k, col in _car_map.items():
        if k in data:
            sets.append(f"{col} = %s")
            params.append((data[k] or "").strip() or None)
    if "event_code" in data:
        ev = data["event_code"]
        if ev in (None, "", "null"):
            sets.append("EventCode = NULL")
        else:
            try:
                ev = int(ev)
                if not ((600000 <= ev <= 699999) or (800000 <= ev <= 899999)):
                    return jsonify({"error": "EventCode must be 6xxxxx or 8xxxxx."}), 400
                sets.append("EventCode = %s"); params.append(ev)
            except (TypeError, ValueError):
                return jsonify({"error": "EventCode must be an integer."}), 400
    user = get_current_user()
    sets.append("UpdatedBy = %s"); params.append(user)
    sets.append("UpdatedAt = SYSUTCDATETIME()")
    params.append(kid)
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            f"UPDATE WO_KB_Entries SET {', '.join(sets)} WHERE KBID = %s",
            tuple(params),
        )
        # Replace tickboxes if provided
        if "tickboxes" in data:
            cursor.execute("DELETE FROM WO_KB_Tickboxes WHERE KBID=%s", (kid,))
            for i, lbl in enumerate(data["tickboxes"] or []):
                lbl = str(lbl).strip()
                if not lbl:
                    continue
                cursor.execute("""
                    INSERT INTO WO_KB_Tickboxes (KBID, SeqNum, Label)
                    VALUES (%s, %s, %s)
                """, (kid, i + 1, lbl[:500]))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/kb/<int:kid>", methods=["DELETE"])
@require_roles(*MANAGER_ROLES)
def api_kb_delete(kid):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        snap = _row_snapshot(cursor, "SELECT * FROM WO_KB_Entries WHERE KBID=%s", (kid,))
        tbs = []
        cursor.execute("SELECT * FROM WO_KB_Tickboxes WHERE KBID=%s", (kid,))
        _cols = [d[0] for d in cursor.description]
        for r in cursor.fetchall():
            tbs.append({c: v for c, v in zip(_cols, r)})
        if snap is not None:
            snap["_tickboxes"] = tbs
        cursor.execute("DELETE FROM WO_KB_Tickboxes WHERE KBID=%s", (kid,))
        cursor.execute("DELETE FROM WO_KB_Entries WHERE KBID=%s", (kid,))
        log_deletion(cursor, "kb", kid,
                     f"KB entry {(snap or {}).get('Title') or ('#'+str(kid))}",
                     snap, False, get_current_user())
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/kb/suggest")
@api_login_required
def api_kb_suggest():
    """?event_code=8xxxxx → ranked list of KB entries matching that EventCode
    (UseCount DESC), each with its tickbox list. Used by the WO-creation form
    to pre-fill Diagnosis / ProposedFix / Tasks."""
    ev = request.args.get("event_code")
    if not ev:
        return jsonify([])
    try:
        ev = int(ev)
    except (TypeError, ValueError):
        return jsonify({"error": "event_code must be an integer."}), 400
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT KBID, EventCode, Title, UseCount,
                   Symptom, DiagnosticConfirmation, RootCause,
                   CorrectiveAction, PreventiveAction, VerificationOfCompletion
            FROM WO_KB_Entries WHERE EventCode = %s
            ORDER BY UseCount DESC, KBID DESC
        """, (ev,))
        rows = cursor.fetchall()
        out = []
        for r in rows:
            kb_id = int(r[0])
            cursor.execute("""
                SELECT SeqNum, Label FROM WO_KB_Tickboxes
                WHERE KBID = %s ORDER BY SeqNum, TBID
            """, (kb_id,))
            tbs = [{"seq": x[0], "label": x[1]} for x in cursor.fetchall()]
            out.append({
                "id": kb_id, "event_code": int(r[1]) if r[1] is not None else None,
                "title": r[2], "use_count": int(r[3]),
                "symptom":                    r[4],
                "diagnostic_confirmation":    r[5],
                "root_cause":                 r[6],
                "corrective_action":          r[7],
                "preventive_action":          r[8],
                "verification_of_completion": r[9],
                "tickboxes": tbs,
            })
        conn.close()
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/kb/<int:kid>/use", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_kb_use(kid):
    """Increment UseCount when a manager accepts a KB suggestion into a WO."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE WO_KB_Entries SET UseCount = UseCount + 1 WHERE KBID = %s",
            (kid,),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Admin: test-data cleanup (admin only, remove after stress test) ───────────

def _cascade_delete_joborder(cursor, jid):
    """Delete a WO and everything attached. Returns list of SP item IDs to remove."""
    sp_ids = []
    # WO-direct images
    cursor.execute(
        "SELECT SPItemID FROM WO_Images WHERE ParentType='joborder' AND ParentID=%s",
        (jid,),
    )
    sp_ids.extend([r[0] for r in cursor.fetchall() if r[0]])
    cursor.execute(
        "DELETE FROM WO_Images WHERE ParentType='joborder' AND ParentID=%s",
        (jid,),
    )
    # Task images
    cursor.execute("""
        SELECT i.SPItemID FROM WO_Images i
        INNER JOIN WO_JobOrderTasks t ON t.TaskID = i.ParentID
        WHERE i.ParentType='task' AND t.JobOrderID = %s
    """, (jid,))
    sp_ids.extend([r[0] for r in cursor.fetchall() if r[0]])
    cursor.execute("""
        DELETE FROM WO_Images WHERE ParentType='task' AND ParentID IN
            (SELECT TaskID FROM WO_JobOrderTasks WHERE JobOrderID = %s)
    """, (jid,))
    # Tasks, activity, complaint unlink, the WO
    cursor.execute("DELETE FROM WO_JobOrderTasks WHERE JobOrderID=%s", (jid,))
    cursor.execute(
        "DELETE FROM WO_Activity WHERE ParentType='joborder' AND ParentID=%s",
        (jid,),
    )
    cursor.execute(
        "UPDATE WO_Complaints SET JobOrderID=NULL, StatusCode=0 WHERE JobOrderID=%s",
        (jid,),
    )
    cursor.execute("DELETE FROM WO_JobOrders WHERE JobOrderID=%s", (jid,))
    return sp_ids


def _cascade_delete_complaint(cursor, cid):
    """Delete a complaint, its linked WO (if any), and all children."""
    sp_ids = []
    cursor.execute("SELECT JobOrderID FROM WO_Complaints WHERE ComplaintID=%s", (cid,))
    row = cursor.fetchone()
    if row and row[0]:
        sp_ids.extend(_cascade_delete_joborder(cursor, int(row[0])))
    cursor.execute(
        "SELECT SPItemID FROM WO_Images WHERE ParentType='complaint' AND ParentID=%s",
        (cid,),
    )
    sp_ids.extend([r[0] for r in cursor.fetchall() if r[0]])
    cursor.execute(
        "DELETE FROM WO_Images WHERE ParentType='complaint' AND ParentID=%s",
        (cid,),
    )
    cursor.execute(
        "DELETE FROM WO_Activity WHERE ParentType='complaint' AND ParentID=%s",
        (cid,),
    )
    cursor.execute("DELETE FROM WO_Complaints WHERE ComplaintID=%s", (cid,))
    return sp_ids


def _delete_sp_items(sp_ids):
    """Best-effort SP cleanup. Failures logged but not raised."""
    for sp_id in sp_ids:
        if not sp_id:
            continue
        try:
            sp.delete_item(sp_id)
        except Exception as e:
            print(f"[admin_delete] SP delete failed for {sp_id}: {e}")


def _row_snapshot(cursor, sql, params):
    """Return the first matching row as a {column: value} dict (for the deletion log)."""
    cursor.execute(sql, params)
    row = cursor.fetchone()
    if not row:
        return None
    cols = [d[0] for d in cursor.description]
    return {c: v for c, v in zip(cols, row)}


@workorders_bp.route("/admin/complaint/<int:cid>", methods=["DELETE"])
@require_roles(ROLE_ADMIN)
def api_admin_delete_complaint(cid):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        snap = _row_snapshot(cursor, "SELECT * FROM WO_Complaints WHERE ComplaintID=%s", (cid,))
        sp_ids = _cascade_delete_complaint(cursor, cid)
        log_deletion(cursor, "complaint", cid,
                     f"Complaint {(snap or {}).get('DisplayID') or ('#' + str(cid))}",
                     snap, False, get_current_user())
        conn.commit()
        conn.close()
    except Exception as e:
        return jsonify({"error": f"Delete failed: {str(e)}"}), 500
    _delete_sp_items(sp_ids)
    return jsonify({"ok": True, "sp_deleted": len(sp_ids)})


@workorders_bp.route("/admin/joborder/<int:jid>", methods=["DELETE"])
@require_roles(ROLE_ADMIN)
def api_admin_delete_joborder(jid):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        snap = _row_snapshot(cursor, "SELECT * FROM WO_JobOrders WHERE JobOrderID=%s", (jid,))
        sp_ids = _cascade_delete_joborder(cursor, jid)
        log_deletion(cursor, "joborder", jid,
                     f"Work order {(snap or {}).get('DisplayID') or ('#' + str(jid))}",
                     snap, False, get_current_user())
        conn.commit()
        conn.close()
    except Exception as e:
        return jsonify({"error": f"Delete failed: {str(e)}"}), 500
    _delete_sp_items(sp_ids)
    return jsonify({"ok": True, "sp_deleted": len(sp_ids)})


@workorders_bp.route("/admin/movementorder/<int:mid>", methods=["DELETE"])
@require_roles(ROLE_ADMIN)
def api_admin_delete_movementorder(mid):
    """Delete a movement order. If it was COMPLETED, first UNWIND its effects so
    the machine and its dated history return to the pre-move state — makes testing
    clean and lets you correct a wrongly-completed move. Only safe to undo the most
    recent move for a machine (no later moves/edits layered on top)."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT MovementType, MachineCode, FromLocation, FromLat, FromLon, StatusCode
            FROM WO_MovementOrders WITH (UPDLOCK, HOLDLOCK) WHERE MovementOrderID = %s
        """, (mid,))
        row = cursor.fetchone()

        if row and int(row[5]) == 2:   # completed → revert its effects (fail-closed)
            mtype, code, from_loc, from_lat, from_lon, _ = row
            code = str(code)

            # GUARD: only auto-undo the machine's MOST RECENT completed move. Undoing
            # an older move would corrupt a history that later moves have built on.
            cursor.execute("""
                SELECT TOP 1 MovementOrderID FROM WO_MovementOrders
                WHERE CAST(MachineCode AS NVARCHAR(50)) = %s AND StatusCode = 2
                ORDER BY CompletedAt DESC, MovementOrderID DESC
            """, (code,))
            latest = cursor.fetchone()
            if not latest or int(latest[0]) != mid:
                conn.close()
                return jsonify({"error": "This isn't the machine's most recent completed move — "
                                         "undo the later move(s) first."}), 409

            # GUARD: if this machine's history was hand-edited (record-move / historic /
            # corrective), the move's interval may have been split — auto-undo could
            # mis-locate it. Fail closed and require manual correction.
            cursor.execute("""
                SELECT COUNT(*) FROM MachineLocationHistory
                WHERE MachineCode = %s AND Source IN ('record-move','historic','corrective')
            """, (code,))
            if cursor.fetchone()[0] > 0:
                conn.close()
                return jsonify({"error": "This machine has manually-recorded moves in its history, "
                                         "so auto-undo is disabled to avoid corrupting it. Remove the "
                                         "recorded move from the History card first, or adjust manually."}), 409

            # 1. Undo the dated history split — branch on TYPE, never guess.
            if mtype in ("deploy", "relocate"):
                cursor.execute("""
                    SELECT HistoryID, ValidFromOle FROM MachineLocationHistory
                    WHERE MovementOrderID = %s AND ValidToOle IS NULL
                """, (mid,))
                hrow = cursor.fetchone()
                if not hrow:
                    conn.close()
                    return jsonify({"error": "This move's history was already changed — "
                                             "can't auto-undo; adjust history manually."}), 409
                cut = hrow[1]
                # delete the interval this move OPENED first (so there's never two open
                # intervals — the filtered unique index is never violated),
                cursor.execute("DELETE FROM MachineLocationHistory WHERE HistoryID = %s", (hrow[0],))
                # then re-open the single interval it had CLOSED at the same cutoff
                # (target by HistoryID so we can never re-open two rows).
                cursor.execute("""
                    SELECT TOP 1 HistoryID FROM MachineLocationHistory
                    WHERE MachineCode = %s AND ValidToOle = %s ORDER BY ValidFromOle DESC
                """, (code, cut))
                prior = cursor.fetchone()
                if prior:
                    cursor.execute("UPDATE MachineLocationHistory SET ValidToOle = NULL WHERE HistoryID = %s",
                                   (prior[0],))
            else:   # retrieve: it only CLOSED the open interval (inserted nothing)
                cursor.execute("""
                    SELECT COUNT(*) FROM MachineLocationHistory
                    WHERE MachineCode = %s AND ValidToOle IS NULL
                """, (code,))
                if cursor.fetchone()[0] != 0:
                    conn.close()
                    return jsonify({"error": "Unexpected open interval — can't auto-undo this retrieve."}), 409
                cursor.execute("""
                    SELECT TOP 1 HistoryID FROM MachineLocationHistory
                    WHERE MachineCode = %s ORDER BY ValidToOle DESC
                """, (code,))
                rr = cursor.fetchone()
                if rr:
                    cursor.execute("UPDATE MachineLocationHistory SET ValidToOle = NULL WHERE HistoryID = %s",
                                   (rr[0],))
            # 2. Restore MachineLookup to the pre-move state.
            if mtype == "relocate":
                cursor.execute("""
                    UPDATE MachineLookup SET MachineName=%s, Latitude=%s, Longitude=%s, IsActive=1
                    WHERE MachineCode=%s
                """, (from_loc, from_lat, from_lon, code))
            elif mtype == "deploy":
                cursor.execute("UPDATE MachineLookup SET IsActive=0 WHERE MachineCode=%s", (code,))
            elif mtype == "retrieve":
                cursor.execute("""
                    UPDATE MachineLookup
                    SET IsActive=1, DecommissionedAt=NULL, DecommissionReason=NULL, MachineName=%s
                    WHERE MachineCode=%s
                """, (from_loc, code))

        snap = _row_snapshot(cursor, "SELECT * FROM WO_MovementOrders WHERE MovementOrderID=%s", (mid,))
        log_deletion(cursor, "movementorder", mid,
                     f"Movement {(snap or {}).get('DisplayID') or ('#'+str(mid))}"
                     + (" (completed move reverted on delete)" if row and int(row[5]) == 2 else ""),
                     snap, False, get_current_user())
        cursor.execute(
            "DELETE FROM WO_Activity WHERE ParentType='movementorder' AND ParentID=%s", (mid,))
        cursor.execute("DELETE FROM WO_MovementOrders WHERE MovementOrderID=%s", (mid,))
        conn.commit()
        conn.close()
    except Exception as e:
        try:
            conn.rollback(); conn.close()
        except Exception:
            pass
        return jsonify({"error": f"Delete failed: {str(e)}"}), 500
    return jsonify({"ok": True})


@workorders_bp.route("/admin/wipe-all", methods=["POST"])
@require_roles(ROLE_ADMIN)
def api_admin_wipe_all():
    """Nuke ALL fault report + WO data. Preserves KB unless include_kb=true.
    Resets WO_Counters unless include_counters=false.
    Body: {include_kb: bool, reset_counters: bool}"""
    data = request.get_json(silent=True) or {}
    include_kb     = bool(data.get("include_kb", False))
    reset_counters = bool(data.get("reset_counters", True))
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT SPItemID FROM WO_Images WHERE SPItemID IS NOT NULL")
        sp_ids = [r[0] for r in cursor.fetchall()]
        counts = {}
        for t in ("WO_Complaints", "WO_JobOrders", "WO_MovementOrders", "WO_Images", "WO_Activity"):
            cursor.execute(f"SELECT COUNT(*) FROM {t}")
            counts[t] = int(cursor.fetchone()[0])
        log_deletion(cursor, "wipe-all", None,
                     f"Wiped ALL WO data (include_kb={include_kb})",
                     {"counts": counts, "include_kb": include_kb, "reset_counters": reset_counters},
                     False, get_current_user())
        cursor.execute("DELETE FROM WO_Images")
        cursor.execute("DELETE FROM WO_Activity")
        cursor.execute("DELETE FROM WO_JobOrderTasks")
        cursor.execute("UPDATE WO_Complaints SET JobOrderID=NULL")  # break FK first
        cursor.execute("DELETE FROM WO_JobOrders")
        cursor.execute("DELETE FROM WO_Complaints")
        cursor.execute("DELETE FROM WO_MovementOrders")
        if include_kb:
            cursor.execute("DELETE FROM WO_KB_Tickboxes")
            cursor.execute("DELETE FROM WO_KB_Entries")
        if reset_counters:
            cursor.execute("DELETE FROM WO_Counters")
        conn.commit()
        conn.close()
    except Exception as e:
        return jsonify({"error": f"Wipe failed: {str(e)}"}), 500
    _delete_sp_items(sp_ids)
    return jsonify({"ok": True, "sp_deleted": len(sp_ids)})


@workorders_bp.route("/manager/machine/<path:code>/recorded-move/<int:hid>", methods=["DELETE"])
@require_roles(*MANAGER_ROLES)
def api_delete_recorded_move(code, hid):
    """Remove a manually-recorded past move: delete its history interval and extend
    the following interval back over the gap (re-absorb). Logged + reversible."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        snap = _row_snapshot(cursor,
            "SELECT * FROM MachineLocationHistory WHERE HistoryID=%s AND MachineCode=%s AND Source='record-move'",
            (hid, str(code)))
        if not snap:
            conn.close()
            return jsonify({"error": "Recorded move not found (only manually-recorded moves can be removed here)."}), 404
        absorb_start = snap["ValidFromOle"]
        cut          = snap["ValidToOle"]
        cursor.execute("""
            SELECT TOP 1 HistoryID FROM MachineLocationHistory
            WHERE MachineCode=%s AND ValidFromOle=%s ORDER BY ValidToOle
        """, (str(code), cut))
        nxt = cursor.fetchone()
        if not nxt:
            conn.close()
            return jsonify({"error": "History changed since this move was recorded — can't safely remove it."}), 409
        cursor.execute("UPDATE MachineLocationHistory SET ValidFromOle=%s WHERE HistoryID=%s",
                       (absorb_start, nxt[0]))
        cursor.execute("DELETE FROM MachineLocationHistory WHERE HistoryID=%s", (hid,))
        log_deletion(cursor, "recorded-move", hid,
                     f"{code}: removed recorded move '{snap.get('LocationName')}'",
                     snap, True, get_current_user())
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        try:
            conn.rollback(); conn.close()
        except Exception:
            pass
        return jsonify({"error": f"Delete failed: {str(e)}"}), 500


@workorders_bp.route("/admin/deleted-log")
@require_roles(ROLE_ADMIN)
def api_deleted_log():
    """Audit trail of every delete action, newest first."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 300 LogID, EntityType, EntityKey, Summary, Reversible,
                   DeletedBy, DeletedAt, RestoredBy, RestoredAt
            FROM WO_DeletedLog ORDER BY LogID DESC
        """)
        cols = [d[0] for d in cursor.description]
        rows = [{c: v for c, v in zip(cols, r)} for r in cursor.fetchall()]
        conn.close()
        for o in rows:
            o["DeletedAt"]  = _iso(o.get("DeletedAt"))
            o["RestoredAt"] = _iso(o.get("RestoredAt"))
            o["Reversible"] = bool(o.get("Reversible")) and not o.get("RestoredAt")
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/admin/deleted-log/<int:log_id>/restore", methods=["POST"])
@require_roles(ROLE_ADMIN)
def api_deleted_log_restore(log_id):
    """One-click restore for reversible log entries (currently: recorded-move).
    Other entity types keep their full snapshot in the log for manual restore."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT EntityType, Snapshot, Reversible, RestoredAt FROM WO_DeletedLog WHERE LogID=%s",
                       (log_id,))
        row = cursor.fetchone()
        if not row:
            conn.close(); return jsonify({"error": "Log entry not found."}), 404
        etype, snap_json, reversible, restored_at = row
        if restored_at is not None:
            conn.close(); return jsonify({"error": "Already restored."}), 409
        if not reversible or not snap_json:
            conn.close(); return jsonify({"error": "This entry can't be auto-restored; its snapshot is kept for manual restore."}), 400
        snap = json.loads(snap_json)

        if etype == "recorded-move":
            code = str(snap["MachineCode"]); absorb_start = snap["ValidFromOle"]; cut = snap["ValidToOle"]
            cursor.execute("""
                SELECT TOP 1 HistoryID FROM MachineLocationHistory
                WHERE MachineCode=%s AND ValidFromOle=%s ORDER BY ValidToOle
            """, (code, absorb_start))
            nxt = cursor.fetchone()
            if not nxt:
                conn.close(); return jsonify({"error": "History changed since deletion — can't auto-restore."}), 409
            cursor.execute("UPDATE MachineLocationHistory SET ValidFromOle=%s WHERE HistoryID=%s", (cut, nxt[0]))
            cursor.execute("""
                INSERT INTO MachineLocationHistory
                    (MachineCode, LocationName, Latitude, Longitude, ValidFromOle, ValidToOle, Source)
                VALUES (%s, %s, %s, %s, %s, %s, 'record-move')
            """, (code, snap["LocationName"], snap.get("Latitude"), snap.get("Longitude"), absorb_start, cut))
        else:
            conn.close(); return jsonify({"error": f"Restore not implemented for '{etype}'."}), 400

        cursor.execute("UPDATE WO_DeletedLog SET RestoredBy=%s, RestoredAt=SYSUTCDATETIME(), Reversible=0 WHERE LogID=%s",
                       (get_current_user(), log_id))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        try:
            conn.rollback(); conn.close()
        except Exception:
            pass
        return jsonify({"error": f"Restore failed: {str(e)}"}), 500


# ── Images (SP-backed proxy) ──────────────────────────────────────────────────

@workorders_bp.route("/images", methods=["POST"])
@api_login_required
def api_image_upload():
    """Body: {parent_type, parent_id, stage, image_data_url}.
    Uploads to SP, inserts WO_Images row, returns the new image id."""
    data = request.get_json(silent=True) or {}
    parent_type = (data.get("parent_type") or "").strip().lower()
    if parent_type not in ("complaint", "joborder", "task"):
        return jsonify({"error": "parent_type must be complaint, joborder, or task."}), 400
    try:
        parent_id = int(data.get("parent_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "parent_id required."}), 400
    stage = (data.get("stage") or "before").strip().lower()
    raw, ctype = _decode_data_url(data.get("image_data_url") or "")
    if not raw:
        return jsonify({"error": "image_data_url required (data: URL)."}), 400

    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        new_id = _save_image_to_sp(
            cursor, parent_type, parent_id, stage,
            file_name=data.get("file_name") or f"{stage}.jpg",
            content_type=ctype, raw_bytes=raw, uploaded_by=user,
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": f"Upload failed: {str(e)}"}), 500


@workorders_bp.route("/images/<int:image_id>")
@api_login_required
def api_image_get(image_id):
    """Proxy the SP file body so Easy Auth gates every read.
    Falls back to legacy ImageData if SPItemID is NULL."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ContentType, SPItemID, ImageData, FileName
            FROM WO_Images WHERE ImageID = %s
        """, (image_id,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Image not found."}), 404
        ctype = row[0] or "application/octet-stream"
        sp_id = row[1]
        legacy_bytes = row[2]

        if sp_id:
            body, ctype = sp.download_bytes(sp_id)
        elif legacy_bytes:
            body = bytes(legacy_bytes)
        else:
            return jsonify({"error": "Image has no content."}), 410

        return Response(body, mimetype=ctype)
    except Exception as e:
        return jsonify({"error": f"Image read failed: {str(e)}"}), 500


@workorders_bp.route("/images/<int:image_id>", methods=["DELETE"])
@require_roles(*MANAGER_ROLES)
def api_image_delete(image_id):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT SPItemID FROM WO_Images WHERE ImageID=%s", (image_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Image not found."}), 404
        sp_id = row[0]
        cursor.execute("DELETE FROM WO_Images WHERE ImageID=%s", (image_id,))
        conn.commit()
        conn.close()
        if sp_id:
            try:
                sp.delete_item(sp_id)
            except Exception as e:
                # Log but don't fail the deletion of the DB row.
                print(f"[images/delete] SP delete failed for {sp_id}: {e}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Delete failed: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────────────────────
# PORTED FROM V1 (workorders.py.bak-2026-06-03) — verbatim, unchanged.
# These routes still use text Status/Priority (V1 columns kept alongside V2
# TINYINT columns post-migration). To be modernized in a future pass.
# ─────────────────────────────────────────────────────────────────────────────


def _sync_topup(cursor, machine_code):
    """Replicates the vending dashboard's log_topup behaviour. Returns a status string."""
    now_ole = to_ole_date(datetime.utcnow())
    cursor.execute(
        "SELECT LastTopupTimestamp FROM MachineLookup WHERE MachineCode = %s",
        (machine_code,),
    )
    row = cursor.fetchone()
    if not row:
        return "Machine code not found in vending data — topup not synced."
    current_ole = row[0]
    # NB: '1%%' — pymssql %-formats the query before sending it, so a literal
    # '%' MUST be doubled in any statement that also passes parameters.
    # (This was '1%' and raised TypeError on every call, i.e. delivery
    #  completion has never actually succeeded. Do not "simplify" it back.)
    if current_ole is not None:
        cursor.execute(f"""
            SELECT COUNT(DISTINCT CAST([Date Time] AS FLOAT)) FROM [MasterData Table]
            WHERE CAST([Machine Code] AS NVARCHAR(50)) = %s
              AND LEN(CAST([Event Code] AS NVARCHAR(20))) = 6
              AND CAST([Event Code] AS NVARCHAR(20)) LIKE '1%%'
              AND CAST([Date Time] AS FLOAT) >= {float(current_ole)}
        """, (machine_code,))
    else:
        cursor.execute("""
            SELECT COUNT(DISTINCT CAST([Date Time] AS FLOAT)) FROM [MasterData Table]
            WHERE CAST([Machine Code] AS NVARCHAR(50)) = %s
              AND LEN(CAST([Event Code] AS NVARCHAR(20))) = 6
              AND CAST([Event Code] AS NVARCHAR(20)) LIKE '1%%'
        """, (machine_code,))
    vends_since = int(cursor.fetchone()[0])
    cursor.execute(f"""
        UPDATE MachineLookup
        SET PreviousTopupTimestamp = LastTopupTimestamp,
            LastTopupTimestamp     = {now_ole},
            CountBeforeLastTopup   = %s
        WHERE MachineCode = %s
    """, (vends_since, machine_code))
    return f"Topup logged ({vends_since} vends since previous topup)."


# ── Movement Orders (warehouse ↔ location ↔ location) ────────────────────────

WAREHOUSE_LABEL = "WAREHOUSE"


def _machine_location(cursor, machine_code):
    """Return (name, lat, lon, is_active) for a machine, or (None,...) if not found."""
    cursor.execute("""
        SELECT MachineName, Latitude, Longitude, ISNULL(IsActive, 1)
        FROM MachineLookup WHERE MachineCode = %s
    """, (machine_code,))
    r = cursor.fetchone()
    if not r:
        return (None, None, None, None)
    return (r[0], r[1], r[2], int(r[3]))


@workorders_bp.route("/movementorders", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_movement_create():
    """Body:
        movement_type: 'deploy' | 'relocate' | 'retrieve'
        machine_code:  required
        to_location:   name (NULL for retrieve — defaults to WAREHOUSE)
        to_lat, to_lon: required for deploy + relocate
        notes:         optional
        assigned_to:   driver email
        reason_for_retrieval: optional (retrieve only)
    """
    data = request.get_json(silent=True) or {}
    mtype = (data.get("movement_type") or "").strip().lower()
    if mtype not in MOVEMENT_TYPES:
        return jsonify({"error": "movement_type must be deploy / relocate / retrieve."}), 400

    machine_code = (data.get("machine_code") or "").strip()
    if not machine_code:
        return jsonify({"error": "machine_code required."}), 400

    to_lat = data.get("to_lat"); to_lon = data.get("to_lon")
    to_location = (data.get("to_location") or "").strip() or None
    if mtype == "retrieve":
        to_location = WAREHOUSE_LABEL
        to_lat = None; to_lon = None
    else:
        try:
            to_lat = float(to_lat) if to_lat not in (None, "", "null") else None
            to_lon = float(to_lon) if to_lon not in (None, "", "null") else None
        except (TypeError, ValueError):
            return jsonify({"error": "to_lat and to_lon must be numbers."}), 400
        if to_lat is None or to_lon is None or not to_location:
            return jsonify({"error": "to_location, to_lat, to_lon required for deploy / relocate."}), 400

    user = get_current_user()
    notes              = (data.get("notes") or "").strip() or None
    assigned           = (data.get("assigned_to") or "").strip().lower() or None
    reason             = (data.get("reason_for_retrieval") or "").strip() or None

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cur_name, cur_lat, cur_lon, cur_active = _machine_location(cursor, machine_code)

        # Sanity:
        if mtype == "deploy" and cur_active == 1:
            conn.close()
            return jsonify({"error": "Machine is currently active. Use relocate instead."}), 400
        if mtype in ("relocate", "retrieve") and cur_active != 1:
            conn.close()
            return jsonify({"error": "Machine is not currently active. Use deploy."}), 400

        from_loc = WAREHOUSE_LABEL if (mtype == "deploy" or cur_active != 1) else cur_name
        display_id = allocate_display_id(cursor, "MVO")

        cursor.execute("""
            INSERT INTO WO_MovementOrders
                (MovementType, MachineCode,
                 FromLocation, FromLat, FromLon,
                 ToLocation,   ToLat,   ToLon,
                 Notes, AssignedTo, ReasonForRetrieval,
                 DisplayID, CreatedBy)
            OUTPUT INSERTED.MovementOrderID
            VALUES (%s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s)
        """, (mtype, machine_code,
              from_loc, cur_lat, cur_lon,
              to_location, to_lat, to_lon,
              notes, assigned, reason,
              display_id, user))
        new_id = int(cursor.fetchone()[0])
        _log_activity(cursor, "movementorder", new_id, "created",
                      f"{mtype} {machine_code}: {from_loc} → {to_location}", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id, "display_id": display_id})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/movementorders")
@api_login_required
def api_movement_list():
    """?scope=mine|all  ?status=scheduled|in_progress|completed|all  ?machine_code=…"""
    scope         = (request.args.get("scope") or "mine").strip().lower()
    status_filter = (request.args.get("status") or "all").strip().lower()
    machine_code  = (request.args.get("machine_code") or "").strip() or None
    user          = get_current_user()
    if get_role(user) not in MANAGER_ROLES:
        scope = "mine"

    where, params = [], []
    if scope == "mine":
        where.append("AssignedTo = %s"); params.append(user)
    rev = {v: k for k, v in MOVEMENT_STATUS.items()}
    if status_filter in rev:
        where.append("StatusCode = %s"); params.append(rev[status_filter])
    if machine_code:
        where.append("MachineCode = %s"); params.append(machine_code)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(f"""
            SELECT TOP 200
                MovementOrderID, DisplayID, MovementType, MachineCode,
                FromLocation, FromLat, FromLon,
                ToLocation,   ToLat,   ToLon,
                StatusCode, AssignedTo, CreatedBy, CreatedAt,
                CompletedBy, CompletedAt, Notes, ReasonForRetrieval
            FROM WO_MovementOrders {where_sql}
            ORDER BY CreatedAt DESC, MovementOrderID DESC
        """, tuple(params))
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "id": r[0], "display_id": r[1],
            "movement_type": r[2], "machine_code": r[3],
            "from_location": r[4], "from_lat": r[5], "from_lon": r[6],
            "to_location":   r[7], "to_lat":   r[8], "to_lon":   r[9],
            "status_code":   int(r[10]),
            "status_label":  _label(MOVEMENT_STATUS, r[10]),
            "assigned_to": r[11], "created_by": r[12], "created_at": _iso(r[13]),
            "completed_by": r[14], "completed_at": _iso(r[15]),
            "notes": r[16], "reason_for_retrieval": r[17],
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/movementorders/<int:mid>/complete", methods=["POST"])
@require_roles(*OPERATOR_ROLES)
def api_movement_complete(mid):
    """Mark the movement as completed AND apply the location/active-flag change
    to MachineLookup atomically."""
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT MovementType, MachineCode, ToLocation, ToLat, ToLon,
                   AssignedTo, StatusCode, DisplayID, ReasonForRetrieval
            FROM WO_MovementOrders WHERE MovementOrderID = %s
        """, (mid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Movement order not found."}), 404
        (mtype, machine_code, to_loc, to_lat, to_lon,
         assigned, status_code, display, reason) = row
        if int(status_code) == 2:
            conn.close()
            return jsonify({"error": "Already completed."}), 400
        if get_role(user) not in MANAGER_ROLES and (assigned or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not assigned to you."}), 403

        # Apply MachineLookup change
        if mtype == "deploy":
            cursor.execute("""
                UPDATE MachineLookup
                SET MachineName = %s, Latitude = %s, Longitude = %s,
                    IsActive = 1, DecommissionedAt = NULL, DecommissionReason = NULL
                WHERE MachineCode = %s
            """, (to_loc, to_lat, to_lon, machine_code))
        elif mtype == "relocate":
            cursor.execute("""
                UPDATE MachineLookup
                SET MachineName = %s, Latitude = %s, Longitude = %s
                WHERE MachineCode = %s
            """, (to_loc, to_lat, to_lon, machine_code))
        elif mtype == "retrieve":
            cursor.execute("""
                UPDATE MachineLookup
                SET IsActive = 0,
                    DecommissionedAt = SYSUTCDATETIME(),
                    DecommissionReason = %s
                WHERE MachineCode = %s
            """, (reason, machine_code))

        cursor.execute("""
            UPDATE WO_MovementOrders
            SET StatusCode = 2, CompletedBy = %s, CompletedAt = SYSUTCDATETIME()
            WHERE MovementOrderID = %s
        """, (user, mid))

        # ── Effective-dated location history (sharp cutoff at completion) ──────────
        # Cutoff = now in SGT (UTC+8) as an OLE float, matching vend [Date Time].
        # Close the machine's current OPEN interval, then (deploy/relocate) open a
        # new one. Best-effort: never block the movement completion.
        try:
            cut_ole_sql = "CAST(CONVERT(datetime, DATEADD(HOUR, 8, SYSUTCDATETIME())) AS FLOAT) + 2.0"
            cursor.execute(f"""
                UPDATE MachineLocationHistory
                SET ValidToOle = {cut_ole_sql}
                WHERE MachineCode = %s AND ValidToOle IS NULL
            """, (str(machine_code),))
            if mtype in ("deploy", "relocate"):
                cursor.execute(f"""
                    INSERT INTO MachineLocationHistory
                        (MachineCode, LocationName, Latitude, Longitude, ValidFromOle, ValidToOle, Source, MovementOrderID)
                    VALUES (%s, %s, %s, %s, {cut_ole_sql}, NULL, 'movement', %s)
                """, (str(machine_code), to_loc, to_lat, to_lon, mid))
            # retrieve: leave the interval closed (machine decommissioned, no open row)
        except Exception as he:
            print(f"[movement_complete] history write skipped for {machine_code}: {he}")

        _log_activity(cursor, "movementorder", mid, "completed",
                      f"{mtype} applied to {machine_code}", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500



# ── Dispatch stops (2026-08-11) ───────────────────────────────────────────────
# A "stop" is not a table. It is one machine, one driver, one date, and one or
# more of the orders that already exist: a delivery order for a top-up and/or a
# job order for service. Modelling it as a new table would mean rebuilding the
# operator's list, the Work Order sheet, the signed PDF and the top-up sync to
# read from it. Instead the two new nullable columns (ScheduledDate, RouteSeq)
# turn the existing rows into a dated, ordered round.

STOP_KINDS = ("topup", "service")


def _has_col(cursor, table, col):
    cursor.execute(
        "SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME=%s AND COLUMN_NAME=%s",
        (table, col),
    )
    return cursor.fetchone() is not None


def _sgt_today():
    """KNM runs on SGT. Server clock is UTC, so a round planned at 07:00 local
    must not be filed against yesterday."""
    return (datetime.utcnow() + timedelta(hours=8)).date().isoformat()


def _parse_date(v):
    if not v:
        return None
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date().isoformat()
    except Exception:
        return None


def _next_route_seq(cursor, assigned, day):
    """Next free position in this driver's round. Ties are fine — the board
    sorts by seq then id, so a collision degrades to insertion order."""
    if not assigned or not day:
        return None
    top = 0
    for tbl, idc in (("WO_DeliveryOrders", "DeliveryOrderID"),
                     ("WO_JobOrders", "JobOrderID")):
        try:
            cursor.execute(
                f"SELECT MAX(RouteSeq) FROM {tbl} WHERE AssignedTo=%s AND ScheduledDate=%s",
                (assigned, day),
            )
            r = cursor.fetchone()
            if r and r[0] is not None:
                top = max(top, int(r[0]))
        except Exception:
            return None            # columns not migrated yet
    return top + 1


@workorders_bp.route("/stops", methods=["POST"])
@require_roles(*DISPATCH_ROLES)
def api_stop_create():
    """Create the work a dispatcher ticked for one machine, one driver, one day.

    Body:
        machine_code:  str   (required — must exist in MachineLookup)
        assigned_to:   str   (email; optional, a stop can be planned unassigned)
        scheduled_date:'YYYY-MM-DD' (defaults to today, SGT)
        needs_service: bool — renders as a service request at the top of the sheet
        note:          str   — becomes the job order's Diagnosis so the driver
                               reads it on the Work Order sheet, and the
                               delivery order's Notes.
        priority:      'low' | 'normal' | 'high'

    One stop is ONE WO_DeliveryOrders row. Service is a flag and a note on it,
    not a second order — a job order now only exists for the Tech Support path
    (a customer complaint triaged into a WO).
    """
    data = request.get_json(silent=True) or {}
    code = (data.get("machine_code") or "").strip()
    assigned = (data.get("assigned_to") or "").strip().lower() or None
    note = (data.get("note") or "").strip() or None
    priority = (data.get("priority") or "normal").strip().lower()
    if priority not in ("low", "normal", "high"):
        priority = "normal"
    # "service" is no longer a separate order. It is a flag on this one.
    kinds = [k for k in (data.get("kinds") or []) if k in STOP_KINDS]
    needs_service = bool(data.get("needs_service")) or ("service" in kinds)
    _raw_day = data.get("scheduled_date")
    day = _parse_date(_raw_day) or (None if _raw_day else _sgt_today())
    if day is None:
        # Was: `or _sgt_today()`, which silently relabelled the stop with today's
        # date and still returned success -- the stop then sat on a day nobody
        # was looking at. The PATCH route below has always rejected this.
        return jsonify({"error": "scheduled_date must be YYYY-MM-DD."}), 400
    user = get_current_user()

    if not code:
        return jsonify({"error": "Pick a machine."}), 400
    # Every stop is a visit to the machine; service is an attribute of it.

    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT TOP 1 MachineName, ISNULL(IsActive,1) FROM MachineLookup WHERE MachineCode = %s",
            (code,),
        )
        m = cursor.fetchone()
        if not m:
            conn.close()
            return jsonify({"error": f"Machine {code} is not in the machine registry."}), 404
        name = m[0] or code
        if not int(m[1] or 0):
            conn.close()
            return jsonify({"error": f"{name} is decommissioned; it cannot be scheduled."}), 400

        # Each ALTER in init_workorders_db is swallowed independently, so one
        # table can be migrated while another is not. Probe per table or a
        # half-migrated schema turns into a 500 that loses the whole stop.
        dated_do = _has_col(cursor, "WO_DeliveryOrders", "ScheduledDate")
        seq = _next_route_seq(cursor, assigned, day) if dated_do else None
        created = []

        # ── ONE row per stop ────────────────────────────────────────────────
        # The driver's sheet links one open delivery order per machine per day,
        # so a second would be a ghost carrying a note nobody reads.
        _dupq = ("SELECT TOP 1 DeliveryOrderID, AssignedTo FROM WO_DeliveryOrders "
                 "WHERE MachineCode = %s AND Status <> 'completed'")
        _dupp = [code]
        if dated_do:
            _dupq += " AND ISNULL(CONVERT(VARCHAR(10), ScheduledDate, 23), %s) = %s"
            _dupp += [day, day]
        cursor.execute(_dupq + " ORDER BY CreatedAt", tuple(_dupp))
        dup = cursor.fetchone()
        if dup:
            conn.close()
            return jsonify({
                "error": f"{name} already has an open stop on {day}"
                         + (f" assigned to {dup[1]}" if dup[1] else " (unassigned)")
                         + ". Edit it instead of adding a second.",
                "existing_delivery_order_id": int(dup[0]),
            }), 409

        cols = ["MachineName", "MachineCode", "AssignedTo", "Priority", "CreatedBy"]
        vals = [name, code, assigned, priority, user]
        # One note box on the composer, routed by the tick: a service request
        # renders in the amber box at the top of the sheet, a plain note in the
        # delivery section. Never both, so nothing is duplicated.
        if needs_service:
            cols += ["ServiceNote"]; vals += [note]
        else:
            cols += ["Notes"];       vals += [note]
        if _has_col(cursor, "WO_DeliveryOrders", "NeedsService"):
            cols += ["NeedsService"]; vals += [1 if needs_service else 0]
        if dated_do:
            cols += ["ScheduledDate", "RouteSeq"]; vals += [day, seq]
        cursor.execute(
            f"INSERT INTO WO_DeliveryOrders ({', '.join(cols)}) "
            f"OUTPUT INSERTED.DeliveryOrderID VALUES ({', '.join(['%s'] * len(cols))})",
            tuple(vals),
        )
        did = int(cursor.fetchone()[0])
        _log_activity(cursor, "deliveryorder", did, "created",
                      ("Service stop" if needs_service else "Top-up")
                      + f" scheduled {day} for {assigned or '(unassigned)'} at {name}", user)
        created.append({"kind": "service" if needs_service else "topup", "id": did})

        conn.commit()
        conn.close()
        return jsonify({"ok": True, "created": created, "machine_name": name,
                        "scheduled_date": day, "route_seq": seq,
                        "dated": dated_do, "needs_service": needs_service})
    except Exception as e:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500



@workorders_bp.route("/stops/<kind>/<int:sid>", methods=["PATCH"])
@require_roles(*DISPATCH_ROLES)
def api_stop_update(kind, sid):
    """Edit a stop without destroying and recreating it.

    Body (all optional): note, assigned_to (null to unassign),
    scheduled_date (null to undate), priority.

    NOTE the asymmetry: here `scheduled_date: null` UNDATES the stop, because
    the presence of the key is the signal. On the /assign routes the same field
    is truthiness-tested, so null is a no-op there. Same name, opposite meaning.

    Exists because the alternative — delete and re-add — is destructive, is
    refused once a driver has started, and would lose the order's history.
    """
    data = request.get_json(silent=True) or {}
    user = get_current_user()
    note = data.get("note")
    pri = (data.get("priority") or "").strip().lower() or None
    if pri and pri not in ("low", "normal", "high"):
        return jsonify({"error": "Priority must be low, normal or high."}), 400
    day = _parse_date(data.get("scheduled_date")) if data.get("scheduled_date") else None
    # A date we could not parse must NOT fall through to "undate this stop" —
    # a typo would quietly drop it off every dated board.
    if data.get("scheduled_date") and day is None:
        return jsonify({"error": "scheduled_date must be YYYY-MM-DD."}), 400
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        sets, params, log = [], [], []
        if "assigned_to" in data:
            a = (data.get("assigned_to") or "").strip().lower() or None
            sets.append("AssignedTo = %s"); params.append(a)
            log.append(f"assigned to {a or '(unassigned)'}")
        if "scheduled_date" in data:
            if _has_col(cursor, "WO_DeliveryOrders" if kind == "topup" else "WO_JobOrders", "ScheduledDate"):
                sets.append("ScheduledDate = %s"); params.append(day)
                log.append(f"date {day or '(none)'}")

        if kind == "topup":
            cursor.execute("SELECT Status, AssignedTo FROM WO_DeliveryOrders WHERE DeliveryOrderID=%s", (sid,))
            r = cursor.fetchone()
            if not r:
                conn.close(); return jsonify({"error": "Stop not found."}), 404
            if (r[0] or "").lower() == "completed":
                conn.close(); return jsonify({"error": "This top-up is already completed."}), 400
            if "needs_service" in data:
                ns = 1 if data.get("needs_service") else 0
                sets.append("NeedsService = %s"); params.append(ns)
                log.append("service " + ("required" if ns else "not required"))
                # Carry the existing text across with the flag. Without this the
                # note stays in the column the new mode does not read, and it
                # disappears from the sheet, the PDF and the board with no error.
                if "note" not in data or data.get("note") is None:
                    if ns:
                        sets.append("ServiceNote = COALESCE(NULLIF(LTRIM(RTRIM(ServiceNote)),''), Notes)")
                        sets.append("Notes = NULL")
                    else:
                        sets.append("Notes = COALESCE(NULLIF(LTRIM(RTRIM(Notes)),''), ServiceNote)")
                        sets.append("ServiceNote = NULL")
            if note is not None:
                # Written to whichever field the stop's current mode uses, so a
                # service request never lands in the delivery-notes line and a
                # top-up note never surfaces as an amber service request.
                _ns = data.get("needs_service")
                if _ns is None:
                    cursor.execute("SELECT ISNULL(NeedsService,0) FROM WO_DeliveryOrders WHERE DeliveryOrderID=%s", (sid,))
                    _r2 = cursor.fetchone(); _ns = bool(_r2 and _r2[0])
                col   = "ServiceNote" if _ns else "Notes"
                other = "Notes" if _ns else "ServiceNote"
                sets.append(f"{col} = %s"); params.append((note or "").strip() or None)
                sets.append(f"{other} = NULL")     # never leave a stale copy behind
                log.append("note")
            if pri:
                sets.append("Priority = %s"); params.append(pri); log.append(f"priority {pri}")
            table, idcol, ptype = "WO_DeliveryOrders", "DeliveryOrderID", "deliveryorder"
        elif kind == "service":
            cursor.execute("SELECT StatusCode, ComplaintID, AssignedTo FROM WO_JobOrders WHERE JobOrderID=%s", (sid,))
            r = cursor.fetchone()
            if not r:
                conn.close(); return jsonify({"error": "Stop not found."}), 404
            if int(r[0] or 0) in (2, 3):
                conn.close(); return jsonify({"error": "This job order is in review or closed; it cannot be edited here."}), 400
            if note is not None and r[1]:
                conn.close(); return jsonify({"error": "This job order came from a customer complaint; edit its diagnosis in Tech Support."}), 400
            if note is not None:
                # Diagnosis is what the driver reads on the Work Order sheet.
                sets.append("Diagnosis = %s"); params.append((note or "").strip() or None)
                sets.append("Notes = %s");     params.append((note or "").strip() or None)
                log.append("note")
            if pri:
                sets.append("Priority = %s");     params.append(pri)
                sets.append("PriorityCode = %s"); params.append({"low": 0, "normal": 1, "high": 2}[pri])
                log.append(f"priority {pri}")
            table, idcol, ptype = "WO_JobOrders", "JobOrderID", "joborder"
        else:
            conn.close(); return jsonify({"error": "Unknown stop type."}), 400

        if not sets:
            conn.close(); return jsonify({"ok": True, "changed": False})
        params.append(sid)
        cursor.execute(f"UPDATE {table} SET {', '.join(sets)} WHERE {idcol} = %s", tuple(params))
        _log_activity(cursor, ptype, sid, "edited", "Dispatch board: " + ", ".join(log), user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "changed": True})
    except Exception as e:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/stops/<kind>/<int:sid>", methods=["DELETE"])
@require_roles(*DISPATCH_ROLES)
def api_stop_delete(kind, sid):
    """Remove a stop the dispatcher added by mistake.

    Only ever cancels work that is still untouched — a delivery order that is
    still 'open', or a job order still at StatusCode 0. Anything a driver has
    already started or submitted is refused, so this can never erase field work.
    """
    user = get_current_user()
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        sp_ids = []
        if kind == "topup":
            cursor.execute(
                "SELECT Status, MachineName FROM WO_DeliveryOrders WHERE DeliveryOrderID=%s", (sid,))
            r = cursor.fetchone()
            if not r:
                conn.close(); return jsonify({"error": "Stop not found."}), 404
            if (r[0] or "").lower() != "open":
                conn.close(); return jsonify({"error": "This top-up has already been worked; it cannot be removed."}), 400
            # Dispatch can pull a stop after the shift has started — only a
            # COMPLETED location is protected (the Status check above). If a
            # driver has it open on an unsubmitted sheet, detach the link first
            # so finalize does not later update a row that no longer exists.
            cursor.execute("""
                UPDATE WO_VisitSessions SET LinkedDeliveryOrderID = NULL
                WHERE LinkedDeliveryOrderID = %s
                  AND Status NOT IN ('signed', 'pending_email_signature')
            """, (sid,))
            snap = _row_snapshot(cursor, "SELECT * FROM WO_DeliveryOrders WHERE DeliveryOrderID=%s", (sid,))
            cursor.execute("DELETE FROM WO_DeliveryOrderLines WHERE DeliveryOrderID=%s", (sid,))
            cursor.execute("DELETE FROM WO_Activity WHERE ParentType='deliveryorder' AND ParentID=%s", (sid,))
            cursor.execute("DELETE FROM WO_DeliveryOrders WHERE DeliveryOrderID=%s", (sid,))
            log_deletion(cursor, "deliveryorder", sid,
                         f"Top-up at {(snap or {}).get('MachineName') or sid}",
                         snap, False, user)
        elif kind == "service":
            cursor.execute("SELECT StatusCode, ComplaintID FROM WO_JobOrders WHERE JobOrderID=%s", (sid,))
            r = cursor.fetchone()
            if not r:
                conn.close(); return jsonify({"error": "Stop not found."}), 404
            if int(r[0] or 0) != 0:
                conn.close(); return jsonify({"error": "This service job has been started or submitted; it cannot be removed."}), 400
            if r[1]:
                conn.close(); return jsonify({"error": "This job order came from a customer complaint. Close it in Tech Support instead."}), 400
            # Same rule for service: removable until the location is completed.
            # Unhook it from any unsubmitted sheet rather than refusing.
            cursor.execute("""
                DELETE FROM WO_VisitSession_JobOrders
                WHERE JobOrderID = %s AND VisitID IN (
                    SELECT VisitID FROM WO_VisitSessions
                    WHERE Status NOT IN ('signed', 'pending_email_signature'))
            """, (sid,))
            # Full cascade + a restorable audit record, exactly like the admin
            # delete route. A bare DELETE would orphan tasks, images, the
            # SharePoint files behind them, and the activity trail.
            snap = _row_snapshot(cursor, "SELECT * FROM WO_JobOrders WHERE JobOrderID=%s", (sid,))
            sp_ids = _cascade_delete_joborder(cursor, sid)
            log_deletion(cursor, "joborder", sid,
                         f"Work order {(snap or {}).get('DisplayID') or ('#' + str(sid))}",
                         snap, False, user)
        else:
            conn.close(); return jsonify({"error": "Unknown stop type."}), 400
        conn.commit()
        conn.close()
        if sp_ids:
            _delete_sp_items(sp_ids)
        return jsonify({"ok": True})
    except Exception as e:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ══════════════════════════════════════════════════════════════════════════════
# SALES SCHEDULE  (2026-08-11)
# ══════════════════════════════════════════════════════════════════════════════
# Sales keys next week's top-ups and deliveries on a Mon–Sat calendar. What she
# creates is an ORDINARY unassigned WO_DeliveryOrders row with ScheduledDate
# set — the exact thing dispatch already drags onto a driver card. There is no
# separate "request" queue to convert, so nothing can fall between the two
# screens.
#
# A repeat is MATERIALISED, not stored as a rule: one real row per occurrence,
# all carrying the same SeriesID. Every existing read path (driver list, Plan
# board, Work Order sheet, PDF, top-up sync) therefore needs zero knowledge of
# recurrence. The cost is a horizon — occurrences beyond MAX_OCCURRENCES do not
# exist until someone extends them — which is the right trade for a sales team
# that re-keys the round every quarter anyway.

# Sales plans a quarter at a time. The horizon is expressed in WEEKS, not in a
# raw occurrence count, so "12 weeks ahead" means the same thing whichever
# frequency she picks — and a fat-fingered occurrence count cannot materialise
# two years of rows onto the board.
SCHEDULE_HORIZON_WEEKS = 12
MAX_OCCURRENCES = SCHEDULE_HORIZON_WEEKS      # weekly is the densest rule
SCHEDULE_STEP_DAYS = {"weekly": 7, "fortnightly": 14, "fourweekly": 28}
SCHEDULE_RULES = ("none",) + tuple(SCHEDULE_STEP_DAYS)


def _iso_date(d):
    # NOT _iso — that name is already a module-level datetime→"...Z" formatter
    # (line 412) used by ~48 endpoints. Redefining it here would rebind it for
    # all of them and 500 every response carrying a NULL timestamp.
    return d.isoformat()


def _d(s):
    return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()


def _series_dates(start, rule, count):
    """Expand a repeat into concrete dates. `start` is always dates[0].

    Every rule is a fixed stride in whole weeks, so a round can never drift off
    the weekday sales chose — a machine stays in the same column of her calendar
    and on the same day of the driver's week, forever.

    Anything past SCHEDULE_HORIZON_WEEKS from the start date is dropped and
    reported, so the caller can tell her the series was shortened rather than
    letting it silently stop.
    """
    if rule == "none" or count <= 1:
        return [start], []
    step = SCHEDULE_STEP_DAYS.get(rule)
    if not step:
        return [start], []
    horizon = start + timedelta(days=SCHEDULE_HORIZON_WEEKS * 7)
    out, skipped = [start], []
    for i in range(1, count):
        d = start + timedelta(days=step * i)
        if d > horizon:
            skipped.append({"date": _iso_date(d),
                            "reason": f"beyond the {SCHEDULE_HORIZON_WEEKS}-week planning horizon"})
            continue
        out.append(d)
    return out, skipped


def _sched_cols(cursor):
    """Which optional columns this database actually has.

    Each ALTER in init_workorders_db is swallowed independently, so the schema
    can be half-migrated. Probing keeps a partial migration degraded rather than
    a 500 that loses the whole week's keying.
    """
    return {c: _has_col(cursor, "WO_DeliveryOrders", c) for c in
            ("ScheduledDate", "RouteSeq", "NeedsService", "ServiceNote",
             "SeriesID", "SeriesRule", "RequestedBy")}


def _caller_is_sales_only():
    """True when the active role is sales and nothing more.

    Sales is allowed to cancel only what is still unassigned and untouched. A
    field manager hitting the same route keeps full dispatch authority.
    """
    return get_role(get_current_user()) == ROLE_SALES


@workorders_bp.route("/schedule", methods=["GET"])
@require_roles(*SALES_ROLES)
def api_schedule_list():
    """Every dated, still-open stop between ?from= and ?to= (inclusive).

    Feeds both the sales calendar and the dispatch rail, so the two screens can
    never disagree about what was asked for.
    """
    frm = _parse_date(request.args.get("from"))
    to = _parse_date(request.args.get("to"))
    if not frm or not to:
        return jsonify({"error": "from and to are required (YYYY-MM-DD)."}), 400
    if _d(to) < _d(frm):
        frm, to = to, frm
    if (_d(to) - _d(frm)).days > 400:
        return jsonify({"error": "Range too wide — ask for a year or less."}), 400
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        have = _sched_cols(cursor)
        if not have["ScheduledDate"]:
            return jsonify({"stops": [], "dated": False,
                            "warning": "Scheduling columns are not migrated yet."})
        sel = ["DeliveryOrderID", "MachineCode", "MachineName", "AssignedTo",
               "Priority", "Status", "Notes", "CreatedBy", "CreatedAt",
               "CONVERT(VARCHAR(10), ScheduledDate, 23)"]
        opt = [c for c in ("NeedsService", "ServiceNote", "SeriesID",
                           "SeriesRule", "RequestedBy") if have[c]]
        cursor.execute(
            f"SELECT {', '.join(sel + opt)} FROM WO_DeliveryOrders "
            "WHERE Status <> 'completed' AND ScheduledDate >= %s AND ScheduledDate <= %s "
            "ORDER BY ScheduledDate, MachineName", (frm, to))
        rows = cursor.fetchall()
        conn.close()
        out = []
        for r in rows:
            ext = dict(zip(opt, r[len(sel):]))
            note = ext.get("ServiceNote") or r[6]
            out.append({
                "id": int(r[0]), "machine_code": r[1], "machine_name": r[2] or r[1],
                "assigned_to": r[3], "priority": (r[4] or "normal"),
                "status": (r[5] or "open"), "note": note,
                "created_by": ext.get("RequestedBy") or r[7],
                "date": r[9],
                "needs_service": bool(ext.get("NeedsService")),
                "series_id": ext.get("SeriesID"),
                "series_rule": ext.get("SeriesRule"),
            })
        return jsonify({"stops": out, "dated": True})
    except Exception as e:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/schedule/stops", methods=["POST"])
@require_roles(*SALES_ROLES)
def api_schedule_create():
    """Key one stop, optionally repeating.

    Body:
        machine_code:   str (required)
        scheduled_date: 'YYYY-MM-DD' (required — the calendar always supplies it)
        note:           str
        needs_service:  bool
        priority:       'low' | 'normal' | 'high'
        repeat:         'none' | 'weekly' | 'fortnightly' | 'fourweekly'
        occurrences:    int, 1..MAX_OCCURRENCES (total, including the first);
                        also clipped to SCHEDULE_HORIZON_WEEKS from the start

    Stops are created UNASSIGNED — naming a driver is dispatch's call, and the
    /schedule routes never accept an assigned_to even from a field manager, so
    there is one place where assignment happens.

    A date that already holds an open stop for that machine is SKIPPED, not
    fatal: a 12-week series must not be abandoned because week 3 was already
    keyed. The response reports every skip so the calendar can say so.
    """
    data = request.get_json(silent=True) or {}
    code = (data.get("machine_code") or "").strip()
    note = (data.get("note") or "").strip() or None
    needs_service = bool(data.get("needs_service"))
    priority = (data.get("priority") or "normal").strip().lower()
    if priority not in ("low", "normal", "high"):
        priority = "normal"
    rule = (data.get("repeat") or "none").strip().lower()
    if rule not in SCHEDULE_RULES:
        return jsonify({"error": f"Unknown repeat '{rule}'."}), 400
    try:
        occ = int(data.get("occurrences") or 1)
    except (TypeError, ValueError):
        occ = 1
    if rule == "none":
        occ = 1
    occ = max(1, min(occ, MAX_OCCURRENCES))
    day = _parse_date(data.get("scheduled_date"))
    user = get_current_user()

    if not code:
        return jsonify({"error": "Pick a machine."}), 400
    if not day:
        return jsonify({"error": "Pick a date."}), 400
    # A stop keyed into last week can never be driven; it would sit undone on
    # the board forever. Today is allowed — sales does phone in same-day asks.
    if day < _sgt_today():
        return jsonify({"error": "That date has already passed."}), 400

    dates, skipped = _series_dates(_d(day), rule, occ)
    series_id = uuid.uuid4().hex[:32] if (rule != "none" and len(dates) > 1) else None

    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT TOP 1 MachineName, ISNULL(IsActive,1) FROM MachineLookup WHERE MachineCode = %s",
            (code,))
        m = cursor.fetchone()
        if not m:
            conn.close()
            return jsonify({"error": f"Machine {code} is not in the machine registry."}), 404
        name = m[0] or code
        if not int(m[1] or 0):
            conn.close()
            return jsonify({"error": f"{name} is decommissioned; it cannot be scheduled."}), 400

        have = _sched_cols(cursor)
        if not have["ScheduledDate"]:
            conn.close()
            return jsonify({"error": "Scheduling columns are not migrated on this database yet."}), 503

        created = []
        for dt in dates:
            ds = _iso_date(dt)
            # ISNULL matches api_stop_create (workorders.py:3274) and is
            # deliberate: an open stop with NO date shows on the driver's sheet
            # on EVERY day, and WO_VisitSessions carries a single
            # LinkedDeliveryOrderID — so an undated carry-over really does
            # collide with any date sales picks, and one of the two rows could
            # never be closed. It is reported, not swallowed: `skipped` comes
            # back in the payload and the calendar toasts "N date(s) skipped
            # (already booked)". A whole series can therefore come back empty —
            # that is the loud symptom of an undated ghost row on that machine,
            # and the fix is to date or close the ghost, not to book over it.
            cursor.execute(
                "SELECT TOP 1 DeliveryOrderID, AssignedTo FROM WO_DeliveryOrders "
                "WHERE MachineCode = %s AND Status <> 'completed' "
                "AND ISNULL(CONVERT(VARCHAR(10), ScheduledDate, 23), %s) = %s",
                (code, ds, ds))
            dup = cursor.fetchone()
            if dup:
                skipped.append({"date": ds,
                                "reason": "already has an open stop"
                                          + (f" for {dup[1]}" if dup[1] else " (unassigned)"),
                                "existing_id": int(dup[0])})
                continue
            cols = ["MachineName", "MachineCode", "AssignedTo", "Priority",
                    "CreatedBy", "ScheduledDate"]
            vals = [name, code, None, priority, user, ds]
            # One note box, routed by the service tick — exactly as the Plan
            # board does it, so a stop sales keyed and a stop dispatch keyed are
            # indistinguishable downstream.
            if needs_service and have["ServiceNote"]:
                cols += ["ServiceNote"]; vals += [note]
            else:
                cols += ["Notes"];       vals += [note]
            if have["NeedsService"]:
                cols += ["NeedsService"]; vals += [1 if needs_service else 0]
            if series_id and have["SeriesID"]:
                cols += ["SeriesID"]; vals += [series_id]
            if have["SeriesRule"]:
                cols += ["SeriesRule"]; vals += [rule]
            if have["RequestedBy"]:
                cols += ["RequestedBy"]; vals += [user]
            cursor.execute(
                f"INSERT INTO WO_DeliveryOrders ({', '.join(cols)}) "
                f"OUTPUT INSERTED.DeliveryOrderID VALUES ({', '.join(['%s'] * len(cols))})",
                tuple(vals))
            did = int(cursor.fetchone()[0])
            _log_activity(cursor, "deliveryorder", did, "created",
                          f"Sales schedule: {'service + top-up' if needs_service else 'top-up'} "
                          f"at {name} on {ds}"
                          + (f" (series {rule})" if series_id else ""), user)
            created.append({"id": did, "date": ds})

        conn.commit()
        conn.close()
        return jsonify({"ok": True, "machine_name": name, "created": created,
                        "skipped": skipped, "series_id": series_id, "repeat": rule})
    except Exception as e:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


def _cancel_stop_rows(cursor, ids, user, reason):
    """Delete open delivery orders by id, with the same care as /stops DELETE.

    Detaches any unsubmitted Work Order sheet first — a driver holding the sheet
    open must not have finalize later update a row that no longer exists — and
    snapshots each row to WO_DeletedLog so nothing vanishes without a trail.
    """
    done = []
    for did in ids:
        cursor.execute("SELECT Status, MachineName FROM WO_DeliveryOrders "
                       "WHERE DeliveryOrderID=%s", (did,))
        r = cursor.fetchone()
        if not r or (r[0] or "").lower() != "open":
            continue
        cursor.execute("""
            UPDATE WO_VisitSessions SET LinkedDeliveryOrderID = NULL
            WHERE LinkedDeliveryOrderID = %s
              AND Status NOT IN ('signed', 'pending_email_signature')
        """, (did,))
        snap = _row_snapshot(cursor,
                             "SELECT * FROM WO_DeliveryOrders WHERE DeliveryOrderID=%s", (did,))
        cursor.execute("DELETE FROM WO_DeliveryOrderLines WHERE DeliveryOrderID=%s", (did,))
        cursor.execute("DELETE FROM WO_Activity WHERE ParentType='deliveryorder' AND ParentID=%s", (did,))
        cursor.execute("DELETE FROM WO_DeliveryOrders WHERE DeliveryOrderID=%s", (did,))
        log_deletion(cursor, "deliveryorder", did,
                     f"{reason}: {(snap or {}).get('MachineName') or did}", snap, False, user)
        done.append(did)
    return done


@workorders_bp.route("/schedule/stops/<int:did>", methods=["DELETE"])
@require_roles(*SALES_ROLES)
def api_schedule_delete(did):
    """Cancel one scheduled stop.

    Sales may only cancel what is still unassigned — once dispatch has put a
    driver's name on it, it belongs to the round and only dispatch can pull it.
    """
    user = get_current_user()
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Status, AssignedTo FROM WO_DeliveryOrders "
                       "WHERE DeliveryOrderID=%s", (did,))
        r = cursor.fetchone()
        if not r:
            conn.close(); return jsonify({"error": "Stop not found."}), 404
        if (r[0] or "").lower() != "open":
            conn.close()
            return jsonify({"error": "This stop has already been worked; it cannot be cancelled."}), 400
        if r[1] and _caller_is_sales_only():
            conn.close()
            return jsonify({"error": f"Already assigned to {r[1]} — ask dispatch to pull it."}), 403
        gone = _cancel_stop_rows(cursor, [did], user, "Schedule cancel")
        conn.commit(); conn.close()
        return jsonify({"ok": True, "deleted": gone})
    except Exception as e:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/schedule/series/<series_id>", methods=["DELETE"])
@require_roles(*SALES_ROLES)
def api_schedule_delete_series(series_id):
    """Cancel the remaining occurrences of a repeat.

    ?from=YYYY-MM-DD limits it to that date onward (the calendar always sends
    the occurrence that was clicked), so history and any stop a driver is
    already carrying stay untouched. Assigned occurrences are LEFT ALONE and
    reported back rather than silently skipped.

    ?from= is caller-supplied, so for a sales-only caller it is floored at
    today: she cannot reach backwards and wipe occurrences ops still has to
    run. Dispatch keeps the full range — clearing a stale past series is
    legitimate board cleanup.

    NOT a complete backstop on its own: DELETE /schedule/stops/<did> carries no
    date check, so the same range can be cleared one row at a time. See the
    open item in the handoff note.
    """
    user = get_current_user()
    frm = _parse_date(request.args.get("from")) or _sgt_today()
    # 2026-08-11 rev2 — ?from= is caller-supplied. Sales must not be able to
    # reach backwards past today and wipe occurrences ops still has to run, so
    # her floor is today no matter what she sends. Dispatch keeps the full
    # range: clearing a stale past series is a legitimate board cleanup.
    if _caller_is_sales_only():
        frm = max(frm, _sgt_today())
    sid = (series_id or "").strip()
    if not sid:
        return jsonify({"error": "No series."}), 400
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        if not _has_col(cursor, "WO_DeliveryOrders", "SeriesID"):
            conn.close()
            return jsonify({"error": "Series columns are not migrated on this database yet."}), 503
        cursor.execute(
            "SELECT DeliveryOrderID, AssignedTo, CONVERT(VARCHAR(10), ScheduledDate, 23) "
            "FROM WO_DeliveryOrders WHERE SeriesID = %s AND Status = 'open' "
            "AND ScheduledDate >= %s", (sid, frm))
        rows = cursor.fetchall()
        sales_only = _caller_is_sales_only()
        ids, kept = [], []
        for r in rows:
            if r[1] and sales_only:
                kept.append({"id": int(r[0]), "date": r[2], "assigned_to": r[1]})
            else:
                ids.append(int(r[0]))
        gone = _cancel_stop_rows(cursor, ids, user, f"Series cancel ({sid[:8]})")
        conn.commit(); conn.close()
        return jsonify({"ok": True, "deleted": gone, "kept_assigned": kept})
    except Exception as e:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/movementorders/<int:mid>/assign", methods=["POST"])
@require_roles(*DISPATCH_ROLES)
def api_movement_assign(mid):
    """Reassign a movement order.

    This route did not exist before 2026-08-10: a movement's assignee could only
    ever be set at creation, which is why the Plan tab told dispatch to "assign
    via the Movements tab" for something no tab could do. Mirrors
    api_delivery_assign.

    A completed movement (StatusCode 2) is refused — its location cutover has
    already been applied, so moving it to another driver would be meaningless.
    """
    data = request.get_json(silent=True) or {}
    assigned = (data.get("assigned_to") or "").strip().lower() or None
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT StatusCode, AssignedTo FROM WO_MovementOrders WHERE MovementOrderID = %s",
            (mid,),
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Movement order not found."}), 404
        sc = int(row[0] or 0)
        if sc == 2:
            conn.close()
            return jsonify({"error": "Movement already completed; cannot reassign."}), 400
        # in_progress belongs to the driver who started it. Handing it over
        # without resetting would leave them unable to complete it (both /status
        # and /complete gate on AssignedTo) while the new owner inherits an
        # order that already reads in_progress. Reset to scheduled and say so.
        reset = sc == 1
        _sets = ["AssignedTo = %s", "StatusCode = %s"]
        _params = [assigned, 0 if reset else sc]
        _day = _parse_date(data.get("scheduled_date"))
        if _day and _has_col(cursor, "WO_MovementOrders", "ScheduledDate"):
            _sets.append("ScheduledDate = %s"); _params.append(_day)
        _params.append(mid)
        cursor.execute(
            f"UPDATE WO_MovementOrders SET {', '.join(_sets)} WHERE MovementOrderID = %s",
            tuple(_params),
        )
        _log_activity(cursor, "movementorder", mid, "assigned",
                      f"Assigned to {assigned or '(unassigned)'}"
                      + (f" (was {row[1]})" if row[1] else "")
                      + (" — reset from in_progress to scheduled" if reset else ""), user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/movementorders/<int:mid>/status", methods=["POST"])
@require_roles(*OPERATOR_ROLES)
def api_movement_status(mid):
    """Driver marks in_progress (1). Status 2 = complete is via /complete."""
    data = request.get_json(silent=True) or {}
    try:
        sc = int(data.get("status_code"))
    except (TypeError, ValueError):
        return jsonify({"error": "status_code required."}), 400
    if sc not in (0, 1):
        return jsonify({"error": "Use 0 or 1 here; use /complete to finish."}), 400
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT AssignedTo, StatusCode FROM WO_MovementOrders WHERE MovementOrderID=%s", (mid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found."}), 404
        if get_role(user) not in MANAGER_ROLES and (row[0] or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not assigned to you."}), 403
        cursor.execute("UPDATE WO_MovementOrders SET StatusCode=%s WHERE MovementOrderID=%s", (sc, mid))
        _log_activity(cursor, "movementorder", mid, "status",
                      _label(MOVEMENT_STATUS, sc), user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Delivery Orders ───────────────────────────────────────────────────────────

@workorders_bp.route("/deliveryorders", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_delivery_create():
    data = request.get_json(silent=True) or {}
    machine_name = (data.get("machine_name") or "").strip()
    machine_code = (data.get("machine_code") or "").strip() or None
    notes        = (data.get("notes") or "").strip() or None
    assigned     = (data.get("assigned_to") or "").strip().lower() or None
    priority     = (data.get("priority") or "normal").strip().lower()
    if priority not in ("low", "normal", "high"):
        priority = "normal"
    if not machine_name:
        return jsonify({"error": "Machine name is required."}), 400
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO WO_DeliveryOrders
                (MachineName, MachineCode, Notes, AssignedTo, Priority, CreatedBy)
            OUTPUT INSERTED.DeliveryOrderID
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (machine_name, machine_code, notes, assigned, priority, user))
        new_id = cursor.fetchone()[0]
        _log_activity(cursor, "deliveryorder", new_id, "created",
                      f"Delivery order created for {machine_name}", user)
        if assigned:
            _log_activity(cursor, "deliveryorder", new_id, "assigned",
                          f"Assigned to {assigned}", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/deliveryorders")
@api_login_required
def api_delivery_list():
    tab = (request.args.get("tab") or "mine").strip().lower()
    user = get_current_user()

    where, params = [], []
    if tab == "mine":
        where.append("AssignedTo = %s"); params.append(user)
        where.append("Status <> 'completed'")
    elif tab == "open":
        where.append("Status <> 'completed'")
    elif tab == "completed":
        where.append("Status = 'completed'")
    elif tab == "unassigned":
        where.append("AssignedTo IS NULL")
        where.append("Status <> 'completed'")

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sql = f"""
        SELECT TOP 300 DeliveryOrderID, MachineName, MachineCode, AssignedTo,
               Priority, Status, CreatedBy, CreatedAt, CompletedAt
        FROM WO_DeliveryOrders
        {where_sql}
        ORDER BY CASE Priority WHEN 'high' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END,
                 CreatedAt DESC
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "id": r[0], "machine_name": r[1], "machine_code": r[2],
            "assigned_to": r[3], "priority": r[4], "status": r[5],
            "created_by": r[6], "created_at": _iso(r[7]),
            "completed_at": _iso(r[8]),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/deliveryorders/<int:did>")
@api_login_required
def api_delivery_detail(did):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DeliveryOrderID, MachineName, MachineCode, Notes, AssignedTo,
                   Priority, Status,
                   Item1Qty, Item2Qty, Item3Qty, Item4Qty,
                   Item5Qty, Item6Qty, Item7Qty, Item8Qty,
                   RecipientName, CreatedBy, CreatedAt, CompletedBy, CompletedAt
            FROM WO_DeliveryOrders WHERE DeliveryOrderID = %s
        """, (did,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Delivery order not found."}), 404
        activity = _activity_for(cursor, "deliveryorder", did)
        # Quantities live in WO_DeliveryOrderLines whenever the order went
        # through a visit work order. The legacy Item1Qty..Item8Qty columns only
        # cover the first 8 of 17 catalogue items, so prefer the lines table and
        # fall back to the legacy columns for pre-visit-era orders.
        line_items = []
        try:
            cursor.execute("""
                SELECT i.Name, l.QtyOrdered, l.QtyDelivered, i.Unit, i.Content
                FROM WO_DeliveryOrderLines l
                INNER JOIN WO_DeliveryItems i ON i.ItemID = l.ItemID
                WHERE l.DeliveryOrderID = %s
                ORDER BY i.SortOrder, i.ItemID
            """, (did,))
            for n, (name, qo, qd, unit, content) in enumerate(cursor.fetchall(), start=1):
                line_items.append({"index": n, "label": name, "unit": unit,
                                   "content": content,
                                   "quantity": int(qd if qd is not None else (qo or 0)),
                                   "quantity_ordered": int(qo or 0)})
        except Exception:
            line_items = []
        conn.close()
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

    items = line_items or [{"index": i + 1, "label": DELIVERY_ITEMS[i],
                            "quantity": int(row[7 + i] or 0)} for i in range(8)]
    return jsonify({
        "id": row[0], "machine_name": row[1], "machine_code": row[2],
        "notes": row[3], "assigned_to": row[4], "priority": row[5],
        "status": row[6], "items": items, "recipient_name": row[15],
        "created_by": row[16], "created_at": _iso(row[17]),
        "completed_by": row[18], "completed_at": _iso(row[19]),
        "activity": activity,
    })


@workorders_bp.route("/deliveryorders/<int:did>", methods=["PATCH"])
@require_roles(*OPERATOR_ROLES)
def api_delivery_update(did):
    data = request.get_json(silent=True) or {}
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT AssignedTo, Status FROM WO_DeliveryOrders WHERE DeliveryOrderID = %s",
            (did,),
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Delivery order not found."}), 404
        role = get_role(user)
        if role == ROLE_OPERATOR and (row[0] or "").lower() != user:
            conn.close()
            return jsonify({"error": "You can only update your own delivery orders."}), 403

        quantities = data.get("quantities") or []
        if not isinstance(quantities, list) or len(quantities) != 8:
            return jsonify({"error": "Provide all 8 item quantities."}), 400
        clean = []
        for q in quantities:
            try:
                v = int(q)
            except (TypeError, ValueError):
                return jsonify({"error": "Quantities must be whole numbers."}), 400
            if v < 0 or v > 20:
                return jsonify({"error": "Each quantity must be between 0 and 20."}), 400
            clean.append(v)
        recipient = (data.get("recipient_name") or "").strip()[:255] or None

        cursor.execute("""
            UPDATE WO_DeliveryOrders
            SET Item1Qty=%s, Item2Qty=%s, Item3Qty=%s, Item4Qty=%s,
                Item5Qty=%s, Item6Qty=%s, Item7Qty=%s, Item8Qty=%s,
                RecipientName=%s
            WHERE DeliveryOrderID = %s
        """, (*clean, recipient, did))
        _log_activity(cursor, "deliveryorder", did, "updated",
                      "Quantities / recipient updated", user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/deliveryorders/<int:did>/complete", methods=["POST"])
@require_roles(*OPERATOR_ROLES)
def api_delivery_complete(did):
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT AssignedTo, MachineCode, RecipientName, Status FROM WO_DeliveryOrders "
            "WHERE DeliveryOrderID = %s",
            (did,),
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Delivery order not found."}), 404
        # Already completed (e.g. signed off via a visit work order) — completing
        # again would re-run _sync_topup and wipe the top-up baseline.
        if (row[3] or "").lower() == "completed":
            conn.close()
            return jsonify({"error": "This delivery order is already completed."}), 400
        role = get_role(user)
        if role == ROLE_OPERATOR and (row[0] or "").lower() != user:
            conn.close()
            return jsonify({"error": "You can only complete your own delivery orders."}), 403
        if not row[2]:
            conn.close()
            return jsonify({"error": "Add the recipient name (typed signature) before completing."}), 400

        cursor.execute("""
            UPDATE WO_DeliveryOrders
            SET Status='completed', CompletedBy=%s, CompletedAt=SYSUTCDATETIME()
            WHERE DeliveryOrderID = %s
        """, (user, did))
        _log_activity(cursor, "deliveryorder", did, "completed",
                      f"Signed for by {row[2]}", user)
        topup_note = None
        if row[1]:
            topup_note = _sync_topup(cursor, str(row[1]))
            _log_activity(cursor, "deliveryorder", did, "topup_synced", topup_note, user)
        else:
            topup_note = "No machine code on file — topup not synced to vending data."
            _log_activity(cursor, "deliveryorder", did, "topup_sync_skipped", topup_note, user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "topup_note": topup_note})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/deliveryorders/<int:did>/assign", methods=["POST"])
@require_roles(*DISPATCH_ROLES)
def api_delivery_assign(did):
    data = request.get_json(silent=True) or {}
    assigned = (data.get("assigned_to") or "").strip().lower() or None
    priority = data.get("priority")
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM WO_DeliveryOrders WHERE DeliveryOrderID = %s", (did,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({"error": "Delivery order not found."}), 404
        sets, params = ["AssignedTo = %s"], [assigned]
        if priority and priority in ("low", "normal", "high"):
            sets.append("Priority = %s"); params.append(priority)
        _day = _parse_date(data.get("scheduled_date"))
        if _day:
            sets.append("ScheduledDate = %s"); params.append(_day)
        params.append(did)
        cursor.execute(
            f"UPDATE WO_DeliveryOrders SET {', '.join(sets)} WHERE DeliveryOrderID = %s",
            tuple(params),
        )
        detail = f"Assigned to {assigned or '(unassigned)'}"
        if priority:
            detail += f"; priority -> {priority}"
        _log_activity(cursor, "deliveryorder", did, "assigned", detail, user)
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Assignment candidates ─────────────────────────────────────────────────────

@workorders_bp.route("/assignment/delivery-candidates")
@require_roles(*DISPATCH_ROLES)
def api_assign_delivery_candidates():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ml.MachineName, ml.MachineCode, ml.LastTopupTimestamp,
                   COUNT(v.t) AS VendsSince
            FROM MachineLookup ml
            LEFT JOIN (
                SELECT CAST([Machine Code] AS NVARCHAR(50)) AS code,
                       CAST([Date Time] AS FLOAT) AS t
                FROM [MasterData Table]
                WHERE LEN(CAST([Event Code] AS NVARCHAR(20))) = 6
                  AND CAST([Event Code] AS NVARCHAR(20)) LIKE '1%'
                GROUP BY CAST([Machine Code] AS NVARCHAR(50)), CAST([Date Time] AS FLOAT)
            ) v ON v.code = CAST(ml.MachineCode AS NVARCHAR(50))
               AND (ml.LastTopupTimestamp IS NULL OR v.t >= ml.LastTopupTimestamp)
            WHERE ml.MachineCode IS NOT NULL
            GROUP BY ml.MachineName, ml.MachineCode, ml.LastTopupTimestamp
        """)
        rows = cursor.fetchall()
        conn.close()
        out = []
        for r in rows:
            last_dt = from_ole_date(r[2])
            out.append({
                "machine_name": r[0],
                "machine_code": str(r[1]) if r[1] is not None else None,
                "last_topup":   last_dt.strftime("%Y-%m-%d %H:%M") if last_dt else None,
                "vends_since":  int(r[3]) if r[3] is not None else 0,
            })
        out.sort(key=lambda x: x["vends_since"], reverse=True)
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/assignment/joborder-candidates")
@require_roles(*DISPATCH_ROLES)
def api_assign_joborder_candidates():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ml.MachineName, ml.MachineCode,
                   ISNULL((
                       SELECT COUNT(*) FROM WO_Complaints c
                       WHERE c.StatusCode <> 2
                         AND (
                             (c.MachineCode IS NOT NULL AND c.MachineCode = ml.MachineCode)
                             OR (c.MachineCode IS NULL AND c.MachineName = ml.MachineName)
                         )
                   ), 0) AS OpenComplaints
            FROM MachineLookup ml
        """)
        rows = cursor.fetchall()
        conn.close()
        out = [{
            "machine_name": r[0],
            "machine_code": str(r[1]) if r[1] is not None else None,
            "open_complaints": int(r[2]),
        } for r in rows]
        out.sort(key=lambda x: x["open_complaints"], reverse=True)
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Manager views ────────────────────────────────────────────────────────────

@workorders_bp.route("/manager/overview")
@require_roles(*MANAGER_ROLES)
def api_manager_overview():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT JobOrderID, DisplayID, MachineName, AssignedTo,
                   PriorityCode, StatusCode, CreatedAt
            FROM WO_JobOrders WHERE StatusCode <> 2
            ORDER BY PriorityCode DESC, CreatedAt
        """)
        jobs = [{
            "id": r[0], "display_id": r[1], "machine_name": r[2], "assigned_to": r[3],
            "priority_code":  int(r[4]), "priority_label": _label(PRIORITY, r[4]),
            "status_code":    int(r[5]), "status_label":   _label(JOBORDER_STATUS, r[5]),
            "created_at": _iso(r[6]),
        } for r in cursor.fetchall()]

        cursor.execute("""
            SELECT DeliveryOrderID, MachineName, AssignedTo, Priority, Status, CreatedAt
            FROM WO_DeliveryOrders WHERE Status <> 'completed'
            ORDER BY CASE Priority WHEN 'high' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END,
                     CreatedAt
        """)
        deliveries = [{
            "id": r[0], "machine_name": r[1], "assigned_to": r[2],
            "priority": r[3], "status": r[4], "created_at": _iso(r[5]),
        } for r in cursor.fetchall()]
        conn.close()
        return jsonify({"job_orders": jobs, "delivery_orders": deliveries})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/manager/machines")
@require_roles(*OPERATOR_ROLES)
def api_manager_machines_search():
    q = (request.args.get("q") or "").strip()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        if q:
            cursor.execute("""
                SELECT MachineName, MachineCode FROM MachineLookup
                WHERE MachineCode IS NOT NULL
                  AND (MachineName LIKE %s OR CAST(MachineCode AS NVARCHAR(50)) LIKE %s)
                ORDER BY MachineName
            """, (f"%{q}%", f"%{q}%"))
        else:
            cursor.execute("""
                SELECT MachineName, MachineCode FROM MachineLookup
                WHERE MachineCode IS NOT NULL ORDER BY MachineName
            """)
        rows = cursor.fetchall()
        conn.close()
        seen = {}
        for r in rows:
            code = str(r[1]) if r[1] is not None else None
            if not code:
                continue
            if code not in seen:
                seen[code] = {"code": code, "names": []}
            if r[0] and r[0] not in seen[code]["names"]:
                seen[code]["names"].append(r[0])
        out = [{"code": v["code"], "name": " / ".join(v["names"]) or v["code"]}
               for v in seen.values()]
        # Machine History browses the full fleet (client-side search filters this
        # list), so return all machines. Cap only when it's an actual typeahead query.
        return jsonify(out[:50] if q else out)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/manager/machine/<path:code>/record-move", methods=["POST"])
@require_roles(*MANAGER_ROLES)
def api_record_move(code):
    """Reusable: record that a machine was at a PREVIOUS location until a given date,
    splitting its dated MachineLocationHistory so vends before the date read the old
    location and vends from the date read whatever the timeline already shows.

    Use this for machines that were moved via admin relabel (current location already
    correct) — no per-machine SQL. Body:
        previous_location : required (name the machine was at before the move)
        moved_on          : required 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM' (SGT)
        previous_lat/lon  : optional
    """
    data = request.get_json(silent=True) or {}
    prev_loc = (data.get("previous_location") or "").strip()
    moved_on = (data.get("moved_on") or "").strip()
    if not prev_loc or not moved_on:
        return jsonify({"error": "previous_location and moved_on are required."}), 400

    def _f(v):
        try:
            return float(v) if v not in (None, "", "null") else None
        except (TypeError, ValueError):
            return None
    prev_lat = _f(data.get("previous_lat"))
    prev_lon = _f(data.get("previous_lon"))

    dt = None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(moved_on, fmt); break
        except ValueError:
            continue
    if dt is None:
        return jsonify({"error": "moved_on must be YYYY-MM-DD or 'YYYY-MM-DD HH:MM' (SGT)."}), 400
    cut = to_ole_date(dt)   # vend [Date Time] is an SGT OLE float; move date is SGT too

    try:
        conn = get_connection()
        cursor = conn.cursor()
        # Find the interval that CONTAINS the move date (usually the current open one).
        cursor.execute("""
            SELECT TOP 1 HistoryID, LocationName, ValidFromOle, ValidToOle
            FROM MachineLocationHistory WITH (UPDLOCK, HOLDLOCK)
            WHERE MachineCode = %s
              AND %s >= ValidFromOle
              AND (ValidToOle IS NULL OR %s < ValidToOle)
            ORDER BY ValidFromOle DESC
        """, (str(code), cut, cut))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "No location interval covers that date. Check the machine has a current location."}), 400
        hid, cur_name, vfrom, vto = row
        vfrom = vfrom if vfrom is not None else 0.0
        if cut <= vfrom:
            conn.close()
            return jsonify({"error": "Move date must be AFTER the start of the covering interval."}), 400
        if prev_loc == cur_name:
            conn.close()
            return jsonify({"error": f"Previous location is the same as the current one ({cur_name})."}), 400

        # Split: the covering interval now starts at the move date; insert the old
        # location for the period before it. One transaction; open-interval invariant
        # preserved (we never touch ValidToOle of the covering row's open end).
        cursor.execute("UPDATE MachineLocationHistory SET ValidFromOle = %s WHERE HistoryID = %s",
                       (cut, hid))
        cursor.execute("""
            INSERT INTO MachineLocationHistory
                (MachineCode, LocationName, Latitude, Longitude, ValidFromOle, ValidToOle, Source)
            VALUES (%s, %s, %s, %s, %s, %s, 'record-move')
        """, (str(code), prev_loc, prev_lat, prev_lon, vfrom, cut))
        _log_activity(cursor, "machine", 0,
                      "location-move-recorded",
                      f"{code}: {prev_loc} until {dt:%Y-%m-%d %H:%M} → {cur_name}",
                      get_current_user())
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        try:
            conn.rollback(); conn.close()   # never leave the UPDATE half-applied
        except Exception:
            pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/manager/equipment-log")
@require_roles(*OPERATOR_ROLES)
def api_equipment_log():
    """Chronological history for one MachineCode."""
    code = (request.args.get("machine_code") or "").strip()
    if not code:
        return jsonify({"error": "machine_code is required."}), 400
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT MachineName, Latitude, Longitude,
                   LastTopupTimestamp, PreviousTopupTimestamp, CountBeforeLastTopup
            FROM MachineLookup WHERE MachineCode = %s
        """, (code,))
        ml_rows = cursor.fetchall()
        if not ml_rows:
            conn.close()
            return jsonify({"error": "Machine not found in MachineLookup."}), 404

        machine_names = sorted({r[0] for r in ml_rows if r[0]})
        last_topup_ole = ml_rows[0][3]
        last_topup    = from_ole_date(last_topup_ole)
        prev_topup    = from_ole_date(ml_rows[0][4])
        vends_before  = int(ml_rows[0][5]) if ml_rows[0][5] is not None else None

        # Vends SINCE last top-up (more actionable than vends_before for the operator)
        if last_topup_ole is not None:
            cursor.execute("""
                SELECT COUNT(DISTINCT CAST(mdt.[Date Time] AS FLOAT)) FROM [MasterData Table] mdt
                WHERE CAST(mdt.[Machine Code] AS NVARCHAR(50)) = %s
                  AND LEN(CAST(mdt.[Event Code] AS NVARCHAR(20))) = 6
                  AND CAST(mdt.[Event Code] AS NVARCHAR(20)) LIKE '1%%'
                  AND CAST(mdt.[Date Time] AS FLOAT) >= %s
            """, (code, float(last_topup_ole)))
        else:
            cursor.execute("""
                SELECT COUNT(DISTINCT CAST(mdt.[Date Time] AS FLOAT)) FROM [MasterData Table] mdt
                WHERE CAST(mdt.[Machine Code] AS NVARCHAR(50)) = %s
                  AND LEN(CAST(mdt.[Event Code] AS NVARCHAR(20))) = 6
                  AND CAST(mdt.[Event Code] AS NVARCHAR(20)) LIKE '1%%'
            """, (code,))
        vends_since = int(cursor.fetchone()[0])

        events = []

        cursor.execute("""
            SELECT ComplaintID, DisplayID, Description, Source,
                   ImpactDescription, ImpactSeverity,
                   StatusCode, SubmitterEmail, SubmittedAt, JobOrderID
            FROM WO_Complaints WHERE MachineCode = %s
        """, (code,))
        complaint_ids = []
        for r in cursor.fetchall():
            complaint_ids.append(r[0])
            summary = (r[2] or "")[:120]
            if r[5] is not None:
                summary += f" (severity {int(r[5])}/5)"
            events.append({
                "at": _iso(r[8]), "type": "complaint", "id": r[0],
                "display_id": r[1],
                "action": "submitted", "by": r[7], "summary": summary,
                "status": _label(COMPLAINT_STATUS, r[6]),
                "linked_job_order": r[9],
            })

        cursor.execute("""
            SELECT JobOrderID, DisplayID, ComplaintID, AssignedTo,
                   PriorityCode, StatusCode,
                   CreatedBy, CreatedAt, CompletedBy, CompletedAt, Notes
            FROM WO_JobOrders WHERE MachineCode = %s
        """, (code,))
        job_ids = []
        for r in cursor.fetchall():
            job_ids.append(r[0])
            link = f" (from complaint #{r[2]})" if r[2] else ""
            events.append({
                "at": _iso(r[7]), "type": "joborder", "id": r[0],
                "display_id": r[1],
                "action": "created", "by": r[6],
                "summary": f"Assigned to {r[3] or 'unassigned'} ({_label(PRIORITY, r[4])} priority){link}",
                "status": _label(JOBORDER_STATUS, r[5]),
            })
            if r[9]:
                events.append({
                    "at": _iso(r[9]), "type": "joborder", "id": r[0],
                    "display_id": r[1],
                    "action": "completed", "by": r[8],
                    "summary": "Completed", "status": "closed",
                })

        cursor.execute("""
            SELECT DeliveryOrderID, AssignedTo, Priority, Status,
                   CreatedBy, CreatedAt, CompletedBy, CompletedAt, RecipientName
            FROM WO_DeliveryOrders WHERE MachineCode = %s
        """, (code,))
        delivery_ids = []
        for r in cursor.fetchall():
            delivery_ids.append(r[0])
            events.append({
                "at": _iso(r[5]), "type": "deliveryorder", "id": r[0],
                "action": "created", "by": r[4],
                "summary": f"Assigned to {r[1] or 'unassigned'} ({r[2]} priority)",
                "status": r[3],
            })
            if r[7]:
                events.append({
                    "at": _iso(r[7]), "type": "deliveryorder", "id": r[0],
                    "action": "completed", "by": r[6],
                    "summary": f"Signed for by {r[8] or '—'}",
                    "status": "completed",
                })

        # Movement orders
        cursor.execute("""
            SELECT MovementOrderID, DisplayID, MovementType,
                   FromLocation, ToLocation, AssignedTo,
                   StatusCode, CreatedBy, CreatedAt, CompletedBy, CompletedAt
            FROM WO_MovementOrders WHERE MachineCode = %s
        """, (code,))
        movement_ids = []
        for r in cursor.fetchall():
            movement_ids.append(r[0])
            events.append({
                "at": _iso(r[8]), "type": "movementorder", "id": r[0],
                "display_id": r[1],
                "action": "created", "by": r[7],
                "summary": f"{r[2]}: {r[3] or '?'} → {r[4] or '?'} (driver: {r[5] or 'unassigned'})",
                "status": _label(MOVEMENT_STATUS, r[6]),
            })
            if r[10]:
                events.append({
                    "at": _iso(r[10]), "type": "movementorder", "id": r[0],
                    "display_id": r[1],
                    "action": "completed", "by": r[9],
                    "summary": f"{r[2]} applied",
                    "status": "completed",
                })

        def _activity_in(parent_type, ids):
            if not ids:
                return
            placeholders = ", ".join(["%s"] * len(ids))
            cursor.execute(f"""
                SELECT ParentID, Action, Detail, ByUser, AtTime
                FROM WO_Activity
                WHERE ParentType = %s AND ParentID IN ({placeholders})
                  AND Action NOT IN ('created', 'submitted', 'completed')
            """, (parent_type,) + tuple(ids))
            for r in cursor.fetchall():
                events.append({
                    "at": _iso(r[4]), "type": parent_type, "id": r[0],
                    "action": r[1], "by": r[3],
                    "summary": r[2] or r[1],
                })

        _activity_in("complaint",     complaint_ids)
        _activity_in("joborder",      job_ids)
        _activity_in("deliveryorder", delivery_ids)
        _activity_in("movementorder", movement_ids)

        # ── Location history (every place this machine has been) ──────────────────
        location_history = []
        try:
            cursor.execute("""
                SELECT LocationName, ValidFromOle, ValidToOle, Source, HistoryID
                FROM MachineLocationHistory
                WHERE MachineCode = %s
                ORDER BY ValidFromOle DESC
            """, (str(code),))
            for r in cursor.fetchall():
                vf = from_ole_date(r[1]) if r[1] not in (None, 0, 0.0) else None
                vt = from_ole_date(r[2]) if r[2] is not None else None
                location_history.append({
                    "location":   r[0],
                    "from":       vf.strftime("%Y-%m-%d %H:%M") if vf else "—",
                    "to":         vt.strftime("%Y-%m-%d %H:%M") if vt else None,  # None = current
                    "current":    r[2] is None and r[0] != '(decommissioned)',
                    "source":     r[3],
                    "history_id": r[4],
                })
        except Exception as he:
            print(f"[machine_history] location_history skipped: {he}")

        conn.close()
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

    events.sort(key=lambda e: e.get("at") or "", reverse=True)

    return jsonify({
        "machine": {
            "code": code,
            "names": machine_names,
            "last_topup":              _iso(last_topup),
            "previous_topup":          _iso(prev_topup),
            "vends_before_last_topup": vends_before,
            "vends_since_last_topup":  vends_since,
        },
        "counts": {
            "complaints":      len(complaint_ids),
            "job_orders":      len(job_ids),
            "delivery_orders": len(delivery_ids),
            "movement_orders": len(movement_ids),
        },
        "location_history": location_history,
        "events": events[:300],
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  Operator Work Order — unified PDF-template flow (Phase 2 / 2026-06-09)
#  Replaces the existing operator/driver "Tech Support" flow per COO directive.
#  Combines open Service WOs + open Delivery WO per machine into one document.
# ═══════════════════════════════════════════════════════════════════════════════

# ── Service category constants (fixed 4 per the PDF template) ────────────────
# ── Machine attributes (2026-08-10) ───────────────────────────────────────────
# Five physical facts about a machine that change what a driver puts in the van.
# The option lists below are authoritative: the UI renders them verbatim and the
# API rejects anything outside them, so these columns cannot drift into free text.
#
# NULL means "never recorded". That is deliberately NOT the same as the "Absent"
# option, which is a real answer meaning the feature is not fitted.
#
# 2026-08-11 — capture moved to SALES. It used to be the operator's Work Order
# sheet, gated on finalize. That reasoning ("only the person at the machine can
# see this") was correct about the physics and wrong about the people: a field
# tech who is handed data entry stops using the app, and then nothing gets
# recorded at all. Sales already knows the specification when the machine is
# sold, so they enter it once — by hand or by spreadsheet — and the driver is
# never asked. There is NO gate anywhere: an incomplete machine is a sales
# backlog item, never a reason a signed Work Order cannot close.
MACHINE_ATTRS = [
    # (payload key,      DB column,         label,               options)
    ("machine_type",     "MachineType",     "Type of machine",   ["Standing", "Standing-Fr", "Tabletop", "Lite"]),
    ("water_source",     "WaterSource",     "Machine water",     ["Piped", "Carried"]),
    ("tea_syrup",        "TeaSyrup",        "Tea and syrup",     ["Installed", "Absent"]),
    ("payment_terminal", "PaymentTerminal", "Payment terminal",  ["NETS", "Other", "Cash", "Absent"]),
    ("data_connection",  "DataConnection",  "Data connection",   ["Own", "Local WiFi", "Absent"]),
]
MACHINE_ATTR_COLS = [a[1] for a in MACHINE_ATTRS]


def _machine_attrs_spec():
    """Option lists for the UI. DB column names are deliberately not exposed."""
    return [{"key": k, "label": lb, "options": o} for k, _c, lb, o in MACHINE_ATTRS]


def _validate_machine_attrs(payload):
    """PURE. No DB. -> (clean {key: value}, error_message_or_None).

    Called before any cursor.execute in every caller, so a rejection can never
    discard work that was already written in the same request.
    """
    clean = {}
    if not payload:
        return clean, None
    if not isinstance(payload, dict):
        return clean, "machine_attrs must be an object."
    for key, _col, label, options in MACHINE_ATTRS:
        if key not in payload:
            continue
        raw = payload.get(key)
        val = "" if raw is None else str(raw).strip()   # str(): a non-string must not crash
        if not val:
            continue
        if val not in options:
            return {}, f"{label}: '{val}' is not one of {', '.join(options)}."
        clean[key] = val
    return clean, None


def _missing_from(vals):
    """Single source of truth for 'what is still unrecorded'. Used by both the
    UI hint and the finalize gate, so the two cannot drift."""
    return [lb for k, _c, lb, _o in MACHINE_ATTRS if not str(vals.get(k) or "").strip()]


def _read_machine_attrs(cursor, code):
    """-> (values_by_key, machine_row_exists). Missing row -> ({all None}, False)."""
    cursor.execute(
        f"SELECT TOP 1 {', '.join(MACHINE_ATTR_COLS)} FROM MachineLookup WHERE MachineCode = %s",
        (str(code),),
    )
    r = cursor.fetchone()
    if not r:
        return {a[0]: None for a in MACHINE_ATTRS}, False
    return {MACHINE_ATTRS[i][0]: r[i] for i in range(len(MACHINE_ATTRS))}, True


def _write_machine_attrs(cursor, code, clean, user=None, clear=None):
    """Persist already-validated values. -> rows affected.

    Only keys present in `clean` are written, so a partial save never blanks a
    value someone else recorded. Callers MUST check the return: None means
    "nothing was asked of us"; any n <= 0 means the write did not land (0 = no
    such MachineCode, -1 = pymssql could not determine, e.g. a trigger with
    NOCOUNT ON). Either has to surface rather than silently leaving the machine
    incomplete forever.
    """
    clear = set(clear or ())
    if not clean and not clear:
        return None                    # nothing asked of us; not a failed write
    sets, params = [], []
    for key, col, _label, _options in MACHINE_ATTRS:
        if key in clean:
            sets.append(f"{col} = %s")
            params.append(clean[key])
        elif key in clear:
            # Explicit un-set. Only api_machine_attrs_update passes this; the
            # importer never does, so a blank cell still means "no opinion".
            sets.append(f"{col} = NULL")
    if not sets:
        return None
    params.append(str(code))
    cursor.execute(
        f"UPDATE MachineLookup SET {', '.join(sets)} WHERE MachineCode = %s",
        tuple(params),
    )
    n = cursor.rowcount
    if n and user:
        try:
            _log_activity(
                cursor, "machine", 0, "machine_attrs_updated",
                f"{code}: " + ", ".join(
            list(f"{k}={v}" for k, v in sorted(clean.items()))
            + list(f"{k}=(cleared)" for k in sorted(clear))),
                user,
            )
        except Exception as e:                       # never fail a visit over the audit row
            print(f"[machine_attrs] activity log skipped for {code}: {e}")
    return n


# ── Machine properties: the Sales entry points ───────────────────────────────
# SALES_ROLES = MANAGER_ROLES | {sales}. Operators are deliberately absent: this
# is the surface that replaced their data entry, not a second place to do it.

@workorders_bp.route("/machines/attributes")
@require_roles(*SALES_ROLES)
def api_machine_attrs_list():
    """Every machine with its five properties and what is still missing."""
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cols = ", ".join(MACHINE_ATTR_COLS)
        cursor.execute(f"""
            SELECT MachineCode, MachineName, ISNULL(IsActive, 1), {cols}
            FROM MachineLookup
            ORDER BY MachineName, MachineCode
        """)
        rows = cursor.fetchall()
    except Exception as e:
        # Columns are added by init_db, which swallows individual ALTER failures,
        # so name the real cause rather than a generic database error.
        return jsonify({"error": f"Could not read machine properties: {e}",
                        "spec": _machine_attrs_spec(), "machines": []}), 500
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
    out = []
    for r in rows:
        vals = {MACHINE_ATTRS[i][0]: r[3 + i] for i in range(len(MACHINE_ATTRS))}
        out.append({"code": r[0], "name": r[1], "active": bool(r[2]),
                    "attrs": vals, "missing": _missing_from(vals)})
    return jsonify({"spec": _machine_attrs_spec(), "machines": out,
                    "incomplete": sum(1 for m in out if m["missing"] and m["active"])})


@workorders_bp.route("/machines/<path:code>/attributes", methods=["PATCH"])
@require_roles(*SALES_ROLES)
def api_machine_attrs_update(code):
    """Set one machine's properties. Partial: only the keys sent are written."""
    data = request.get_json(silent=True) or {}
    src = data.get("machine_attrs", data)
    if not isinstance(src, dict):
        return jsonify({"error": "machine_attrs must be an object."}), 400
    clean, err = _validate_machine_attrs(src)
    if err:
        return jsonify({"error": err}), 400
    # "— not recorded —" has to actually un-set the value. _validate_machine_attrs
    # DROPS blanks, which is right for the import (an empty cell means "no
    # opinion") and wrong here: a sales user correcting a mis-entered value would
    # otherwise get "saved" and the old value still on the machine, with no way
    # anywhere in the product to clear it.
    nulls = [k for k, _c, _lb, _o in MACHINE_ATTRS
             if k in src and not str(src[k] or "").strip()]
    if not clean and not nulls:
        return jsonify({"error": "Nothing to save."}), 400
    user = get_current_user()
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        n = _write_machine_attrs(cursor, code, clean, user, clear=nulls)
        if n is not None and n <= 0:
            return jsonify({"error": f"{code} is not in the machine registry."}), 404
        # Read BEFORE commit: a transient failure on the read-back would
        # otherwise report "save failed" for a save that had already landed, and
        # the user would do it twice.
        vals, exists = _read_machine_attrs(cursor, code)
        if not exists:
            return jsonify({"error": f"{code} is not in the machine registry."}), 404
        conn.commit()
        return jsonify({"ok": True, "code": code, "attrs": vals,
                        "missing": _missing_from(vals)})
    except Exception as e:
        return jsonify({"error": f"Database error: {e}"}), 500
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


# Header -> payload key. Both the DB column name and the human label are
# accepted, because sales will export the template AND hand-build sheets.
_ATTR_HEADERS = {}
for _k, _c, _lb, _o in MACHINE_ATTRS:
    for _alias in (_k, _c, _lb):
        _ATTR_HEADERS[_alias.strip().lower().replace(" ", "").replace("_", "")] = _k
# ORDERED, not a set: set iteration order depends on PYTHONHASHSEED, which
# gunicorn randomises per worker, so a sheet with both "Machine" (a name) and
# "MachineCode" would match a different column on every upload.
_CODE_HEADERS = ("machinecode", "machineid", "code", "machine")


def _canon_option(key, val):
    """Case- and space-insensitive match onto the exact option.

    Strict matching is right for the API and wrong for a spreadsheet: "piped"
    and "Local Wifi" are not mistakes a person needs told about, and rejecting
    the row throws away its four good values too. Returns the input unchanged
    when nothing matches, so a genuine typo still fails validation.
    """
    for k, _c, _lb, options in MACHINE_ATTRS:
        if k == key:
            t = " ".join(val.split()).lower()
            for o in options:
                if o.lower() == t:
                    return o
    return val


def _norm_header(h):
    """Fold a header cell to its key. Everything from "(" on is dropped so the
    template's own "Machine water (Piped/Carried)" matches "Machine water"."""
    t = str(h or "").split("(")[0]
    return t.strip().lower().replace(" ", "").replace("_", "")


def _parse_upload(filename, blob):
    """-> (rows, error). rows are dicts keyed by the normalised header.

    .xlsx via openpyxl; .csv/.txt natively. Legacy .xls is a different binary
    format openpyxl cannot read, so it is refused by name with instructions
    rather than failing with a confusing parser error.
    """
    name = (filename or "").lower()
    if name.endswith(".xls"):
        return None, ("Legacy .xls is not supported. Open it in Excel and use "
                      "File > Save As > Excel Workbook (.xlsx), or export as CSV.")
    if name.endswith(".csv") or name.endswith(".txt"):
        import csv as _csv
        try:
            text = blob.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = blob.decode("latin-1")
        rdr = _csv.reader(io.StringIO(text))
        table = [r for r in rdr]
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            import openpyxl
        except ImportError:
            return None, ("Spreadsheet support is not installed on the server "
                          "(openpyxl). Save the file as CSV and upload that.")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        except Exception as e:
            return None, f"Could not read that workbook: {e}"
        if not wb.sheetnames:
            wb.close()
            return None, "That workbook has no sheets."
        # First sheet that actually has a machine-code column. Hand-built
        # workbooks routinely open on a cover or notes tab, and blaming the
        # headers for that sends the user to fix the wrong thing.
        table, used = [], None
        for sn in wb.sheetnames:
            rows_ = [[c for c in row] for row in wb[sn].iter_rows(values_only=True)]
            rows_ = [r for r in rows_ if any(str(c or "").strip() for c in r)]
            if rows_ and any(_norm_header(c) in _CODE_HEADERS for c in rows_[0]):
                table, used = rows_, sn
                break
            if not table:
                table, used = rows_, sn
        wb.close()
    else:
        return None, "Upload a .xlsx or .csv file."
    table = [r for r in table if any(str(c or "").strip() for c in r)]
    if not table:
        return None, "That file has no rows."
    headers = [_norm_header(c) for c in table[0]]
    if not any(h in _CODE_HEADERS for h in headers):
        return None, ("No machine-code column. The first row must contain a "
                      "MachineCode header — download the template. "
                      "(Read the headers as: " + ", ".join(str(c) for c in table[0][:6]) + ")")
    rows = []
    for raw in table[1:]:
        rows.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
    return rows, None


@workorders_bp.route("/machines/attributes/template.csv")
@require_roles(*SALES_ROLES)
def api_machine_attrs_template():
    """The exact headers the importer accepts, with each option list in the
    header cell so the file is self-documenting without a second data row."""
    # Headers only. An instruction row would be read back as data row 2 on every
    # import and reported as "not in the machine registry" forever. The option
    # lists live in the header text, which _norm_header strips back to the key.
    heads = ["MachineCode"] + [f"{lb} ({'/'.join(o)})" for _k, _c, lb, o in MACHINE_ATTRS]
    body = ",".join(f'"{x}"' for x in heads) + "\n"
    return Response(body, mimetype="text/csv", headers={
        "Content-Disposition": "attachment; filename=machine-properties-template.csv"})


@workorders_bp.route("/machines/attributes/import", methods=["POST"])
@require_roles(*SALES_ROLES)
def api_machine_attrs_import():
    """Upload a sheet of properties. Matched strictly on MachineCode.

    Defaults to a DRY RUN: send apply=1 to write. Every row gets an outcome, so
    a partial import is legible instead of mysterious. A row that names an
    unknown machine is reported and skipped — never guessed at by name.
    """
    # Checked BEFORE reading: f.read() on a 600 MB body materialises it in the
    # worker and takes the whole app down with it.
    if (request.content_length or 0) > 6 * 1024 * 1024:
        return jsonify({"error": "That file is too large. Export the sheet as CSV, "
                                 "or split it."}), 413
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded."}), 400
    blob = f.read()
    if len(blob) > 4 * 1024 * 1024:
        return jsonify({"error": "File is over 4 MB. Split it or export as CSV."}), 413
    rows, err = _parse_upload(f.filename, blob)
    if err:
        return jsonify({"error": err}), 400
    apply = request.args.get("apply") in ("1", "true", "yes")
    user = get_current_user()

    results, ok, skipped = [], 0, 0
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT MachineCode FROM MachineLookup")
        known = {str(r[0]).strip().lower(): str(r[0]) for r in cursor.fetchall()}

        for i, row in enumerate(rows, start=2):        # row 1 is the header
            code_raw = ""
            for h in _CODE_HEADERS:   # ordered: most specific first
                if row.get(h) not in (None, ""):
                    code_raw = str(row[h]).strip()
                    break
            # Excel turns a numeric code into 54020502.0 — strip that, it is not
            # a different machine.

            if not code_raw:
                results.append({"row": i, "code": "", "status": "skipped",
                                "detail": "no machine code"}); skipped += 1; continue
            # Match the raw value FIRST: a code that genuinely ends in ".0"
            # must not be rewritten into a different machine's code. Only if
            # that misses do we treat it as Excel's numeric formatting.
            real = known.get(code_raw.lower())
            if not real and code_raw.endswith(".0") and code_raw[:-2].isdigit():
                real = known.get(code_raw[:-2].lower())
            if not real:
                results.append({"row": i, "code": code_raw, "status": "skipped",
                                "detail": "not in the machine registry (if the code has "
                                          "leading zeros, format that column as Text in Excel)"})
                skipped += 1; continue
            payload = {}
            for h, v in row.items():
                key = _ATTR_HEADERS.get(h)
                if key and str(v or "").strip():
                    payload[key] = _canon_option(key, str(v).strip())
            if not payload:
                results.append({"row": i, "code": real, "status": "skipped",
                                "detail": "no property values"}); skipped += 1; continue
            clean, verr = _validate_machine_attrs(payload)
            if verr:
                results.append({"row": i, "code": real, "status": "error",
                                "detail": verr}); skipped += 1; continue
            if apply:
                n = _write_machine_attrs(cursor, real, clean, user)
                # Commit in batches. One transaction over a few thousand rows
                # escalates to a table lock on MachineLookup, which every
                # driver's sheet and every board refresh reads through.
                if ok and ok % 200 == 0:
                    conn.commit()
                if not n or n <= 0:
                    results.append({"row": i, "code": real, "status": "error",
                                    "detail": "write did not land"}); skipped += 1; continue
            results.append({"row": i, "code": real,
                            "status": "would set" if not apply else "set",
                            "detail": ", ".join(f"{k}={v}" for k, v in sorted(clean.items()))})
            ok += 1
        if apply:
            conn.commit()
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({"error": f"Database error: {e}"}), 500
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
    return jsonify({"applied": apply, "ok": ok, "skipped": skipped,
                    "rows": len(rows), "results": results})


SERVICE_CATEGORIES = [
    ("PMC", "General Machine Preventive Maintenance and Cleaning"),
    ("CMR", "Machine Corrective Maintenance and Recovery"),
    ("INR", "Machine Installation or Replacement"),
    ("OTH", "Other Services"),
]


# ── Admin: WO_DeliveryItems CRUD (config) ────────────────────────────────────

@workorders_bp.route("/admin/delivery-items")
@api_login_required
def api_delivery_items_list():
    """All items; admins see inactive too. Used by operator visit form + admin UI."""
    include_inactive = (
        request.args.get("include_inactive") in ("1", "true", "yes")
        and get_role(get_current_user()) == ROLE_ADMIN
    )
    try:
        conn = get_connection()
        cursor = conn.cursor()
        sql = """
            SELECT ItemID, Name, Unit, Content, SortOrder, IsActive,
                   CreatedBy, CreatedAt, UpdatedBy, UpdatedAt
            FROM WO_DeliveryItems
        """
        if not include_inactive:
            sql += " WHERE IsActive = 1"
        sql += " ORDER BY SortOrder, ItemID"
        cursor.execute(sql)
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            "id": r[0], "name": r[1], "unit": r[2], "content": r[3],
            "sort_order": int(r[4]), "is_active": bool(r[5]),
            "created_by": r[6], "created_at": _iso(r[7]),
            "updated_by": r[8], "updated_at": _iso(r[9]),
        } for r in rows])
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/admin/delivery-items", methods=["POST"])
@require_roles(ROLE_ADMIN)
def api_delivery_items_create():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    unit = (data.get("unit") or "").strip()
    content = (data.get("content") or "").strip()
    if not name or not unit or not content:
        return jsonify({"error": "name, unit, content required."}), 400
    try:
        sort_order = int(data.get("sort_order") or 999)
    except (TypeError, ValueError):
        sort_order = 999
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO WO_DeliveryItems (Name, Unit, Content, SortOrder, CreatedBy)
            OUTPUT INSERTED.ItemID
            VALUES (%s, %s, %s, %s, %s)
        """, (name[:255], unit[:50], content[:50], sort_order, user))
        new_id = int(cursor.fetchone()[0])
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/admin/delivery-items/<int:iid>", methods=["PATCH"])
@require_roles(ROLE_ADMIN)
def api_delivery_items_update(iid):
    data = request.get_json(silent=True) or {}
    sets, params = [], []
    for fld, col in (("name", "Name"), ("unit", "Unit"), ("content", "Content")):
        if fld in data:
            sets.append(f"{col} = %s")
            params.append((str(data[fld]) or "").strip()[:255] or None)
    if "sort_order" in data:
        try:
            sets.append("SortOrder = %s")
            params.append(int(data["sort_order"]))
        except (TypeError, ValueError):
            pass
    if "is_active" in data:
        sets.append("IsActive = %s")
        params.append(1 if data["is_active"] else 0)
    if not sets:
        return jsonify({"error": "No fields to update."}), 400
    user = get_current_user()
    sets.append("UpdatedBy = %s"); params.append(user)
    sets.append("UpdatedAt = SYSUTCDATETIME()")
    params.append(iid)
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            f"UPDATE WO_DeliveryItems SET {', '.join(sets)} WHERE ItemID = %s",
            tuple(params),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/admin/delivery-items/<int:iid>", methods=["DELETE"])
@require_roles(ROLE_ADMIN)
def api_delivery_items_delete(iid):
    """Soft-delete: IsActive=0 (preserves historical lines)."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE WO_DeliveryItems SET IsActive=0 WHERE ItemID=%s",
            (iid,),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── Operator: color-coded location list ──────────────────────────────────────

@workorders_bp.route("/operator/locations")
@require_roles(*OPERATOR_ROLES)
def api_operator_locations():
    """Returns distinct machines where the current operator has open work.
    Color codes:
      'red'   = at least one open service JobOrder assigned to me
      'blue'  = an open delivery WO assigned to me
      'both'  = both red + blue
    """
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        # Future-dated stops belong to a later round and must not show today.
        # Guarded: if the column has not been added yet, fall back to the
        # pre-scheduling behaviour rather than failing the driver's whole list.
        _dt = "AND (ScheduledDate IS NULL OR ScheduledDate <= %s)"
        _jo_dt = _dt if _has_col(cursor, "WO_JobOrders", "ScheduledDate") else ""
        _do_dt = _dt if _has_col(cursor, "WO_DeliveryOrders", "ScheduledDate") else ""
        _today = _sgt_today()

        # Service WOs with summary fields so operator gets an overview before tapping in.
        cursor.execute("""
            SELECT MachineCode, MachineName, JobOrderID, DisplayID, PriorityCode,
                   StatusCode, Diagnosis, ProposedFix, AttachedKBID, CreatedAt
            FROM WO_JobOrders
            WHERE AssignedTo = %s AND StatusCode IN (0, 1) """ + _jo_dt + """
            ORDER BY MachineCode, PriorityCode DESC, CreatedAt
        """, (user,) + ((_today,) if _jo_dt else ()))
        service_rows = cursor.fetchall()

        # Pre-fetch KB titles in one shot (small set)
        kb_titles = {}
        kb_ids = [r[8] for r in service_rows if r[8] is not None]
        if kb_ids:
            ph = ",".join(["%s"] * len(kb_ids))
            cursor.execute(
                f"SELECT KBID, Title FROM WO_KB_Entries WHERE KBID IN ({ph})",
                tuple(kb_ids),
            )
            kb_titles = {int(r[0]): r[1] for r in cursor.fetchall()}

        cursor.execute("""
            SELECT MachineCode, MachineName, DeliveryOrderID, Priority,
                   Notes, CreatedAt
            FROM WO_DeliveryOrders
            WHERE AssignedTo = %s AND Status <> 'completed' """ + _do_dt + """
            ORDER BY MachineCode, CreatedAt
        """, (user,) + ((_today,) if _do_dt else ()))
        delivery_rows = cursor.fetchall()

        # Per-delivery-order line totals (so the card can say "8 items")
        do_ids = [r[2] for r in delivery_rows]
        item_counts = {}
        if do_ids:
            ph = ",".join(["%s"] * len(do_ids))
            cursor.execute(f"""
                SELECT DeliveryOrderID, SUM(QtyOrdered) AS UnitsOrdered, COUNT(*) AS LineCount
                FROM WO_DeliveryOrderLines
                WHERE DeliveryOrderID IN ({ph})
                GROUP BY DeliveryOrderID
            """, tuple(do_ids))
            item_counts = {int(r[0]): {"units": int(r[1] or 0), "lines": int(r[2] or 0)}
                           for r in cursor.fetchall()}

        conn.close()
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

    # Group everything by machine
    by_code = {}
    for r in service_rows:
        code = str(r[0]) if r[0] is not None else None
        if not code: continue
        rec = by_code.setdefault(code, {"name": r[1], "services": [], "deliveries": []})
        diag = (r[6] or "").strip()
        fix  = (r[7] or "").strip()
        kb_title = kb_titles.get(int(r[8])) if r[8] is not None else None
        if kb_title:
            summary = kb_title
        elif diag:
            summary = diag[:80] + ("…" if len(diag) > 80 else "")
        elif fix:
            summary = fix[:80] + ("…" if len(fix) > 80 else "")
        else:
            summary = "Service requested — details inside"
        rec["services"].append({
            "id":            int(r[2]),
            "display_id":    r[3],
            "priority_code": int(r[4]),
            "priority_label": _label(PRIORITY, r[4]),
            "status_code":   int(r[5]),
            "status_label":  _label(JOBORDER_STATUS, r[5]),
            "summary":       summary,
        })

    for r in delivery_rows:
        code = str(r[0]) if r[0] is not None else None
        if not code: continue
        rec = by_code.setdefault(code, {"name": r[1], "services": [], "deliveries": []})
        ic = item_counts.get(int(r[2]), {"units": 0, "lines": 0})
        if ic["lines"]:
            summary = f"{ic['lines']} item line{'s' if ic['lines'] != 1 else ''}, {ic['units']} unit{'s' if ic['units'] != 1 else ''} planned"
        else:
            summary = "Top-up scheduled — open to add items on site"
        rec["deliveries"].append({
            "id":         int(r[2]),
            "priority":   r[3],
            "summary":    summary,
            "notes":      (r[4] or "")[:120],
        })

    out = []
    for code in sorted(by_code):
        rec = by_code[code]
        has_svc = bool(rec["services"])
        has_del = bool(rec["deliveries"])
        color = "both" if (has_svc and has_del) else ("red" if has_svc else "blue")
        out.append({
            "machine_code": code,
            "machine_name": rec["name"] or code,
            "color":        color,
            "open_service_count":  len(rec["services"]),
            "open_delivery_count": len(rec["deliveries"]),
            "services":            rec["services"],
            "deliveries":          rec["deliveries"],
        })
    return jsonify(out)


# ── Operator: location detail (combined view) ────────────────────────────────

@workorders_bp.route("/operator/location/<path:code>")
@require_roles(*OPERATOR_ROLES)
def api_operator_location_detail(code):
    """Returns the operator's view of one machine:
      - Open service JobOrders assigned to me (with KB ref + diagnosis + proposed fix)
      - The single open delivery WO assigned to me (if any) with its lines
      - The 17 active delivery items (for the PDF-template grid)
    """
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        # Work scheduled for a later day is not part of today's visit. Without
        # this, openWorkOrder passes every open job order into visits/start and
        # finalize flips them all to pending_review — silently completing work
        # nobody has done. Guarded so a pre-migration schema still serves.
        _dt = "AND (ScheduledDate IS NULL OR ScheduledDate <= %s)"
        _jo_dt = _dt if _has_col(cursor, "WO_JobOrders", "ScheduledDate") else ""
        _do_dt = _dt if _has_col(cursor, "WO_DeliveryOrders", "ScheduledDate") else ""
        _today = _sgt_today()
        # init_workorders_db swallows each ALTER independently, so these two can
        # be missing while everything else is fine. Unguarded, they would fail
        # EVERY machine's sheet fleet-wide.
        _svc_cols = ("ISNULL(NeedsService,0), ServiceNote"
                     if (_has_col(cursor, "WO_DeliveryOrders", "NeedsService")
                         and _has_col(cursor, "WO_DeliveryOrders", "ServiceNote"))
                     else "CAST(0 AS BIT), NULL")

        # Machine info
        cursor.execute("""
            SELECT TOP 1 MachineName, ISNULL(IsActive, 1)
            FROM MachineLookup WHERE MachineCode = %s
        """, (code,))
        m = cursor.fetchone()
        if not m:
            conn.close()
            return jsonify({"error": "Machine not found."}), 404
        machine = {"code": code, "name": m[0], "is_active": bool(m[1])}
        # Machine attributes: current values + the authoritative option lists.
        # Wrapped because init_db swallows ALTER TABLE failures; if the columns
        # are not there yet the sheet must still open, just without this block.
        # Machine properties are Sales-owned and are not shown on this sheet.

        # Service JobOrders for me, open
        cursor.execute("""
            SELECT JobOrderID, DisplayID, ComplaintID, Notes,
                   PriorityCode, StatusCode,
                   Diagnosis, ProposedFix, EventCode, AttachedKBID,
                   OnSiteObservations, OnSiteChanges, TechnicianComments,
                   CreatedBy, CreatedAt
            FROM WO_JobOrders
            WHERE MachineCode = %s AND AssignedTo = %s AND StatusCode IN (0, 1) """ + _jo_dt + """
            ORDER BY PriorityCode DESC, CreatedAt
        """, (code, user) + ((_today,) if _jo_dt else ()))
        service_wos = []
        rows_for_complaint_lookup = []
        for r in cursor.fetchall():
            wo = {
                "id": r[0], "display_id": r[1], "complaint_id": r[2], "notes": r[3],
                "priority_code": int(r[4]), "priority_label": _label(PRIORITY, r[4]),
                "status_code":   int(r[5]), "status_label":   _label(JOBORDER_STATUS, r[5]),
                "diagnosis": r[6], "proposed_fix": r[7],
                "event_code": int(r[8]) if r[8] is not None else None,
                "attached_kb_id": int(r[9]) if r[9] is not None else None,
                "on_site_observations": r[10],
                "on_site_changes":      r[11],
                "technician_comments":  r[12],
                "created_by": r[13], "created_at": _iso(r[14]),
                "attached_kb": None,
                "complaint_summary": None,
                "complaint_reporter": None,
                "complaint_urgency": None,
            }
            # Pull KB ref summary if attached
            if wo["attached_kb_id"]:
                cursor.execute("""
                    SELECT KBID, Title, Symptom, DiagnosticConfirmation, RootCause,
                           CorrectiveAction, PreventiveAction, VerificationOfCompletion
                    FROM WO_KB_Entries WHERE KBID = %s
                """, (wo["attached_kb_id"],))
                kr = cursor.fetchone()
                if kr:
                    wo["attached_kb"] = {
                        "id": kr[0], "title": kr[1],
                        "symptom": kr[2], "diagnostic_confirmation": kr[3],
                        "root_cause": kr[4], "corrective_action": kr[5],
                        "preventive_action": kr[6], "verification_of_completion": kr[7],
                    }
            # Pull the source complaint's verbatim issue (so operator gets the
            # original report on the WO card without having to open the modal).
            if wo["complaint_id"]:
                cursor.execute("""
                    SELECT Description, ReportedBy, PerceivedUrgency
                    FROM WO_Complaints WHERE ComplaintID = %s
                """, (wo["complaint_id"],))
                cr = cursor.fetchone()
                if cr:
                    wo["complaint_summary"]  = cr[0]
                    wo["complaint_reporter"] = cr[1]
                    wo["complaint_urgency"]  = _label(PRIORITY, cr[2]) if cr[2] is not None else None
            service_wos.append(wo)

        # Open delivery WO for me (max 1)
        cursor.execute("""
            SELECT TOP 1 DeliveryOrderID, Notes, Priority, Status,
                         CreatedBy, CreatedAt, """ + _svc_cols + """
            FROM WO_DeliveryOrders
            WHERE MachineCode = %s AND AssignedTo = %s AND Status <> 'completed' """ + _do_dt + """
            ORDER BY """ + ("CASE WHEN ScheduledDate = %s THEN 0 ELSE 1 END, " if _do_dt else "") + """CreatedAt
        """, (code, user) + ((_today, _today) if _do_dt else ()))
        drow = cursor.fetchone()
        delivery_wo = None
        if drow:
            delivery_wo = {
                "needs_service": bool(drow[6]) if len(drow) > 6 else False,
                "service_note":  drow[7] if len(drow) > 7 else None,
                "id": drow[0], "notes": drow[1], "priority": drow[2],
                "status": drow[3], "created_by": drow[4], "created_at": _iso(drow[5]),
                "lines": [],
            }
            cursor.execute("""
                SELECT l.LineID, l.ItemID, l.QtyOrdered, l.QtyDelivered,
                       i.Name, i.Unit, i.Content, i.SortOrder
                FROM WO_DeliveryOrderLines l
                INNER JOIN WO_DeliveryItems i ON i.ItemID = l.ItemID
                WHERE l.DeliveryOrderID = %s
                ORDER BY i.SortOrder, i.ItemID
            """, (drow[0],))
            for lr in cursor.fetchall():
                delivery_wo["lines"].append({
                    "line_id": lr[0], "item_id": lr[1],
                    "qty_ordered": int(lr[2] or 0),
                    "qty_delivered": int(lr[3]) if lr[3] is not None else None,
                    "name": lr[4], "unit": lr[5], "content": lr[6],
                    "sort_order": int(lr[7]),
                })

        # All active delivery items (for the grid — qty defaults to 0 if no line)
        cursor.execute("""
            SELECT ItemID, Name, Unit, Content, SortOrder
            FROM WO_DeliveryItems WHERE IsActive = 1
            ORDER BY SortOrder, ItemID
        """)
        items = [{
            "id": r[0], "name": r[1], "unit": r[2], "content": r[3],
            "sort_order": int(r[4]),
        } for r in cursor.fetchall()]

        conn.close()
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

    return jsonify({
        "machine":       machine,
        "service_wos":   service_wos,
        "delivery_wo":   delivery_wo,
        "delivery_items": items,
        "service_categories": [
            {"code": code_, "label": label_} for code_, label_ in SERVICE_CATEGORIES
        ],
    })


# ── Visit Sessions: start, patch, finalize, detail, PDF ──────────────────────

def _visit_row_to_json(cursor, vrow):
    """Format a WO_VisitSessions row as JSON. Pulls junction + delivery lines."""
    vid = int(vrow[0])
    out = {
        "id": vid, "display_id": vrow[1],
        "machine_code": vrow[2], "machine_name": vrow[3],
        "operator_email": vrow[4],
        "visit_date": str(vrow[5]) if vrow[5] else None,
        "dispense_counter": int(vrow[6]) if vrow[6] is not None else None,
        "services": {
            "PMC": {"done": bool(vrow[7]),  "remarks": vrow[8]},
            "CMR": {"done": bool(vrow[9]),  "remarks": vrow[10]},
            "INR": {"done": bool(vrow[11]), "remarks": vrow[12]},
            "OTH": {"done": bool(vrow[13]), "remarks": vrow[14]},
        },
        "receiving_name": vrow[15], "receiving_date": str(vrow[16]) if vrow[16] else None,
        "service_name":   vrow[17], "service_date":   str(vrow[18]) if vrow[18] else None,
        "customer_unavailable":        bool(vrow[19]),
        "customer_unavailable_reason": vrow[20],
        "status":      vrow[21],
        "pdf_url":     vrow[22],
        "linked_delivery_order_id": int(vrow[23]) if vrow[23] is not None else None,
        "created_at":   _iso(vrow[24]),
        "updated_at":   _iso(vrow[25]),
        "submitted_at": _iso(vrow[26]),
        "signed_at":    _iso(vrow[27]),
        "linked_job_order_ids": [],
    }
    cursor.execute(
        "SELECT JobOrderID FROM WO_VisitSession_JobOrders WHERE VisitID = %s",
        (vid,),
    )
    out["linked_job_order_ids"] = [int(r[0]) for r in cursor.fetchall()]
    return out


def _select_visit_sql():
    return """
        SELECT VisitID, DisplayID, MachineCode, MachineNameSnap,
               OperatorEmail, VisitDate, DispenseCounter,
               Svc_PMC_Done, Svc_PMC_Remarks,
               Svc_CMR_Done, Svc_CMR_Remarks,
               Svc_INR_Done, Svc_INR_Remarks,
               Svc_OTH_Done, Svc_OTH_Remarks,
               ReceivingName, ReceivingDate,
               ServiceName,   ServiceDate,
               CustomerUnavailable, CustomerUnavailableReason,
               Status, PDFSPWebURL,
               LinkedDeliveryOrderID,
               CreatedAt, UpdatedAt, SubmittedAt, SignedAt
        FROM WO_VisitSessions
    """


@workorders_bp.route("/visits/start", methods=["POST"])
@require_roles(*OPERATOR_ROLES)
def api_visit_start():
    """Create a draft visit for current operator + machine_code.

    Job order ids are re-validated server-side, INCLUDING the scheduled date —
    a stale tab or a crafted payload must not be able to pull a future-dated job
    into today's visit, where finalize would flip it to pending_review.
    Body: {machine_code, service_job_order_ids: [int...], delivery_order_id: int|null}
    Returns: full visit detail."""
    data = request.get_json(silent=True) or {}
    code = (data.get("machine_code") or "").strip()
    if not code:
        return jsonify({"error": "machine_code required."}), 400
    sjo_ids = data.get("service_job_order_ids") or []
    do_id   = data.get("delivery_order_id")
    user = get_current_user()
    # Visit date is the operator's LOCAL (SGT, UTC+8) day — a 07:00 SGT visit
    # must not be stamped as yesterday, and reopening a sheet at 08:30 SGT
    # must reuse the same visit rather than start a new one.
    today = (datetime.utcnow() + timedelta(hours=8)).date()
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT MachineName FROM MachineLookup WHERE MachineCode=%s", (code,))
        mrow = cursor.fetchone()
        machine_name = mrow[0] if mrow else None

        # Reuse an existing UNFINALISED visit for this operator+machine+day instead of
        # creating a duplicate draft every time the sheet is opened. (Finalised
        # visits are immutable, so a second trip on the same day starts a new one.)
        cursor.execute("""
            SELECT TOP 1 VisitID FROM WO_VisitSessions
            WHERE MachineCode = %s AND OperatorEmail = %s AND VisitDate = %s
              AND (Status IS NULL OR Status NOT IN ('signed', 'pending_email_signature'))
            ORDER BY VisitID DESC
        """, (code, user, today))
        exist = cursor.fetchone()

        if exist:
            vid = int(exist[0])
            if do_id is not None:
                cursor.execute(
                    "UPDATE WO_VisitSessions SET LinkedDeliveryOrderID = %s "
                    "WHERE VisitID = %s AND LinkedDeliveryOrderID IS NULL",
                    (do_id, vid),
                )
        else:
            display_id = allocate_display_id(cursor, "VIS")
            cursor.execute("""
                INSERT INTO WO_VisitSessions
                    (DisplayID, MachineCode, MachineNameSnap, OperatorEmail,
                     VisitDate, LinkedDeliveryOrderID)
                OUTPUT INSERTED.VisitID
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (display_id, code, machine_name, user, today, do_id))
            vid = int(cursor.fetchone()[0])

        for jid in sjo_ids:
            try:
                jid_int = int(jid)
            except (TypeError, ValueError):
                continue
            # Only link job orders that are genuinely open on THIS machine and
            # (for non-managers) assigned to the caller — stops a crafted payload
            # force-closing someone else's work order at finalize.
            if get_role(user) in MANAGER_ROLES:
                cursor.execute(
                    "SELECT 1 FROM WO_JobOrders WHERE JobOrderID=%s "
                    "AND MachineCode=%s AND StatusCode IN (0,1)",
                    (jid_int, code),
                )
            else:
                cursor.execute(
                    "SELECT 1 FROM WO_JobOrders WHERE JobOrderID=%s "
                    "AND MachineCode=%s AND StatusCode IN (0,1) AND LOWER(AssignedTo)=%s",
                    (jid_int, code, user.lower()),
                )
            if not cursor.fetchone():
                continue
            cursor.execute("""
                IF NOT EXISTS (SELECT 1 FROM WO_VisitSession_JobOrders
                               WHERE VisitID=%s AND JobOrderID=%s)
                INSERT INTO WO_VisitSession_JobOrders (VisitID, JobOrderID) VALUES (%s, %s)
            """, (vid, jid_int, vid, jid_int))

        cursor.execute(_select_visit_sql() + " WHERE VisitID = %s", (vid,))
        row = cursor.fetchone()
        out = _visit_row_to_json(cursor, row)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "visit": out})
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/visits/<int:vid>", methods=["GET"])
@require_roles(*OPERATOR_ROLES)
def api_visit_detail(vid):
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(_select_visit_sql() + " WHERE VisitID = %s", (vid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Visit not found."}), 404
        if get_role(user) not in MANAGER_ROLES and (row[4] or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not your visit."}), 403
        out = _visit_row_to_json(cursor, row)
        conn.close()
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/visits/<int:vid>", methods=["PATCH"])
@require_roles(*OPERATOR_ROLES)
def api_visit_update(vid):
    """Save draft. Body any of:
        dispense_counter:int, services:{PMC|CMR|INR|OTH: {done, remarks}},
        delivery_lines:[{line_id?:int, item_id:int, qty_delivered:int}],
        service_name:str, service_date:'YYYY-MM-DD',
        customer_unavailable:bool, customer_unavailable_reason:str
    """
    data = request.get_json(silent=True) or {}
    user = get_current_user()
    # Validated BEFORE opening a connection: a bad attribute value can then never
    # cost the technician the draft he just typed on site.
    attrs_clean, attrs_err = _validate_machine_attrs(data.get("machine_attrs"))
    if attrs_err:
        return jsonify({"error": attrs_err}), 400
    # Same reasoning, same place: ServiceDate went into the UPDATE verbatim and
    # the date box on the sheet is a plain text input now. Checked BEFORE
    # get_connection() -- an early return after it would leak the connection,
    # which every other exit in this function avoids with an explicit close.
    if data.get("service_date") and _parse_date(data["service_date"]) is None:
        return jsonify({"error": "Service date must be YYYY-MM-DD."}), 400
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT OperatorEmail, Status, LinkedDeliveryOrderID, MachineCode FROM WO_VisitSessions WHERE VisitID=%s",
            (vid,),
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Visit not found."}), 404
        if get_role(user) not in MANAGER_ROLES and (row[0] or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not your visit."}), 403
        if row[1] in ("signed", "pending_email_signature"):
            conn.close()
            return jsonify({"error": "Visit already finalised; cannot edit."}), 400
        linked_do = row[2]

        sets, params = [], []
        if "dispense_counter" in data:
            try:
                sets.append("DispenseCounter = %s")
                params.append(int(data["dispense_counter"]) if data["dispense_counter"] not in (None, "") else None)
            except (TypeError, ValueError):
                return jsonify({"error": "dispense_counter must be integer."}), 400
        svc = data.get("services") or {}
        for code_ in ("PMC", "CMR", "INR", "OTH"):
            if code_ in svc:
                sets.append(f"Svc_{code_}_Done = %s")
                params.append(1 if svc[code_].get("done") else 0)
                sets.append(f"Svc_{code_}_Remarks = %s")
                params.append((svc[code_].get("remarks") or "").strip() or None)
        if "service_name" in data:
            sets.append("ServiceName = %s")
            params.append((data["service_name"] or "").strip() or None)
        if "service_date" in data:
            sets.append("ServiceDate = %s")
            params.append(data["service_date"] or None)
        if "customer_unavailable" in data:
            sets.append("CustomerUnavailable = %s")
            params.append(1 if data["customer_unavailable"] else 0)
        if "customer_unavailable_reason" in data:
            sets.append("CustomerUnavailableReason = %s")
            params.append((data["customer_unavailable_reason"] or "").strip() or None)

        # Machine attributes ride along with the draft and commit with it.
        # A draft save must NEVER fail: this fires on every keystroke-driven
        # autosave, and losing the technician's counter and remarks because of a
        # registry problem he cannot fix would be far worse than the problem.
        # So a bad machine code is REPORTED in the 200 body and hard-failed later
        # at finalize, where it is actionable and nothing is yet at stake.
        attrs_warning = None
        if attrs_clean:
            if not row[3]:
                attrs_warning = "This visit has no machine code, so the machine details were not saved."
            else:
                _n = None   # machine attributes are Sales-owned; see MACHINE_ATTRS
                if _n is not None and _n <= 0:
                    attrs_warning = (f"Machine {row[3]} may not be in the machine registry — "
                                     "could not be confirmed. Tell a manager.")

        if sets:
            sets.append("UpdatedAt = SYSUTCDATETIME()")
            params.append(vid)
            cursor.execute(
                f"UPDATE WO_VisitSessions SET {', '.join(sets)} WHERE VisitID = %s",
                tuple(params),
            )

        # Upsert delivery lines (qty_delivered) — operator records actual qty.
        # If no linked delivery WO exists and the operator entered any qty,
        # lazy-create an ad-hoc delivery WO so the replenishment is recorded
        # in the normal delivery history.
        lines = data.get("delivery_lines") or []
        non_zero_lines = [
            ln for ln in lines
            if (ln.get("qty_delivered") not in (None, "", 0, "0"))
        ]
        if non_zero_lines and linked_do is None:
            # Look up machine name for the WO record
            cursor.execute(
                "SELECT MachineCode, MachineNameSnap FROM WO_VisitSessions WHERE VisitID=%s",
                (vid,),
            )
            vrow = cursor.fetchone()
            mcode = vrow[0]
            mname = vrow[1] or mcode
            cursor.execute("""
                INSERT INTO WO_DeliveryOrders
                    (MachineName, MachineCode, Notes,
                     AssignedTo, Priority, Status, CreatedBy)
                OUTPUT INSERTED.DeliveryOrderID
                VALUES (%s, %s, %s,
                        %s, 'normal', 'open', %s)
            """, (mname, mcode,
                  "Ad-hoc replenishment recorded by operator at site visit.",
                  user, user))
            linked_do = int(cursor.fetchone()[0])
            cursor.execute(
                "UPDATE WO_VisitSessions SET LinkedDeliveryOrderID = %s WHERE VisitID = %s",
                (linked_do, vid),
            )
            _log_activity(cursor, "deliveryorder", linked_do, "adhoc_created",
                          f"Ad-hoc delivery from visit {vid}", user)

        if lines and linked_do is not None:
            for ln in lines:
                try:
                    item_id = int(ln.get("item_id"))
                    qty = int(ln.get("qty_delivered") or 0)
                except (TypeError, ValueError):
                    continue
                if qty < 0 or qty > 999:
                    conn.close()
                    return jsonify({"error": f"Quantity for item {item_id} must be 0-999."}), 400
                cursor.execute("""
                    UPDATE WO_DeliveryOrderLines SET QtyDelivered = %s
                    WHERE DeliveryOrderID = %s AND ItemID = %s
                """, (qty, linked_do, item_id))
                cursor.execute("""
                    IF NOT EXISTS (
                        SELECT 1 FROM WO_DeliveryOrderLines
                        WHERE DeliveryOrderID = %s AND ItemID = %s
                    )
                    INSERT INTO WO_DeliveryOrderLines
                        (DeliveryOrderID, ItemID, QtyOrdered, QtyDelivered)
                    VALUES (%s, %s, 0, %s)
                """, (linked_do, item_id, linked_do, item_id, qty))

        conn.commit()
        conn.close()
        out = {"ok": True}
        if attrs_warning:
            out["machine_attrs_warning"] = attrs_warning
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500


# ── PDF generation (ReportLab — mirrors the KNM Work Order paper template) ────

def _pdf_esc(v):
    """reportlab Paragraphs are mini-HTML: a bare & or < in free text raises."""
    return (str(v or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _build_visit_pdf(visit, service_wos, delivery_wo, delivery_items_with_qty):
    """Return PDF bytes for a finalised visit, modelled on KNM's 'Work Order 11May26.pdf'.
    delivery_items_with_qty: [{name, unit, content, qty_delivered}]
    """
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, KeepTogether)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=18*mm, rightMargin=18*mm,
                            topMargin=15*mm, bottomMargin=15*mm,
                            title=f"Work Order {visit.get('display_id') or visit['id']}")
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=ss["Heading1"], fontSize=18, alignment=2, textColor=colors.black, spaceAfter=4)
    cn = ParagraphStyle("cn", parent=ss["Normal"], fontSize=10, leading=12)
    sm = ParagraphStyle("sm", parent=ss["Normal"], fontSize=8,  leading=10, textColor=colors.grey)
    sec = ParagraphStyle("sec", parent=ss["Heading2"], fontSize=12, spaceBefore=8, spaceAfter=4, textColor=colors.black)
    body = ParagraphStyle("body", parent=ss["Normal"], fontSize=9, leading=11)

    story = []

    # Header
    header_data = [
        [Paragraph("<b>Kopi Near Me Pte. Ltd.</b>", cn), Paragraph("WORK ORDER", h1)],
        [Paragraph("Kopi Near Me Pte. Ltd.<br/>18 Kim Chuan Terrace, #03-03, Singapore 537040", sm),
         Paragraph("sales@kopinearme.com", sm)],
    ]
    story.append(Table(header_data, colWidths=[110*mm, 60*mm],
                       style=TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")])))
    story.append(Spacer(1, 4*mm))

    # Visit meta — Company / Date / Counter / DisplayID
    machine_name = visit.get("machine_name") or visit.get("machine_code")
    visit_date   = visit.get("visit_date")  or ""
    counter      = visit.get("dispense_counter")
    display_id   = visit.get("display_id") or f"VIS-{visit['id']}"

    meta = [
        [Paragraph("<b>Company Name</b>", body), Paragraph(machine_name, body),
         Paragraph("<b>Date:</b>", body), Paragraph(str(visit_date), body)],
        [Paragraph("<b>Dispense Counter:</b>", body),
         Paragraph(str(counter) if counter is not None else "—", body),
         Paragraph("<b>WO ID:</b>", body), Paragraph(display_id, body)],
    ]
    story.append(Table(meta, colWidths=[36*mm, 60*mm, 24*mm, 50*mm],
                       style=TableStyle([
                           ("LINEBELOW", (0,0), (-1,-1), 0.4, colors.grey),
                           ("FONTSIZE", (0,0), (-1,-1), 9),
                           ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                           ("TOPPADDING",    (0,0), (-1,-1), 3),
                       ])))
    story.append(Spacer(1, 4*mm))

    # Service requested highlight (if any open service WO)
    if service_wos:
        for wo in service_wos:
            bits = []
            if wo.get("display_id"):  bits.append(f"<b>{wo['display_id']}</b>")
            if wo.get("diagnosis"):   bits.append(f"<b>Diagnosis:</b> {_pdf_esc(wo['diagnosis'])}")
            if wo.get("proposed_fix"):bits.append(f"<b>Proposed fix:</b> {_pdf_esc(wo['proposed_fix'])}")
            if wo.get("attached_kb"): bits.append(f"<b>Ref KB:</b> {wo['attached_kb'].get('title')}")
            txt = "<br/>".join(bits)
            box = Table([[Paragraph("⚠ <b>Service requested to fix issue</b>", body),
                          Paragraph(txt, body)]],
                        colWidths=[55*mm, 115*mm],
                        style=TableStyle([
                            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#fef3c7")),
                            ("BOX",        (0,0), (-1,-1), 0.7, colors.HexColor("#b45309")),
                            ("VALIGN",     (0,0), (-1,-1), "TOP"),
                            ("LEFTPADDING",  (0,0), (-1,-1), 6),
                            ("RIGHTPADDING", (0,0), (-1,-1), 6),
                            ("TOPPADDING",   (0,0), (-1,-1), 5),
                            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
                        ]))
            story.append(box)
            story.append(Spacer(1, 2*mm))
        story.append(Spacer(1, 2*mm))

    # SERVICES section
    story.append(Paragraph("SERVICES", sec))
    sv = visit.get("services", {})
    svc_rows = [["Description", "Completed", "Remarks"]]
    for code_, label_ in SERVICE_CATEGORIES:
        v = sv.get(code_) or {}
        svc_rows.append([
            Paragraph(label_, body),
            "✓" if v.get("done") else "☐",
            Paragraph(v.get("remarks") or "", body),
        ])
    svc_t = Table(svc_rows, colWidths=[80*mm, 22*mm, 68*mm])
    svc_t.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#f9fafb")),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("LINEBELOW",   (0,0), (-1,-1), 0.3, colors.lightgrey),
        ("ALIGN",       (1,1), (1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    story.append(svc_t)
    story.append(Spacer(1, 4*mm))

    # DELIVERY section
    story.append(Paragraph("DELIVERY", sec))
    dl_rows = [["Description", "Unit", "Content", "Qty"]]
    for it in delivery_items_with_qty:
        qty = it.get("qty_delivered")
        dl_rows.append([
            Paragraph(it["name"], body),
            it["unit"], it["content"],
            str(qty) if qty is not None else "—",
        ])
    dl_t = Table(dl_rows, colWidths=[80*mm, 22*mm, 30*mm, 38*mm])
    dl_t.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#f9fafb")),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 9),
        ("LINEBELOW",    (0,0), (-1,-1), 0.3, colors.lightgrey),
        ("ALIGN",        (1,1), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ]))
    story.append(dl_t)
    story.append(Spacer(1, 6*mm))

    # Acknowledgement + signatures
    story.append(Paragraph(
        "The Receiving Personnel has checked and verified that services rendered and goods delivered are in good order.",
        body
    ))
    story.append(Spacer(1, 4*mm))

    if visit.get("customer_unavailable"):
        unavail_box = Table([[Paragraph(
            f"⚠ <b>Customer unavailable on site.</b><br/>Reason: {visit.get('customer_unavailable_reason') or '—'}<br/>"
            "PDF to be emailed to contact person for sign-off.", body
        )]], colWidths=[170*mm],
        style=TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#fee2e2")),
            ("BOX", (0,0), (-1,-1), 0.7, colors.HexColor("#b91c1c")),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
            ("RIGHTPADDING",  (0,0), (-1,-1), 8),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(unavail_box)
        story.append(Spacer(1, 4*mm))

    sigs = [
        [Paragraph("<b>Receiving Personnel</b>", body),
         Paragraph("<b>Service Personnel</b>", body)],
        [Paragraph(f"Signature: <i>{visit.get('receiving_name') or '— pending —'}</i>", body),
         Paragraph(f"Signature: <i>{visit.get('service_name') or '— pending —'}</i>", body)],
        [Paragraph(f"Name / Date: {visit.get('receiving_name') or '—'} / {visit.get('receiving_date') or '—'}", body),
         Paragraph(f"Name / Date: {visit.get('service_name') or '—'} / {visit.get('service_date') or '—'}", body)],
    ]
    story.append(Table(sigs, colWidths=[85*mm, 85*mm],
                       style=TableStyle([
                           ("LINEABOVE", (0,1), (-1,1), 0.4, colors.grey),
                           ("VALIGN", (0,0), (-1,-1), "TOP"),
                           ("TOPPADDING", (0,0), (-1,-1), 6),
                           ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                       ])))

    doc.build(story)
    return buf.getvalue()


@workorders_bp.route("/visits/<int:vid>/finalize", methods=["POST"])
@require_roles(*OPERATOR_ROLES)
def api_visit_finalize(vid):
    """Finalise the visit: validate, save signature(s), generate PDF, upload to SP,
    transition linked WOs.

    Body:
        receiving_name: str|null (if customer present)
        receiving_date: 'YYYY-MM-DD'
        service_name:   str (operator typed)
        service_date:   'YYYY-MM-DD'
        customer_unavailable: bool
        customer_unavailable_reason: str
    """
    data = request.get_json(silent=True) or {}
    user = get_current_user()
    conn = None
    skipped_job_orders = []
    attrs_clean, attrs_err = _validate_machine_attrs(data.get("machine_attrs"))
    if attrs_err:
        return jsonify({"error": attrs_err}), 400
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(_select_visit_sql() + " WHERE VisitID = %s", (vid,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Visit not found."}), 404
        if get_role(user) not in MANAGER_ROLES and (row[4] or "").lower() != user.lower():
            conn.close()
            return jsonify({"error": "Not your visit."}), 403
        if row[21] in ("signed", "pending_email_signature"):
            conn.close()
            return jsonify({"error": "Already finalised."}), 400

        cust_unavail   = bool(data.get("customer_unavailable"))
        cust_reason    = (data.get("customer_unavailable_reason") or "").strip() or None
        receiving_name = (data.get("receiving_name") or "").strip() or None
        receiving_date = data.get("receiving_date") or None
        service_name   = (data.get("service_name") or "").strip() or None
        service_date   = data.get("service_date") or None

        # Both are written verbatim into the visit row further down. Every other
        # date this blueprint accepts is format-checked; these two were the gap,
        # and the front end can now hand over whatever was typed into a text box.
        for _lbl, _v in (("Receiving date", receiving_date), ("Service date", service_date)):
            if _v and _parse_date(_v) is None:
                conn.close()
                return jsonify({"error": f"{_lbl} must be YYYY-MM-DD."}), 400

        if not service_name:
            conn.close()
            return jsonify({"error": "Service Personnel name (you) is required."}), 400

        # Machine attributes are Sales' job now (see MACHINE_ATTRS). The old
        # gate refused to close a signed Work Order until the technician filled
        # five dropdowns. Nothing here may ever block a finalize again.
        #
        # The write is kept only so an older cached copy of the sheet that still
        # posts machine_attrs is not silently discarded. A failure is logged and
        # ignored — this is a courtesy, not a requirement of the visit.
        _vmcode = row[2]
        if attrs_clean and _vmcode:
            try:
                _write_machine_attrs(cursor, _vmcode, attrs_clean, user)
            except Exception as e:
                print(f"[machine_attrs] legacy write skipped for {_vmcode}: {e}")

        if cust_unavail and not cust_reason:
            conn.close()
            return jsonify({"error": "Reason required when customer unavailable."}), 400
        if not cust_unavail and not receiving_name:
            conn.close()
            return jsonify({"error": "Receiving Personnel name required (or mark customer unavailable)."}), 400

        new_status = "pending_email_signature" if cust_unavail else "signed"

        cursor.execute("""
            UPDATE WO_VisitSessions
            SET ReceivingName = %s, ReceivingDate = %s,
                ServiceName   = %s, ServiceDate   = %s,
                CustomerUnavailable = %s, CustomerUnavailableReason = %s,
                Status = %s, SubmittedAt = SYSUTCDATETIME(),
                SignedAt = CASE WHEN %s = 'signed' THEN SYSUTCDATETIME() ELSE NULL END
            WHERE VisitID = %s
        """, (receiving_name, receiving_date,
              service_name, service_date,
              1 if cust_unavail else 0, cust_reason,
              new_status, new_status, vid))

        # Refetch the visit + supporting data needed for PDF
        cursor.execute(_select_visit_sql() + " WHERE VisitID = %s", (vid,))
        vrow = cursor.fetchone()
        visit = _visit_row_to_json(cursor, vrow)

        # Linked service WOs (snapshot)
        service_wos = []
        if visit["linked_job_order_ids"]:
            ph = ",".join(["%s"] * len(visit["linked_job_order_ids"]))
            cursor.execute(f"""
                SELECT JobOrderID, DisplayID, Diagnosis, ProposedFix, AttachedKBID
                FROM WO_JobOrders WHERE JobOrderID IN ({ph})
            """, tuple(visit["linked_job_order_ids"]))
            for r in cursor.fetchall():
                wo = {"id": r[0], "display_id": r[1], "diagnosis": r[2],
                      "proposed_fix": r[3], "attached_kb": None}
                if r[4] is not None:
                    cursor.execute(
                        "SELECT Title FROM WO_KB_Entries WHERE KBID = %s",
                        (r[4],),
                    )
                    kt = cursor.fetchone()
                    if kt:
                        wo["attached_kb"] = {"title": kt[0]}
                service_wos.append(wo)

        # A merged stop carries its service request on the delivery order rather
        # than a job order. Feed it through the same shape so the signed PDF
        # still shows the amber "Service requested" box.
        if visit["linked_delivery_order_id"]:
            try:
                cursor.execute("""
                    SELECT ISNULL(NeedsService,0), ServiceNote
                    FROM WO_DeliveryOrders WHERE DeliveryOrderID = %s
                """, (visit["linked_delivery_order_id"],))
                _dr = cursor.fetchone()
                if _dr and _dr[0]:
                    service_wos.append({
                        "id": None, "display_id": None,
                        "diagnosis": _dr[1] or "Service requested by dispatch",
                        "proposed_fix": None, "attached_kb": None,
                    })
            except Exception as e:
                print(f"[visit finalize] merged service note skipped: {e}")

        # Delivery items (all active + the qty_delivered from lines if linked)
        delivery_items_with_qty = []
        cursor.execute("""
            SELECT ItemID, Name, Unit, Content, SortOrder
            FROM WO_DeliveryItems WHERE IsActive = 1
            ORDER BY SortOrder, ItemID
        """)
        items_meta = [(r[0], r[1], r[2], r[3]) for r in cursor.fetchall()]
        qty_map = {}
        if visit["linked_delivery_order_id"]:
            cursor.execute("""
                SELECT ItemID, QtyDelivered
                FROM WO_DeliveryOrderLines WHERE DeliveryOrderID = %s
            """, (visit["linked_delivery_order_id"],))
            qty_map = {int(r[0]): (int(r[1]) if r[1] is not None else None)
                       for r in cursor.fetchall()}
        for iid, name, unit, content in items_meta:
            delivery_items_with_qty.append({
                "name": name, "unit": unit, "content": content,
                "qty_delivered": qty_map.get(int(iid)),
            })

        # Transition linked WOs:
        #   Each linked JobOrder → StatusCode = 2 (pending_review) so manager reviews.
        #   Re-check it is still open (and still the caller's, for non-managers) —
        #   it may have been reassigned or closed since the visit was started.
        may_force = get_role(user) in MANAGER_ROLES
        for jid in visit["linked_job_order_ids"]:
            if may_force:
                cursor.execute("""
                    UPDATE WO_JobOrders
                    SET StatusCode = 2, CompletedBy = %s, CompletedAt = SYSUTCDATETIME()
                    WHERE JobOrderID = %s AND StatusCode IN (0, 1)
                """, (user, jid))
            else:
                cursor.execute("""
                    UPDATE WO_JobOrders
                    SET StatusCode = 2, CompletedBy = %s, CompletedAt = SYSUTCDATETIME()
                    WHERE JobOrderID = %s AND StatusCode IN (0, 1)
                      AND LOWER(AssignedTo) = %s
                """, (user, jid, user.lower()))
            if cursor.rowcount:
                _log_activity(cursor, "joborder", jid, "visit_submit",
                              f"Linked to visit {visit['display_id']}", user)
            else:
                skipped_job_orders.append(jid)
                _log_activity(cursor, "joborder", jid, "visit_submit_skipped",
                              f"Visit {visit['display_id']} submitted but job order was "
                              f"no longer open/assigned to {user}", user)

        # Linked DeliveryOrder → completed. The goods physically changed hands
        # even when the customer was unavailable to sign, so complete it either
        # way; the visit's own 'pending_email_signature' status is what flags the
        # manager for sign-off. (Previously this left the DO open forever with no
        # path to close it.) Guarded on Status <> 'completed' so a second submit
        # or a manager also hitting /deliveryorders/<id>/complete cannot
        # double-run _sync_topup and destroy the top-up baseline.
        if visit["linked_delivery_order_id"]:
            cursor.execute("""
                UPDATE WO_DeliveryOrders
                SET Status = 'completed', CompletedBy = %s, CompletedAt = SYSUTCDATETIME(),
                    RecipientName = COALESCE(%s, RecipientName)
                WHERE DeliveryOrderID = %s AND Status <> 'completed'
            """, (user, receiving_name, visit["linked_delivery_order_id"]))
            if cursor.rowcount:
                _log_activity(cursor, "deliveryorder", visit["linked_delivery_order_id"],
                              "visit_submit",
                              f"Linked to visit {visit['display_id']}"
                              + (" (customer unavailable)" if cust_unavail else ""), user)
                # Mirror /deliveryorders/<id>/complete: keep MachineLookup's refill
                # clock in step, otherwise "last refill" / "vends since" go stale.
                # Only when something was actually delivered. Every stop is a
                # delivery-order row now, so a service-only visit would
                # otherwise stamp LastTopupTimestamp and overwrite
                # PreviousTopupTimestamp — unrecoverable, and it drives
                # "vends since last top-up" across the whole fleet.
                _delivered = 0
                try:
                    cursor.execute("""
                        SELECT ISNULL(SUM(ISNULL(QtyDelivered,0)),0)
                        FROM WO_DeliveryOrderLines WHERE DeliveryOrderID = %s
                    """, (visit["linked_delivery_order_id"],))
                    _dr = cursor.fetchone()
                    _delivered = int(_dr[0] or 0) if _dr else 0
                except Exception:
                    _delivered = 0
                if not _delivered:
                    try:
                        cursor.execute("""
                            SELECT ISNULL(Item1Qty,0)+ISNULL(Item2Qty,0)+ISNULL(Item3Qty,0)
                                 + ISNULL(Item4Qty,0)+ISNULL(Item5Qty,0)+ISNULL(Item6Qty,0)
                                 + ISNULL(Item7Qty,0)+ISNULL(Item8Qty,0)
                            FROM WO_DeliveryOrders WHERE DeliveryOrderID = %s
                        """, (visit["linked_delivery_order_id"],))
                        _lr = cursor.fetchone()
                        _delivered = int(_lr[0] or 0) if _lr else 0
                    except Exception:
                        _delivered = 0
                if visit["machine_code"] and _delivered:
                    topup_note = _sync_topup(cursor, str(visit["machine_code"]))
                    _log_activity(cursor, "deliveryorder", visit["linked_delivery_order_id"],
                                  "topup_synced", topup_note, user)
                elif visit["machine_code"]:
                    _log_activity(cursor, "deliveryorder", visit["linked_delivery_order_id"],
                                  "topup_skipped",
                                  "Nothing delivered — refill clock left untouched", user)

        # Commit the record BEFORE touching SharePoint. The upload is a 60-120s
        # external HTTP call; doing it inside the transaction held write locks on
        # MachineLookup / WO_JobOrders / WO_DeliveryOrders for its whole duration.
        conn.commit()
        conn.close()
        conn = None

        # Build PDF + upload to SharePoint (best-effort, own short connection)
        pdf_ready = False
        conn2 = None
        try:
            pdf_bytes = _build_visit_pdf(visit, service_wos, None, delivery_items_with_qty)
            now = datetime.utcnow()
            sp_item_id, web_url, _path = sp.upload_bytes(
                kind="workorder",
                display_id=visit["display_id"] or f"VIS-{vid}",
                year=now.year, month=now.month,
                file_name=f"{visit['display_id'] or 'VIS-' + str(vid)}.pdf",
                data=pdf_bytes,
                content_type="application/pdf",
            )
            conn2 = get_connection()
            c2 = conn2.cursor()
            c2.execute(
                "UPDATE WO_VisitSessions SET PDFSPItemID = %s, PDFSPWebURL = %s WHERE VisitID = %s",
                (sp_item_id, web_url, vid),
            )
            conn2.commit()
            pdf_ready = True
        except Exception as e:
            # PDF/SP failure must not lose the signature; log and continue
            print(f"[visit_finalize] PDF/SP upload failed for visit {vid}: {e}")
        finally:
            if conn2 is not None:
                try:
                    conn2.close()
                except Exception:
                    pass

        return jsonify({"ok": True, "status": new_status,
                        "pdf_ready": pdf_ready,
                        "skipped_job_orders": skipped_job_orders})
    except Exception as e:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass
        return jsonify({"error": f"Database error: {str(e)}"}), 500


@workorders_bp.route("/visits/<int:vid>/pdf")
@require_roles(*OPERATOR_ROLES)
def api_visit_pdf(vid):
    """Proxy the SP-hosted PDF so Easy Auth gates the read."""
    user = get_current_user()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT OperatorEmail, PDFSPItemID FROM WO_VisitSessions WHERE VisitID = %s
        """, (vid,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Visit not found."}), 404
        if get_role(user) not in MANAGER_ROLES and (row[0] or "").lower() != user.lower():
            return jsonify({"error": "Not your visit."}), 403
        if not row[1]:
            return jsonify({"error": "PDF not yet generated."}), 404
        body, ctype = sp.download_bytes(row[1])
        return Response(body, mimetype=ctype or "application/pdf")
    except Exception as e:
        return jsonify({"error": f"PDF read failed: {str(e)}"}), 500


@workorders_bp.route("/visits")
@require_roles(*OPERATOR_ROLES)
def api_visit_list():
    """List visits — operator sees their own; manager sees all."""
    user = get_current_user()
    scope = (request.args.get("scope") or "mine").strip().lower()
    if get_role(user) not in MANAGER_ROLES:
        scope = "mine"
    where, params = [], []
    if scope == "mine":
        where.append("OperatorEmail = %s")
        params.append(user)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(_select_visit_sql() + f" {where_sql} ORDER BY VisitDate DESC, VisitID DESC")
        rows = cursor.fetchall()
        out = [_visit_row_to_json(cursor, r) for r in rows]
        conn.close()
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500
# ── Google Calendar feed (read-only) ─────────────────────────────────────────
# Sales keys the round in Google Calendar. gcal_feed polls an Apps Script web
# app on a BACKGROUND thread; this route only ever reads that in-memory cache,
# so it can never block the single gunicorn worker on an outbound call.

@workorders_bp.route("/schedule/gcal", methods=["GET"])
@require_roles(*SALES_ROLES)
def api_schedule_gcal():
    try:
        import gcal_feed
    except Exception as e:
        return jsonify({"enabled": False, "ok": False, "stops": [],
                        "count": 0, "error": f"feed module unavailable: {e}"})
    return jsonify(gcal_feed.snapshot(
        _parse_date(request.args.get("from")),
        _parse_date(request.args.get("to")),
    ))
